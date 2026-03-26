from flask import Flask, render_template, request, redirect, flash, session,send_file,url_for,jsonify
import mysql.connector
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime,time, timedelta
import pandas as pd
import io
import re
import random
import string
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import yagmail
import os
from werkzeug.utils import secure_filename
from flask import send_file, make_response
import io
import xlsxwriter
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
import json
from edu_ai.keyword_engine.config_builder import build_keyword_config
from edu_ai.evaluation_engine import evaluate_hybrid_answer
import hashlib
from dotenv import load_dotenv
from pdf_qa_engine import extract_answers_from_pdf

# Load environment variables from .env file
load_dotenv()


app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Replace with a secure key in production


ALLOWED_EXTENSIONS = {'xls', 'xlsx'}


# MySQL Configuration
db_config = {
    'host': '127.0.0.1',  # Better than 'localhost' on some Windows setups
    'user': 'root',
    'port': 3307,        # User confirmed 3307 (as integer)
    'password': '',      # Replace with your MySQL password
    'database': 'edusystem_db',
    'use_pure': True     # Force pure Python implementation to avoid binary crashes on Windows
}

# ------------------------- Test DB Route ---------------------------
@app.route('/test_db')
def test_db():
    try:
        conn = mysql.connector.connect(**db_config)
        cur = conn.cursor()
        cur.execute("SELECT DATABASE()")
        db_name = cur.fetchone()
        cur.close()
        conn.close()
        return f"Database Connection Successful! Connected to: {db_name[0]}"
    except Exception as e:
        return f"Database Connection Failed: {str(e)}"

# ------------------------- Global Config ---------------------------
import sys

# ---------------- EMAIL SENDING FUNCTION ----------------
def send_email(receiver_email, subject, body):
    try:
        # Login with your Gmail (use app password, not raw Gmail password)
        yag = yagmail.SMTP("www.gautam.2005@gmail.com", "yilt qrum ypgz tlgt")
        yag.send(
            to=receiver_email,
            subject=subject,
            contents=body
        )
        print(f"Email sent to {receiver_email}")
        return True
    except Exception as e:
        print(f"Error sending email to {receiver_email}: {e}")
        return False

# ---------------- PASSWORD GENERATOR ----------------
def generate_password(length=12):
    chars = string.ascii_letters + string.digits + "!@#$%^&*()?"
    return ''.join(random.choice(chars) for _ in range(length))


@app.route('/')
def index():
    return render_template('landing_page.html')


@app.route('/change_password')
def request_password_change():
    uid = session['user_id']   # this is users.uid

    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()

    cursor.execute("SELECT email FROM users WHERE uid = %s", (uid,))
    row = cursor.fetchone()
    if not row:
        flash("User not found", "error")
        return redirect('/login')

    email = row[0]

    otp = str(random.randint(100000, 999999))
    otp_hash = hashlib.sha256(otp.encode()).hexdigest()
    expires_at = datetime.now() + timedelta(minutes=5)

    cursor.execute("""
        INSERT INTO password_reset_otp (user_id, otp_hash, expires_at)
        VALUES (%s, %s, %s)
        ON DUPLICATE KEY UPDATE
            otp_hash = %s,
            expires_at = %s,
            verified = 0
    """, (uid, otp_hash, expires_at, otp_hash, expires_at))

    conn.commit()
    cursor.close()
    conn.close()

    send_email(
        email,
        "EduAI Password Reset OTP",
        f"Your OTP is: {otp}\nThis OTP is valid for 5 minutes."
    )

    return redirect('/verify_otp')


@app.route('/verify_otp', methods=['GET', 'POST'])
def verify_otp():
    uid = session['user_id']

    if request.method == 'POST':
        otp_input = request.form['otp'].strip()
        otp_hash = hashlib.sha256(otp_input.encode()).hexdigest()

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT expires_at
            FROM password_reset_otp
            WHERE user_id = %s
              AND otp_hash = %s
              AND verified = 0
        """, (uid, otp_hash))

        row = cursor.fetchone()

        if not row or row[0] < datetime.now():
            cursor.close()
            conn.close()
            flash("Invalid or expired OTP", "error")
            return redirect('/verify_otp')

        cursor.execute("""
            UPDATE password_reset_otp
            SET verified = 1
            WHERE user_id = %s
        """, (uid,))

        conn.commit()
        cursor.close()
        conn.close()

        return redirect('/reset_password')

    return render_template('verify_otp.html')



from flask import request
from datetime import datetime
from werkzeug.security import generate_password_hash

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    uid = session['user_id']

    if request.method == 'POST':
        password = request.form['password']
        confirm = request.form['confirm_password']

        if password != confirm:
            flash("Passwords do not match", "error")
            return redirect('/reset_password')

        if len(password) < 8:
            flash("Password must be at least 8 characters", "error")
            return redirect('/reset_password')

        hashed_password = generate_password_hash(password)

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # 🔹 Get user email
        cursor.execute("SELECT email FROM users WHERE uid = %s", (uid,))
        email = cursor.fetchone()[0]

        # 🔹 Update password
        cursor.execute("""
            UPDATE users
            SET password = %s,
                password_status = 1,
                modified_at = NOW()
            WHERE uid = %s
        """, (hashed_password, uid))

        # 🔹 Cleanup OTP
        cursor.execute("DELETE FROM password_reset_otp WHERE user_id = %s", (uid,))
        conn.commit()

        cursor.close()
        conn.close()

        # 🔹 SECURITY EMAIL
        ip_address = request.remote_addr or "Unknown IP"
        user_agent = request.headers.get('User-Agent', 'Unknown Device')
        timestamp = datetime.now().strftime('%d %b %Y, %I:%M %p')

        email_body = f"""
Hello,

Your EduAI account password was successfully changed.

🔐 Password Change Details:
• Date & Time: {timestamp}
• IP Address: {ip_address}
• Device / Browser: {user_agent}

If YOU made this change, no further action is required.

⚠️ If you did NOT change your password:
• Immediately contact support
• Reset your password again
• Secure your account

— EduAI Security Team
"""

        send_email(
            receiver_email=email,
            subject="EduAI | Password Changed Successfully",
            body=email_body
        )

        flash("Password changed successfully. Please login again.", "success")
        session.clear()
        return redirect('/login')

    return render_template('reset_password.html')



@app.route('/login', methods=['GET', 'POST'])
def login():
    print(f"DEBUG: login() called, method={request.method}") # ABSOLUTE FIRST LINE
    if request.method == 'POST':
        email    = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()

        print(f"Login attempt: {email}") # Debugging

        if not email or not password:
            flash("Email and password are required!", "error")
            return redirect('/login')

        try:
            # ---------------------------------------------------------
            # 1. Validate user from `users` table
            # ---------------------------------------------------------
            print(f"Connecting to MySQL at {db_config['host']}:{db_config['port']}...")
            conn = mysql.connector.connect(**db_config)
            print("Connected successfully. Fetching user...")
            
            cur  = conn.cursor(dictionary=True)
            cur.execute(
                "SELECT uid, email, password, role FROM users WHERE email = %s",
                (email,)
            )
            user = cur.fetchone()
            cur.close()
            conn.close()
            print(f"Fetch complete. User found: {bool(user)}")

            if not user:
                print("User not found in database.") # Debugging
                flash("Invalid email or password.", "error")
                return redirect('/login')

            # ------------------- ADMIN -------------------
            if user['role'] == 'admin':
                if user['password'] == password:                     # plain-text admin
                    print("Admin login detected.")
                    session['user_id'] = user['uid']
                    session['email'] = user['email']
                    session['role'] = 'admin'
                    flash("Welcome back, Admin!", "success")
                    return redirect('/admin_dashboard')
                else:
                    flash("Invalid password.", "error")
                    return redirect('/login')

            # ------------------- HASHED PASSWORD (teacher/student) -------------------
            if not check_password_hash(user['password'], password):
                print("Password hash check failed.") # Debugging
                flash("Invalid password.", "error")
                return redirect('/login')

            print(f"Login successful for {user['role']}: {email}") # Debugging
            
            # Explicit session assignment (avoid update() just in case)
            session['user_id'] = user['uid']
            session['email'] = user['email']
            session['role'] = user['role']

            # ---------------------------------------------------------
            # 2. TEACHER PROFILE CHECK (unchanged)
            # ---------------------------------------------------------
            if user['role'] == 'teacher':
                conn = mysql.connector.connect(**db_config)
                cur  = conn.cursor(dictionary=True)
                cur.execute("SELECT * FROM teachers WHERE teacher_id = %s", (user['uid'],))
                teacher = cur.fetchone()
                cur.close()
                conn.close()

                required = [
                    'full_name','dob','last_degree','contact','gender',
                    'address','expertise','subjects_taught','experience_years',
                    'industry_experience_years','research_papers'
                ]
                incomplete = any(teacher.get(f) in (None, '', 'Not Provided') for f in required) if teacher else True

                if incomplete:
                    print("Teacher profile incomplete, redirecting...") # Debugging
                    flash("Please complete your profile first.", "info")
                    return redirect(url_for('complete_teacher_profile'))

                flash("Welcome back, Teacher!", "success")
                return redirect('/teacher_dashboard')

            # ---------------------------------------------------------
            # 3. STUDENT PROFILE CHECK – **IN-LINE**, just like teacher
            # ---------------------------------------------------------
            if user['role'] == 'student':
                conn = mysql.connector.connect(**db_config)
                cur  = conn.cursor(dictionary=True)

                # Ensure row exists with defaults (runs only once)
                cur.execute("""
                    INSERT IGNORE INTO students
                    (student_id, department, university)
                    VALUES (%s, %s, %s)
                """, (user['uid'],
                      "Department of Computer Science",
                      "Gujarat University"))
                conn.commit()

                # Now fetch the fields the student must fill
                cur.execute("""
                    SELECT full_name, contact, dob, gender, address
                    FROM students
                    WHERE student_id = %s
                """, (user['uid'],))
                student = cur.fetchone()
                cur.close()
                conn.close()

                # These are the ONLY fields the student fills
                required_student_fields = ['full_name', 'contact', 'dob', 'gender', 'address']

                # If any of them are missing → send to profile page
                if not student or any(student.get(f) in (None, '', 'Not Provided') for f in required_student_fields):
                    print("Student profile incomplete, redirecting...") # Debugging
                    flash("Please complete your profile first.", "info")
                    return redirect(url_for('complete_student_profile'))

                flash("Welcome back, Student!", "success")
                return redirect('/student_dashboard')

        except mysql.connector.Error as err:
            print(f"DATABASE ERROR during login: {err}") # Debugging
            flash(f"Database error: {err}", "error")
            return redirect('/login')
        except Exception as e:
            import traceback
            print(f"UNEXPECTED ERROR during login: {e}") # Debugging
            traceback.print_exc() # Print full stack trace
            flash(f"An unexpected error occurred: {e}", "error")
            return redirect('/login')

    # GET → show login page
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('email', None)
    session.pop('role', None)
    flash("You have been logged out.", "success")
    return redirect('/')



from datetime import datetime, time, timedelta
from collections import Counter
import mysql.connector
from flask import flash, redirect, render_template, session

@app.route('/teacher_dashboard')
def teacher_dashboard():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')
    
    teacher_id = session['user_id']  # current teacher's UID

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Fetch teacher full name
        cursor.execute("SELECT full_name FROM teachers WHERE teacher_id = %s", (teacher_id,))
        teacher_row = cursor.fetchone()
        name = teacher_row['full_name'] if teacher_row else session.get('email')

        # Fetch exams created by this teacher with course, semester, and subject names
        cursor.execute("""
            SELECT 
                e.exam_id,
                e.exam_name,
                e.course_id,
                c.course_name,
                e.semester_id,
                s.semester_name,
                e.subject_id,
                sub.subject_name,
                e.topic,
                e.max_marks,
                e.min_marks,
                e.exam_date,
                e.start_time,
                e.end_time
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.teacher_id = %s
            ORDER BY e.exam_date DESC
        """, (teacher_id,))
        exams = cursor.fetchall()

        # Calculate status for each exam
        current_datetime = datetime.now()
        for exam in exams:
            exam_date = exam['exam_date']
            start_time_delta = exam['start_time']
            end_time_delta = exam['end_time']

            # Convert timedelta to time
            if isinstance(start_time_delta, timedelta):
                total_seconds = int(start_time_delta.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                start_time = time(hours, minutes, seconds)
            else:
                start_time = start_time_delta

            if isinstance(end_time_delta, timedelta):
                total_seconds = int(end_time_delta.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                end_time = time(hours, minutes, seconds)
            else:
                end_time = end_time_delta

            # Format time strings for display
            exam['start_time_str'] = start_time.strftime('%H:%M')
            exam['end_time_str'] = end_time.strftime('%H:%M')

            # Calculate status
            start_dt = datetime.combine(exam_date, start_time)
            end_dt = datetime.combine(exam_date, end_time)
            if current_datetime < start_dt:
                exam['status'] = 'upcoming'
            elif start_dt <= current_datetime <= end_dt:
                exam['status'] = 'ongoing'
            else:
                exam['status'] = 'completed'

        # Calculate status counts
        status_counts = Counter(exam['status'] for exam in exams)

        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        exams = []
        status_counts = Counter()
        name = session.get('email')

    return render_template('teacher_dashboard.html', exams=exams, status_counts=status_counts, name=name)


from datetime import datetime, date, time, timedelta

@app.route('/view_exam/<int:exam_id>')
def view_exam(exam_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Fetch exam details
        cursor.execute("""
            SELECT 
                e.exam_id, e.exam_name, e.course_id, c.course_name,
                e.semester_id, s.semester_name, e.subject_id, sub.subject_name,
                e.topic, e.max_marks, e.min_marks,
                e.exam_date, e.start_time, e.end_time
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.exam_id = %s AND e.teacher_id = %s
        """, (exam_id, session['user_id']))
        exam = cursor.fetchone()

        if not exam:
            flash("Exam not found or you don't have permission to view it.", "error")
            cursor.close()
            conn.close()
            return redirect('/teacher_dashboard')

        # --- Convert time if stored as timedelta ---
        if exam.get("start_time") and isinstance(exam["start_time"], timedelta):
            exam["start_time"] = (datetime.min + exam["start_time"]).time()
        if exam.get("end_time") and isinstance(exam["end_time"], timedelta):
            exam["end_time"] = (datetime.min + exam["end_time"]).time()

        # --- Determine if exam is upcoming ---
        exam_datetime = datetime.combine(exam['exam_date'], exam['end_time'])  # use end_time to be safe
        now = datetime.now()
        exam_is_upcoming = exam_datetime > now
        exam['is_upcoming'] = exam_is_upcoming  # pass to template

        # --- Fetch questions ---
        cursor.execute("""
            SELECT question_id, question_text, model_answer, max_score
            FROM questions
            WHERE exam_id = %s
            ORDER BY question_id
        """, (exam_id,))
        questions = cursor.fetchall()

        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/teacher_dashboard')

    return render_template(
        'view_exam.html',
        exam=exam,
        questions=questions,
        name=session.get('email')
    )


def parse_time(time_input):
    if isinstance(time_input, timedelta):
        # Convert timedelta to time
        total_seconds = int(time_input.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return time(hours, minutes, seconds)
    
    # Handle string input
    time_formats = ['%H:%M', '%H:%M:%S', '%H:%M:%S.%f']
    for fmt in time_formats:
        try:
            return datetime.strptime(time_input, fmt).time()
        except (ValueError, TypeError):
            continue
    raise ValueError("Invalid time format")

@app.route('/edit_exam/<int:exam_id>', methods=['GET', 'POST'])
def edit_exam(exam_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Fetch exam details
        cursor.execute("""
            SELECT 
                e.exam_id,
                e.exam_name,
                e.course_id,
                c.course_name,
                e.semester_id,
                s.semester_name,
                e.subject_id,
                sub.subject_name,
                e.topic,
                e.max_marks,
                e.min_marks,
                e.exam_date,
                e.start_time,
                e.end_time
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.exam_id = %s AND e.teacher_id = %s
        """, (exam_id, session['user_id']))
        exam = cursor.fetchone()

        if not exam:
            flash("Exam not found or you don't have permission to edit it.", "error")
            cursor.close()
            conn.close()
            return redirect('/teacher_dashboard')

        # Convert timedelta to time for template
        exam['start_time'] = parse_time(exam['start_time'])
        exam['end_time'] = parse_time(exam['end_time'])

        # Fetch questions
        cursor.execute("""
            SELECT 
                question_id,
                question_text,
                model_answer,
                max_score
            FROM questions
            WHERE exam_id = %s
            ORDER BY question_id
        """, (exam_id,))
        questions = cursor.fetchall()

        if request.method == 'POST':
            # Handle form submission
            exam_name = request.form.get('exam_name', '').strip()
            topic = request.form.get('topic', '').strip()
            min_passing_percentage = request.form.get('min_passing_percentage', '').strip()
            exam_date = request.form.get('exam_date', '').strip()
            start_time = request.form.get('start_time', '').strip()
            end_time = request.form.get('end_time', '').strip()
            
            # Validation
            errors = []
            if not exam_name or len(exam_name) < 2 or len(exam_name) > 255:
                errors.append("Exam name must be 2–255 characters.")
            if not topic or len(topic) < 2 or len(topic) > 255:
                errors.append("Topic must be 2–255 characters.")
            try:
                min_passing_percentage = float(min_passing_percentage)
                if min_passing_percentage < 0 or min_passing_percentage > 100:
                    errors.append("Minimum passing percentage must be 0–100.")
            except ValueError:
                errors.append("Minimum passing percentage must be a number.")
            try:
                exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
                if exam_date < datetime.now().date():
                    errors.append("Exam date cannot be in the past.")
            except ValueError:
                errors.append("Invalid exam date. Use YYYY-MM-DD.")
            try:
                start_time = parse_time(start_time)
                end_time = parse_time(end_time)
                if start_time >= end_time:
                    errors.append("End time must be after start time.")
            except ValueError:
                errors.append("Invalid start or end time. Use HH:MM format (e.g., 14:30).")
            
            # Validate and collect updated question marks
            new_max_marks = 0
            updated_questions = []
            for q in questions:
                mark_key = f"marks_{q['question_id']}"
                if mark_key in request.form:
                    try:
                        max_score = float(request.form[mark_key])
                        if max_score < 0:
                            errors.append(f"Marks for question {q['question_id']} cannot be negative.")
                        updated_questions.append({
                            'question_id': q['question_id'],
                            'max_score': max_score
                        })
                        new_max_marks += max_score
                    except ValueError:
                        errors.append(f"Invalid mark value for question {q['question_id']}.")
            
            if errors:
                for e in errors:
                    flash(e, "error")
                cursor.close()
                conn.close()
                return render_template('edit_exam.html', exam=exam, questions=questions, name=session.get('email'), errors=errors)
            
            # Calculate min_marks
            min_marks = int((min_passing_percentage / 100) * new_max_marks)
            
            # Update exams table
            try:
                cursor.execute("""
                    UPDATE exams
                    SET exam_name = %s, topic = %s, max_marks = %s, min_marks = %s, exam_date = %s, start_time = %s, end_time = %s
                    WHERE exam_id = %s AND teacher_id = %s
                """, (exam_name, topic, new_max_marks, min_marks, exam_date, start_time, end_time, exam_id, session['user_id']))
                
                # Update questions
                for q in updated_questions:
                    cursor.execute("""
                        UPDATE questions
                        SET max_score = %s
                        WHERE question_id = %s AND exam_id = %s
                    """, (q['max_score'], q['question_id'], exam_id))
                
                conn.commit()
                flash("Exam updated successfully!", "success")
                cursor.close()
                conn.close()
                return redirect('/teacher_dashboard')
            except mysql.connector.Error as err:
                flash(f"Database error: {err}", "error")
                cursor.close()
                conn.close()
                return render_template('edit_exam.html', exam=exam, questions=questions, name=session.get('email'), errors=errors)
        
        cursor.close()
        conn.close()
        return render_template('edit_exam.html', exam=exam, questions=questions, name=session.get('email'))
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/teacher_dashboard')

# Custom Jinja2 filter for formatting time
def strftime(value, format_string):
    if value is None:
        return ""
    if isinstance(value, timedelta):
        # Convert timedelta to time
        total_seconds = int(value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        value = time(hours, minutes, seconds)
    return value.strftime(format_string)

app.jinja_env.filters['strftime'] = strftime



@app.route('/teacher/complete_profile', methods=['GET', 'POST'])
def complete_teacher_profile():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')

    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    # Fetch existing teacher profile
    cursor.execute("SELECT * FROM teachers WHERE teacher_id = %s", (session['user_id'],))
    existing_profile = cursor.fetchone()

    if request.method == 'POST':
        # Case 1: Profile already complete
        if existing_profile and all([
            existing_profile.get('full_name'),
            existing_profile.get('dob'),
            existing_profile.get('last_degree'),
            existing_profile.get('contact'),
            existing_profile.get('gender'),
            existing_profile.get('address'),
            existing_profile.get('expertise'),
            existing_profile.get('subjects_taught'),
            existing_profile.get('experience_years') is not None,
            existing_profile.get('industry_experience_years') is not None,
            existing_profile.get('research_papers') is not None,
            existing_profile.get('department'),
            existing_profile.get('university')
        ]):
            flash("Profile already completed.", "info")
            cursor.close()
            conn.close()
            return redirect('/teacher_dashboard')

        # Case 2: Skip button
        if request.form.get('skip') == '1':
            cursor.close()
            conn.close()
            return redirect('/teacher_dashboard')

        # Get inputs
        full_name = request.form.get('full_name', '').strip()
        dob = request.form.get('dob', '').strip()
        last_degree = request.form.get('last_degree', '').strip()
        contact = request.form.get('contact', '').strip()
        gender = request.form.get('gender', '').strip()
        address = request.form.get('address', '').strip()
        expertise = request.form.get('expertise', '').strip()
        subjects_taught = request.form.get('subjects_taught', '').strip()
        experience_years = request.form.get('experience_years', '').strip()
        industry_experience_years = request.form.get('industry_experience_years', '').strip()
        research_papers = request.form.get('research_papers', '').strip()
        department = request.form.get('department', '').strip()
        university = request.form.get('university', '').strip()

        # ---------- VALIDATION ----------
        errors = []

        # Full name: letters and spaces only
        if not re.match(r'^[A-Za-z ]{3,50}$', full_name):
            errors.append("Full name must be 3–50 alphabetic characters.")

        # DOB: valid date & realistic (e.g., teacher age > 21 years)
        try:
            dob_date = datetime.strptime(dob, '%Y-%m-%d').date()
            age = (datetime.now().date() - dob_date).days // 365
            if age < 21:
                errors.append("Teacher must be at least 21 years old.")
        except ValueError:
            errors.append("Invalid date format for DOB. Use YYYY-MM-DD.")

        # Last degree: not empty
        if len(last_degree) < 2:
            errors.append("Last degree must be at least 2 characters.")

        # Contact: numeric and length 10 digits
        if not re.match(r'^\d{10}$', contact):
            errors.append("Contact number must be exactly 10 digits.")

        # Gender: must be one of fixed values
        if gender not in ['Male', 'Female', 'Other']:
            errors.append("Invalid gender selected.")

        # Address: not empty
        if len(address) < 5:
            errors.append("Address must be at least 5 characters.")

        # Expertise: not empty
        if len(expertise) < 2:
            errors.append("Expertise must be at least 2 characters.")

        # Subjects taught: not empty
        if len(subjects_taught) < 2:
            errors.append("Subjects taught must be at least 2 characters.")

        # Experience years: non-negative integer
        try:
            experience_years = int(experience_years)
            if experience_years < 0:
                errors.append("Experience years cannot be negative.")
        except ValueError:
            errors.append("Experience years must be a valid number.")

        # Industry experience years: non-negative integer
        try:
            industry_experience_years = int(industry_experience_years)
            if industry_experience_years < 0:
                errors.append("Industry experience years cannot be negative.")
        except ValueError:
            errors.append("Industry experience years must be a valid number.")

        # Research papers: non-negative integer
        try:
            research_papers = int(research_papers)
            if research_papers < 0:
                errors.append("Research papers cannot be negative.")
        except ValueError:
            errors.append("Research papers must be a valid number.")

        # Department: not empty
        if len(department) < 2:
            errors.append("Department must be at least 2 characters.")

        # University: not empty
        if len(university) < 2:
            errors.append("University name must be at least 2 characters.")

        # If validation failed
        if errors:
            for e in errors:
                flash(e, "error")
            cursor.close()
            conn.close()
            return render_template('complete_teacher_profile.html', name=session.get('email'))

        # ---------- SAVE TO DB ----------
        try:
            if existing_profile:
                # Update existing
                cursor.execute("""
                    UPDATE teachers 
                    SET full_name=%s, dob=%s, last_degree=%s, contact=%s, gender=%s,
                        address=%s, expertise=%s, subjects_taught=%s,
                        experience_years=%s, industry_experience_years=%s,
                        research_papers=%s, department=%s, university=%s
                    WHERE teacher_id=%s
                """, (
                    full_name, dob_date, last_degree, contact, gender,
                    address, expertise, subjects_taught,
                    experience_years, industry_experience_years,
                    research_papers, department, university, session['user_id']
                ))
            else:
                # Insert new
                cursor.execute("""
                    INSERT INTO teachers (
                        teacher_id, full_name, dob, last_degree, contact, gender, address,
                        expertise, subjects_taught, experience_years, industry_experience_years,
                        research_papers, department, university
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    session['user_id'], full_name, dob_date, last_degree, contact, gender,
                    address, expertise, subjects_taught,
                    experience_years, industry_experience_years,
                    research_papers, department, university
                ))
            conn.commit()
            flash("Profile completed successfully!", "success")
        except mysql.connector.Error as err:
            flash(f"Database error: {err}", "error")

        cursor.close()
        conn.close()
        return redirect('/teacher_dashboard')

    # GET request
    cursor.close()
    conn.close()
    return render_template('complete_teacher_profile.html', name=session.get('email'))




# Add or update these routes in app.py for teacher profile display and edit

from datetime import datetime
import mysql.connector
from flask import render_template, request, redirect, flash, session
import re

@app.route('/teacher/profile')
def teacher_profile_display():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT t.full_name, t.dob, t.last_degree, t.contact, t.gender, t.address, t.expertise, 
                   t.subjects_taught, t.experience_years, t.industry_experience_years, t.research_papers, 
                   t.department, t.university, u.email
            FROM teachers t
            JOIN users u ON t.teacher_id = u.uid
            WHERE t.teacher_id = %s
            """,
            (session['user_id'],)
        )
        profile = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not profile:
            flash("No profile found. Please complete your profile.", "warning")
            return redirect('/teacher/complete_profile')
        
        return render_template('teacher_profile.html', profile=profile, email=profile[13], name=session.get('email'))
    
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/teacher_dashboard')

@app.route('/teacher/edit_profile', methods=['GET', 'POST'])
def teacher_edit_profile():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT t.full_name, t.dob, t.last_degree, t.contact, t.gender, t.address, t.expertise, 
                   t.subjects_taught, t.experience_years, t.industry_experience_years, t.research_papers, 
                   t.department, t.university, u.email
            FROM teachers t
            JOIN users u ON t.teacher_id = u.uid
            WHERE t.teacher_id = %s
            """,
            (session['user_id'],)
        )
        profile = cursor.fetchone()
        
        if request.method == 'POST':
            full_name = request.form.get('full_name', '').strip()
            dob = request.form.get('dob', '').strip()
            last_degree = request.form.get('last_degree', '').strip()
            contact = request.form.get('contact', '').strip()
            gender = request.form.get('gender', '').strip()
            address = request.form.get('address', '').strip()
            expertise = request.form.get('expertise', '').strip()
            subjects_taught = request.form.get('subjects_taught', '').strip()
            experience_years = request.form.get('experience_years', '').strip()
            industry_experience_years = request.form.get('industry_experience_years', '').strip()
            research_papers = request.form.get('research_papers', '').strip()
            department = request.form.get('department', '').strip()
            university = request.form.get('university', '').strip()

            # ---------- VALIDATION ----------
            errors = []

            # Full name: letters and spaces only
            if not re.match(r'^[A-Za-z ]{3,50}$', full_name):
                errors.append("Full name must be 3–50 alphabetic characters.")

            # DOB: valid date & realistic (e.g., teacher age > 21 years)
            try:
                dob_date = datetime.strptime(dob, '%Y-%m-%d').date()
                age = (datetime.now().date() - dob_date).days // 365
                if age < 21:
                    errors.append("Teacher must be at least 21 years old.")
            except ValueError:
                errors.append("Invalid date format for DOB. Use YYYY-MM-DD.")

            # Last degree: not empty
            if len(last_degree) < 2:
                errors.append("Last degree must be at least 2 characters.")

            # Contact: numeric and length 10 digits
            if not re.match(r'^\d{10}$', contact):
                errors.append("Contact number must be exactly 10 digits.")

            # Gender: must be one of fixed values
            if gender not in ['Male', 'Female', 'Other']:
                errors.append("Invalid gender selected.")

            # Address: not empty
            if len(address) < 5:
                errors.append("Address must be at least 5 characters.")

            # Expertise: not empty
            if len(expertise) < 2:
                errors.append("Expertise must be at least 2 characters.")

            # Subjects taught: not empty
            if len(subjects_taught) < 2:
                errors.append("Subjects taught must be at least 2 characters.")

            # Experience years: non-negative integer
            try:
                experience_years = int(experience_years)
                if experience_years < 0:
                    errors.append("Experience years cannot be negative.")
            except ValueError:
                errors.append("Experience years must be a valid number.")

            # Industry experience years: non-negative integer
            try:
                industry_experience_years = int(industry_experience_years)
                if industry_experience_years < 0:
                    errors.append("Industry experience years cannot be negative.")
            except ValueError:
                errors.append("Industry experience years must be a valid number.")

            # Research papers: non-negative integer
            try:
                research_papers = int(research_papers)
                if research_papers < 0:
                    errors.append("Research papers cannot be negative.")
            except ValueError:
                errors.append("Research papers must be a valid number.")

            # Department: not empty
            if len(department) < 2:
                errors.append("Department must be at least 2 characters.")

            # University: not empty
            if len(university) < 2:
                errors.append("University name must be at least 2 characters.")

            if errors:
                for e in errors:
                    flash(e, "error")
                cursor.close()
                conn.close()
                return render_template('teacher_edit_profile.html', profile=profile, email=profile[13] if profile else session.get('email'), name=session.get('email'))

            cursor.execute(
                """
                UPDATE teachers 
                SET full_name=%s, dob=%s, last_degree=%s, contact=%s, gender=%s,
                    address=%s, expertise=%s, subjects_taught=%s,
                    experience_years=%s, industry_experience_years=%s,
                    research_papers=%s, department=%s, university=%s
                WHERE teacher_id=%s
                """,
                (
                    full_name, dob_date, last_degree, contact, gender,
                    address, expertise, subjects_taught,
                    experience_years, industry_experience_years,
                    research_papers, department, university, session['user_id']
                )
            )
            conn.commit()
            flash("Profile updated successfully!", "success")
            cursor.close()
            conn.close()
            return redirect('/teacher/profile')
        
        cursor.close()
        conn.close()
        return render_template('teacher_edit_profile.html', profile=profile, email=profile[13] if profile else session.get('email'), name=session.get('email'))
    
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/teacher_dashboard')
    
    
    

@app.route('/download_question_template')
def download_question_template():
    # Create sample dataframe - Only Question and Maximum Marks (Answer extracted from PDF)
    df = pd.DataFrame({
        "Question": [
            "Explain the concept of Operating Systems.",
            "What is a database index?"
        ],
        "Maximum Marks": [5, 3]
    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Questions')

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Question_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    

    

ALLOWED_PDF_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_pdf_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_PDF_EXTENSIONS


# ────────────────────────────────────────────────────────────────
# SERVER-SIDE TEMP DATA HELPERS
# Flask's cookie-based session has a ~4KB limit. AI-generated model
# answers easily exceed this, causing silent data loss. These helpers
# store the data in a server-side JSON file instead.
# ────────────────────────────────────────────────────────────────
TEMP_EXAM_DIR = os.path.join('static', 'uploads', 'temp_exam_data')

def _save_temp_exam_data(user_id, data):
    """Save exam creation temp data to a server-side JSON file."""
    os.makedirs(TEMP_EXAM_DIR, exist_ok=True)
    filepath = os.path.join(TEMP_EXAM_DIR, f"exam_temp_{user_id}.json")
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

def _load_temp_exam_data(user_id):
    """Load exam creation temp data from a server-side JSON file."""
    filepath = os.path.join(TEMP_EXAM_DIR, f"exam_temp_{user_id}.json")
    if not os.path.exists(filepath):
        return None
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return None

def _delete_temp_exam_data(user_id):
    """Delete the server-side temp file after successful exam creation."""
    filepath = os.path.join(TEMP_EXAM_DIR, f"exam_temp_{user_id}.json")
    if os.path.exists(filepath):
        os.remove(filepath)

@app.route('/create_exam', methods=['GET', 'POST'])
def create_exam():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Fetch courses (assuming teacher has access to all courses; adjust if teacher-specific)
        cursor.execute("SELECT course_id, course_name FROM courses")
        courses = cursor.fetchall()
        
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        courses = []

    if request.method == 'POST':
        if 'upload_excel' in request.form:
            # Step 1: Handle selection, Excel upload + PDF upload, then AI answer extraction
            course_id = request.form.get('course_id')
            semester_id = request.form.get('semester_id')
            subject_id = request.form.get('subject_id')
            topic = request.form.get('topic', '').strip()
            
            # Validation for selections
            errors = []
            if not course_id or not course_id.isdigit():
                errors.append("Course is required.")
            if not semester_id or not semester_id.isdigit():
                errors.append("Semester is required.")
            if not subject_id or not subject_id.isdigit():
                errors.append("Subject is required.")
            if not topic or len(topic) < 2 or len(topic) > 255:
                errors.append("Topic must be 2–255 characters.")
            
            # Validate Excel file
            if 'excel_file' not in request.files:
                errors.append("No Excel file uploaded.")
            
            file = request.files['excel_file']
            if file.filename == '':
                errors.append("No Excel file selected.")
            if file and not allowed_file(file.filename):
                errors.append("Invalid Excel file type. Only XLS/XLSX allowed.")
            
            # Validate PDF file
            if 'pdf_file' not in request.files:
                errors.append("No PDF file uploaded.")
            
            pdf_file = request.files.get('pdf_file')
            if pdf_file and pdf_file.filename == '':
                errors.append("No PDF file selected.")
            if pdf_file and pdf_file.filename and not allowed_pdf_file(pdf_file.filename):
                errors.append("Invalid PDF file type. Only .pdf files are allowed.")
            
            if errors:
                for e in errors:
                    flash(e, "error")
                return render_template('create_exam.html', courses=courses, name=session.get('email'), errors=errors)
            
            # Save the Excel file temporarily
            filename = secure_filename(file.filename)
            upload_folder = 'static/uploads'
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder, exist_ok=True)
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)
            
            # Save the PDF file temporarily
            pdf_filename = secure_filename(pdf_file.filename)
            pdf_path = os.path.join(upload_folder, pdf_filename)
            pdf_file.save(pdf_path)
            
            # Read and validate Excel
            try:
                df = pd.read_excel(file_path)
                # Normalize column names (case-insensitive for validation)
                normalized_columns = [col.strip().lower() for col in df.columns]
                
                # Expected columns - now only Question and Maximum Marks (Answer comes from PDF)
                required_columns = {'question', 'maximum marks'}
                if not required_columns.issubset(normalized_columns):
                    flash("Excel must have columns: Question, Maximum Marks (case-insensitive).", "error")
                    os.remove(file_path)
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
                    return render_template('create_exam.html', courses=courses, name=session.get('email'))
                
                # Create a mapping from normalized to original columns
                column_mapping = {col.strip().lower(): col for col in df.columns}
                # Rename columns to standard internal names
                df = df.rename(columns={
                    column_mapping.get('question', 'Question'): 'question_text',
                    column_mapping.get('maximum marks', 'Maximum Marks'): 'max_score'
                })
                
                # Validate max_score are numbers
                errors = []
                for idx, row in df.iterrows():
                    try:
                        max_score = float(row['max_score'])
                        if max_score < 0:
                            errors.append(f"Invalid maximum marks at row {idx + 2}: must be a non-negative number.")
                    except (ValueError, TypeError):
                        errors.append(f"Invalid maximum marks at row {idx + 2}: must be a number (found '{row['max_score']}').")
                
                if errors:
                    for e in errors:
                        flash(e, "error")
                    os.remove(file_path)
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
                    return render_template('create_exam.html', courses=courses, name=session.get('email'), errors=errors)
                
                # Convert max_score to Python float for JSON serialization
                df['max_score'] = df['max_score'].astype(float)
                
                # Prepare questions data for AI extraction
                questions_for_extraction = []
                for idx, row in df.iterrows():
                    questions_for_extraction.append({
                        'question_text': str(row['question_text']),
                        'max_score': float(row['max_score'])
                    })
                
                # Extract answers from PDF using Groq AI
                try:
                    questions = extract_answers_from_pdf(questions_for_extraction, pdf_path)
                except Exception as ai_err:
                    flash(f"Error extracting answers from PDF: {str(ai_err)}", "error")
                    os.remove(file_path)
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
                    return render_template('create_exam.html', courses=courses, name=session.get('email'))
                
                # Calculate total max_marks
                total_max_marks = float(sum(q['max_score'] for q in questions))
                
                # Store data in server-side temp file (avoids Flask 4KB cookie limit)
                _save_temp_exam_data(session['user_id'], {
                    'course_id': course_id,
                    'semester_id': semester_id,
                    'subject_id': subject_id,
                    'topic': topic,
                    'questions': questions,
                    'max_marks': total_max_marks,
                    'question_excel_path': file_path,
                    'pdf_extracted': True
                })
                
                # Clean up PDF file after extraction
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                
                flash(f"Successfully extracted answers for {len(questions)} questions from the PDF using AI!", "success")
                
                return render_template('create_exam.html', courses=courses, name=session.get('email'), questions=questions, 
                                     course_id=course_id, semester_id=semester_id, subject_id=subject_id, topic=topic, 
                                     max_marks=total_max_marks, show_final_form=False, change_marks=False, pdf_extracted=True)
            
            except Exception as e:
                flash(f"Error processing files: {str(e)}", "error")
                if os.path.exists(file_path):
                    os.remove(file_path)
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                return render_template('create_exam.html', courses=courses, name=session.get('email'))
        
        elif 'confirm_max_marks' in request.form:
            # Step 2: Confirm max_marks, show final form
            temp_data = _load_temp_exam_data(session['user_id'])
            if not temp_data:
                flash("No exam data found. Please start over.", "error")
                return render_template('create_exam.html', courses=courses, name=session.get('email'))
            
            return render_template('create_exam.html', courses=courses, name=session.get('email'), 
                                 questions=temp_data['questions'], course_id=temp_data['course_id'], 
                                 semester_id=temp_data['semester_id'], subject_id=temp_data['subject_id'], 
                                 topic=temp_data['topic'], max_marks=temp_data['max_marks'], 
                                 show_final_form=True, change_marks=False)
        
        elif 'change_max_marks' in request.form:
            # Step 3: Show editable marks table
            temp_data = _load_temp_exam_data(session['user_id'])
            if not temp_data:
                flash("No exam data found. Please start over.", "error")
                return render_template('create_exam.html', courses=courses, name=session.get('email'))
            
            return render_template('create_exam.html', courses=courses, name=session.get('email'), 
                                 questions=temp_data['questions'], course_id=temp_data['course_id'], 
                                 semester_id=temp_data['semester_id'], subject_id=temp_data['subject_id'], 
                                 topic=temp_data['topic'], max_marks=temp_data['max_marks'], 
                                 show_final_form=False, change_marks=True)
        
        elif 'update_marks' in request.form:
            # Step 4: Update marks from editable table
            temp_data = _load_temp_exam_data(session['user_id'])
            if not temp_data:
                flash("No exam data found. Please start over.", "error")
                return render_template('create_exam.html', courses=courses, name=session.get('email'))
            
            questions = temp_data['questions']
            new_max_marks = 0
            errors = []
            
            for idx, q in enumerate(questions):
                mark_key = f"marks_{idx + 1}"
                if mark_key in request.form:
                    try:
                        max_score = float(request.form[mark_key])
                        if max_score < 0:
                            errors.append(f"Marks for question {idx + 1} cannot be negative.")
                        q['max_score'] = float(max_score)
                        new_max_marks += max_score
                    except ValueError:
                        errors.append(f"Invalid mark value for question {idx + 1}.")
            
            if errors:
                for e in errors:
                    flash(e, "error")
                return render_template('create_exam.html', courses=courses, name=session.get('email'), 
                                     questions=questions, course_id=temp_data['course_id'], 
                                     semester_id=temp_data['semester_id'], subject_id=temp_data['subject_id'], 
                                     topic=temp_data['topic'], max_marks=new_max_marks, 
                                     show_final_form=False, change_marks=True, errors=errors)
            
            # Update max_marks in server-side temp file
            temp_data['max_marks'] = float(new_max_marks)
            _save_temp_exam_data(session['user_id'], temp_data)
            
            return render_template('create_exam.html', courses=courses, name=session.get('email'), 
                                 questions=questions, course_id=temp_data['course_id'], 
                                 semester_id=temp_data['semester_id'], subject_id=temp_data['subject_id'], 
                                 topic=temp_data['topic'], max_marks=new_max_marks, 
                                 show_final_form=False, change_marks=False)
        
        elif 'final_submit' in request.form:
            # Step 5: Handle final submission
            temp_data = _load_temp_exam_data(session['user_id'])
            if not temp_data:
                flash("No exam data found. Please start over.", "error")
                return render_template('create_exam.html', courses=courses, name=session.get('email'))
            
            # Remove temp file (equivalent to session.pop)
            _delete_temp_exam_data(session['user_id'])
            course_id = temp_data['course_id']
            semester_id = temp_data['semester_id']
            subject_id = temp_data['subject_id']
            topic = temp_data['topic']
            max_marks = temp_data['max_marks']
            file_path = temp_data['question_excel_path']
            questions = temp_data['questions']
            
            min_passing_percentage = request.form.get('min_passing_percentage', '').strip()
            exam_date = request.form.get('exam_date', '').strip()
            start_time = request.form.get('start_time', '').strip()
            end_time = request.form.get('end_time', '').strip()
            
            # Validation for final fields
            errors = []
            try:
                min_passing_percentage = float(min_passing_percentage)
                if min_passing_percentage < 0 or min_passing_percentage > 100:
                    errors.append("Minimum passing percentage must be 0–100.")
            except ValueError:
                errors.append("Minimum passing percentage must be a number.")
            
            try:
                exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
                if exam_date < datetime.now().date():
                    errors.append("Exam date cannot be in the past.")
            except ValueError:
                errors.append("Invalid exam date. Use YYYY-MM-DD.")
            
            try:
                start_time = datetime.strptime(start_time, '%H:%M').time()
                end_time = datetime.strptime(end_time, '%H:%M').time()
                if start_time >= end_time:
                    errors.append("End time must be after start time.")
            except ValueError:
                errors.append("Invalid start or end time. Use HH:MM.")
            
            if errors:
                for e in errors:
                    flash(e, "error")
                # Restore server-side temp data to allow retry
                _save_temp_exam_data(session['user_id'], temp_data)
                return render_template('create_exam.html', courses=courses, name=session.get('email'), 
                                     questions=questions, course_id=course_id, semester_id=semester_id, 
                                     subject_id=subject_id, topic=topic, max_marks=max_marks, 
                                     show_final_form=True, change_marks=False, errors=errors)
            
            # Calculate min_marks
            min_marks = int((min_passing_percentage / 100) * max_marks)
            
            # Insert into exams table
            try:
                conn = mysql.connector.connect(**db_config)
                cursor = conn.cursor()
                
                exam_name = f"{topic} Exam"
                
                cursor.execute(
                    """
                    INSERT INTO exams (exam_name, teacher_id, course_id, semester_id, subject_id, topic, question_excel_path, 
                                       max_marks, min_marks, exam_date, start_time, end_time)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (exam_name, session['user_id'], course_id, semester_id, subject_id, topic, file_path, 
                     max_marks, min_marks, exam_date, start_time, end_time)
                )
                exam_id = cursor.lastrowid
                
#                # Bulk insert questions WITH keyword config
                for q in questions:
                    # Build keyword config once per question
                    keyword_config = build_keyword_config(
                        expected_answer=q['model_answer'],
                        max_marks=q['max_score']
                    )

                    keyword_config_json = json.dumps(keyword_config)

                    cursor.execute(
                        """
                        INSERT INTO questions (
                            exam_id,
                            course_id,
                            semester_id,
                            subject_id,
                            question_text,
                            model_answer,
                            max_score,
                            keyword_config
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            exam_id,
                            course_id,
                            semester_id,
                            subject_id,
                            q['question_text'],
                            q['model_answer'],
                            q['max_score'],   # ✅ FIXED
                            keyword_config_json
                        )
                    )

                
                conn.commit()
                flash("Exam created successfully!", "success")
                
                cursor.close()
                conn.close()
            except mysql.connector.Error as err:
                flash(f"Database error: {err}", "error")
                # Restore server-side temp data to allow retry
                _save_temp_exam_data(session['user_id'], temp_data)
                return render_template('create_exam.html', courses=courses, name=session.get('email'), 
                                     questions=questions, course_id=course_id, semester_id=semester_id, 
                                     subject_id=subject_id, topic=topic, max_marks=max_marks, 
                                     show_final_form=True, change_marks=False)
            
            # Clean up temporary file
            if os.path.exists(file_path):
                os.remove(file_path)
            
            return redirect('/teacher_dashboard')
        
    return render_template('create_exam.html', courses=courses, name=session.get('email'))




# --------------------------------------------------------------
#  TEACHER – GET SEMESTERS FOR A COURSE
# --------------------------------------------------------------
@app.route('/teacher/get_semesters/<int:course_id>')
def teacher_get_semesters(course_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        conn = mysql.connector.connect(**db_config)
        cur  = conn.cursor(dictionary=True)

        cur.execute(
            """SELECT semester_id, semester_name
               FROM semesters
               WHERE course_id = %s
               ORDER BY semester_name""",
            (course_id,)
        )
        semesters = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({'semesters': semesters})
    except mysql.connector.Error as e:
        return jsonify({'error': str(e)}), 500


# --------------------------------------------------------------
#  TEACHER – GET SUBJECTS FOR A SEMESTER
# --------------------------------------------------------------
@app.route('/teacher/get_subjects/<int:semester_id>')
def teacher_get_subjects(semester_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        conn = mysql.connector.connect(**db_config)
        cur  = conn.cursor(dictionary=True)

        cur.execute(
            """SELECT subject_id, subject_name
               FROM subjects
               WHERE semester_id = %s
               ORDER BY subject_name""",
            (semester_id,)
        )
        subjects = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({'subjects': subjects})
    except mysql.connector.Error as e:
        return jsonify({'error': str(e)}), 500
    
        



from datetime import datetime, date, time, timedelta

def td_to_time(td):
    """Convert MySQL TIME (timedelta) → time object"""
    if td is None or isinstance(td, time):
        return td
    secs = int(td.total_seconds())
    return time(secs // 3600, (secs % 3600) // 60, secs % 60)

@app.route('/view_answers')
def view_answers():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')

    teacher_id = session['user_id']

    # Filters
    course_id = request.args.get('course_id')
    semester_id = request.args.get('semester_id')
    subject_id = request.args.get('subject_id')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Base query
        query = """
            SELECT 
                e.exam_id, e.exam_name, e.topic, e.exam_date, e.start_time, e.end_time,
                e.evaluation_status,
                c.course_name, s.semester_name, sub.subject_name,
                COUNT(sa.answer_id) AS total_submissions,
                COUNT(ev.answer_id) AS evaluated_count
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            LEFT JOIN questions q ON e.exam_id = q.exam_id
            LEFT JOIN student_answers sa ON q.question_id = sa.question_id
            LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
            WHERE e.teacher_id = %s
        """
        params = [teacher_id]

        # Apply filters
        if course_id:
            query += " AND e.course_id = %s"
            params.append(course_id)
        if semester_id:
            query += " AND e.semester_id = %s"
            params.append(semester_id)
        if subject_id:
            query += " AND e.subject_id = %s"
            params.append(subject_id)
        if date_from:
            query += " AND e.exam_date >= %s"
            params.append(date_from)
        if date_to:
            query += " AND e.exam_date <= %s"
            params.append(date_to)

        query += " GROUP BY e.exam_id ORDER BY e.exam_date DESC, e.exam_id DESC"

        cursor.execute(query, params)
        exams = cursor.fetchall()

        # Stats
        total_exams = len(exams)
        total_submissions = sum(e['total_submissions'] or 0 for e in exams)
        total_evaluated = sum(e['evaluated_count'] or 0 for e in exams)
        pending_exams = sum(1 for e in exams if (e['total_submissions'] or 0) > (e['evaluated_count'] or 0))

        now = datetime.now()
        for exam in exams:
            # Time handling
            start_time = td_to_time(exam['start_time'])
            end_time = td_to_time(exam['end_time'])
            exam_end = datetime.combine(exam['exam_date'], end_time) if end_time else None
            exam_ended = exam_end and exam_end < now

            total = exam['total_submissions'] or 0
            evaluated = exam['evaluated_count'] or 0
            all_evaluated = total > 0 and total == evaluated
            is_finalized = exam['evaluation_status'] == 1

            # === Button Logic ===
            can_evaluate = exam_ended and total > 0 and not is_finalized
            can_finalize = exam_ended and all_evaluated and not is_finalized
            is_upcoming = not exam_ended

            # === Status & Badge ===
            if is_finalized:
                status = "Finalized"
                status_class = "finalized"
            elif is_upcoming:
                status = "Upcoming"
                status_class = "upcoming"
            elif all_evaluated:
                status = "Ready to Finalize"
                status_class = "ready"
            elif total > 0:
                status = "In Progress"
                status_class = "pending"
            else:
                status = "No Submissions"
                status_class = "none"

            exam.update({
                'start_time': start_time,
                'end_time': end_time,
                'exam_ended': exam_ended,
                'can_evaluate': can_evaluate,
                'can_finalize': can_finalize,
                'is_finalized': is_finalized,
                'status': status,
                'status_class': status_class
            })

        # Filter options
        cursor.execute("SELECT course_id, course_name FROM courses ORDER BY course_name")
        courses = cursor.fetchall()

        cursor.execute("""
            SELECT DISTINCT s.semester_id, s.semester_name 
            FROM semesters s JOIN exams e ON s.semester_id = e.semester_id 
            WHERE e.teacher_id = %s ORDER BY s.semester_name
        """, (teacher_id,))
        semesters = cursor.fetchall()

        cursor.execute("""
            SELECT DISTINCT sub.subject_id, sub.subject_name 
            FROM subjects sub JOIN exams e ON sub.subject_id = e.subject_id 
            WHERE e.teacher_id = %s ORDER BY sub.subject_name
        """, (teacher_id,))
        subjects = cursor.fetchall()

        cursor.close()
        conn.close()
    except Exception as e:
        flash(f"Database error: {e}", "error")
        exams = courses = semesters = subjects = []
        total_exams = total_submissions = total_evaluated = pending_exams = 0

    return render_template(
        'view_answers.html',
        exams=exams,
        courses=courses, semesters=semesters, subjects=subjects,
        selected_course=course_id, selected_semester=semester_id, selected_subject=subject_id,
        date_from=date_from, date_to=date_to,
        total_exams=total_exams,
        total_submissions=total_submissions,
        total_evaluated=total_evaluated,
        pending_exams=pending_exams,
        name=session.get('email')
    )
    
           

@app.route('/evaluate_exam_list/<int:exam_id>')
def evaluate_exam_list(exam_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        return redirect('/login')

    teacher_id = session['user_id']

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Verify exam
        cursor.execute("""
            SELECT e.exam_name, e.topic, e.max_marks, c.course_name, s.semester_name, 
                   sub.subject_name, e.evaluation_status
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.exam_id = %s AND e.teacher_id = %s
        """, (exam_id, teacher_id))
        exam = cursor.fetchone()
        if not exam:
            flash("Exam not found.", "error")
            return redirect('/view_answers')

        # Fetch students with their scores
        query="""
            SELECT
            st.student_id,
            st.roll_no,
            st.enrollment_no,
            st.full_name,

            COUNT(sa.answer_id) AS answers_given,
            COUNT(ev.answer_id) AS answers_evaluated,

            COALESCE(
                SUM(
                    CASE 
                        WHEN q.exam_id = %s THEN ev.score
                        ELSE 0
                    END
                ), 0
            ) AS total_score,

            r.final_score

        FROM questions q
        JOIN student_answers sa ON sa.question_id = q.question_id
        JOIN students st ON st.student_id = sa.student_id
        LEFT JOIN evaluations ev ON ev.answer_id = sa.answer_id
        LEFT JOIN result r ON r.student_id = st.student_id AND r.exam_id = %s

        WHERE q.exam_id = %s

        GROUP BY st.student_id
        ORDER BY st.roll_no;
        """
        cursor.execute(query, (exam_id, exam_id, exam_id))
    

        students = cursor.fetchall()
        
        # Compute status and final scores
        all_evaluated = True
        for s in students:
            s['all_evaluated'] = s['answers_given'] == s['answers_evaluated']
            s['status'] = 'Evaluated' if s['all_evaluated'] else 'Pending'
            
            # Convert to float and round to 2 decimals
            s['total_score'] = round(float(s['total_score']), 2) if s['total_score'] else 0.0
            print(s['total_score'])
            # Calculate final score (ceil/floor)
            if s['total_score'] > 0:
                fractional = s['total_score'] - int(s['total_score'])
                if fractional >= 0.5:
                    s['calculated_final_score'] = int(s['total_score']) + 1  # ceil
                else:
                    s['calculated_final_score'] = int(s['total_score'])  # floor
            else:
                s['calculated_final_score'] = 0
            
            if not s['all_evaluated']:
                all_evaluated = False

        exam_finalized = exam['evaluation_status'] == 1

        cursor.close()
        conn.close()
    except Exception as e:
        flash(f"Error: {e}", "error")
        students = []
        all_evaluated = False
        exam_finalized = True
        exam = {}

    return render_template(
        'evaluate_exam_list.html',
        exam=exam,
        students=students,
        exam_id=exam_id,
        all_evaluated=all_evaluated,
        exam_finalized=exam_finalized,
        name=session.get('email')
    )
    
    
@app.route('/finalize_exam/<int:exam_id>', methods=['POST'])
def finalize_exam(exam_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        return redirect('/login')

    teacher_id = session['user_id']

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Double-check all are evaluated
        cursor.execute("""
            SELECT COUNT(*) AS pending
            FROM student_answers sa
            JOIN questions q ON sa.question_id = q.question_id
            LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
            WHERE q.exam_id = %s AND ev.answer_id IS NULL
        """, (exam_id,))
        pending = cursor.fetchone()['pending']


        if pending > 0:
            flash("Cannot finalize: Some answers are still pending evaluation.", "error")
        else:
            # Get all students and their scores for this exam
            cursor.execute("""
                    SELECT
    st.student_id,
    st.roll_no,
    st.enrollment_no,
    st.full_name,

    COUNT(sa.answer_id) AS answers_given,
    COUNT(ev.answer_id) AS answers_evaluated,
    COALESCE(SUM(ev.score), 0) AS total_score,
    r.final_score

FROM questions q
JOIN student_answers sa
    ON sa.question_id = q.question_id
JOIN students st
    ON st.student_id = sa.student_id
LEFT JOIN evaluations ev
    ON ev.answer_id = sa.answer_id
LEFT JOIN result r
    ON r.student_id = st.student_id
   AND r.exam_id = %s

WHERE q.exam_id = %s

GROUP BY st.student_id
ORDER BY st.roll_no;

                """, (exam_id, exam_id))

            student_scores = cursor.fetchall()
            
            # Insert scores into result table
            for row in student_scores:
                student_id = row['student_id']
                total_score = row['total_score']
                # Round to 2 decimals
                score = round(float(total_score), 2) if total_score else 0.0
                
                # Calculate final score (ceil/floor)
                fractional = score - int(score)
                if fractional >= 0.5:
                    final_score = int(score) + 1  # ceil
                else:
                    final_score = int(score)  # floor
                
                # Insert or update result
                cursor.execute("""
                    INSERT INTO result (student_id, exam_id, score, final_score)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE score = %s, final_score = %s
                """, (student_id, exam_id, score, final_score, score, final_score))
            
            # Update exam evaluation status
            cursor.execute("""
                UPDATE exams SET evaluation_status = 1 WHERE exam_id = %s AND teacher_id = %s
            """, (exam_id, teacher_id))
            
            conn.commit()
            flash("Evaluation finalized successfully! Scores saved to results.", "success")

        conn.close()
    except Exception as e:
        flash(f"Error finalizing: {e}", "error")

    return redirect(url_for('evaluate_exam_list', exam_id=exam_id))

    
    
@app.route('/evaluate_student/<int:exam_id>/<int:student_id>', methods=['GET', 'POST'])
def evaluate_student(exam_id, student_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        return redirect('/login')

    teacher_id = session['user_id']
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        breakdowns = session.get("keyword_breakdowns", {})

        for key in request.form:
            if key.startswith('score_'):
                q_index = key.split("_")[1]
                answer_id = request.form.get(f'answer_id_{q_index}')
                score = request.form.get(key, '').strip()

                try:
                    score = float(score) if score else 0.0
                    score = round(score, 2)
                    if score < 0:
                        score = 0.0
                except:
                    score = 0.0

                keyword_breakdown_json = json.dumps(
                    breakdowns.get(str(answer_id), {})
                )

                cursor.execute("""
                    INSERT INTO evaluations (answer_id, score, keyword_breakdown)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE
                        score = VALUES(score),
                        keyword_breakdown = VALUES(keyword_breakdown),
                        evaluated_at = NOW()
                """, (answer_id, score, keyword_breakdown_json))

        conn.commit()
        session.pop("keyword_breakdowns", None)  # cleanup
        flash("All answers evaluated successfully!", "success")
        return redirect(url_for('evaluate_exam_list', exam_id=exam_id))

    # Fetch student
    cursor.execute("""
        SELECT full_name, roll_no, enrollment_no
        FROM students WHERE student_id = %s
    """, (student_id,))
    student = cursor.fetchone()
    
    # Fetch exam details
    cursor.execute("""
        SELECT exam_name, topic, max_marks
        FROM exams WHERE exam_id = %s
    """, (exam_id,))
    exam = cursor.fetchone()

    # 🔹 FETCH keyword_config also
    cursor.execute("""
        SELECT 
            sa.answer_id,
            q.question_id,
            q.question_text,
            q.model_answer,
            sa.answer_text,
            q.max_score,
            q.keyword_config,
            COALESCE(ev.score, 0) AS current_score
        FROM student_answers sa
        JOIN questions q ON sa.question_id = q.question_id
        LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
        WHERE sa.student_id = %s AND q.exam_id = %s
        ORDER BY q.question_id
    """, (student_id, exam_id))
    answers = cursor.fetchall()

    # 🔹 HYBRID AI EVALUATION (SEMANTIC + KEYWORD)
    total_ai_score = 0.0

    # initialize temporary storage for analysis breakdowns
    session["keyword_breakdowns"] = {}

    for ans in answers:
        # Load immutable keyword config for pre-extracted keywords
        keyword_config = json.loads(ans['keyword_config'])
        pre_extracted_kws = keyword_config.get("keywords", [])

        # Evaluate student answer using the new Hybrid Engine
        eval_result = evaluate_hybrid_answer(
            student_answer=ans['answer_text'],
            model_answer=ans['model_answer'],
            max_marks=ans['max_score'],
            question_id=ans['question_id'],
            pre_extracted_keywords=pre_extracted_kws
        )
        
        # store temporarily using answer_id as key for DB insertion later
        session["keyword_breakdowns"][str(ans["answer_id"])] = eval_result["analysis"]

        ai_score = eval_result["marks_awarded"]
        ans['ai_score'] = ai_score

        # Use teacher override if exists, else AI score
        ans['current_score'] = (
            round(float(ans['current_score']), 2)
            if ans['current_score']
            else ans['ai_score']
        )

        total_ai_score += ans['ai_score']

    total_ai_score = round(total_ai_score, 2)

    # Final score rounding logic (unchanged)
    fractional = total_ai_score - int(total_ai_score)
    if fractional >= 0.5:
        final_ai_score = int(total_ai_score) + 1
    else:
        final_ai_score = int(total_ai_score)

    cursor.close()
    conn.close()

    return render_template(
        'evaluate_student.html',
        exam_id=exam_id,
        exam=exam,
        student=student,
        answers=answers,
        total_ai_score=total_ai_score,
        final_ai_score=final_ai_score,
        name=session.get('email')
    )

        
    
# def generate_feedback(ai_score, max_score):
#     """
#     Generate intelligent feedback based on AI score
#     """
#     if ai_score is None or max_score == 0:
#         return "No evaluation possible."

#     percentage = (ai_score / max_score) * 100

#     if percentage >= 90:
#         return "Excellent answer! Fully matches the model answer with perfect understanding."
#     elif percentage >= 75:
#         return "Very good! Covers most key points. Minor details missing."
#     elif percentage >= 60:
#         return "Good attempt. Main concepts are there, but needs more clarity or examples."
#     elif percentage >= 40:
#         return "Partial understanding. Some points correct, but major gaps in explanation."
#     elif percentage >= 20:
#         return "Needs improvement. Answer is too brief or off-topic."
#     else:
#         return "Poor response. Does not address the question properly."      
     
    
# from nltk.tokenize import word_tokenize
# from nltk.corpus import stopwords
# import nltk

# # Download once
# nltk.download('punkt', quiet=True)
# nltk.download('punkt_tab', quiet=True)
# nltk.download('stopwords', quiet=True)

# def calculate_ai_score(student_answer, model_answer, max_score):
#     if not student_answer or not model_answer:
#         return 0
#     stop_words = set(stopwords.words('english'))
#     s_tokens = {w.lower() for w in word_tokenize(student_answer) if w.isalnum() and w.lower() not in stop_words}
#     m_tokens = {w.lower() for w in word_tokenize(model_answer) if w.isalnum() and w.lower() not in stop_words}
#     if not m_tokens:
#         return 0
#     overlap = len(s_tokens & m_tokens)
#     precision = overlap / len(s_tokens) if s_tokens else 0
#     recall = overlap / len(m_tokens)
#     f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
#     return round(f1 * max_score, 1)

# @app.route('/evaluate_exam/<int:exam_id>', methods=['GET', 'POST'])
# def evaluate_exam(exam_id):
#     if 'user_id' not in session or session.get('role') != 'teacher':
#         return redirect('/login')

#     teacher_id = session['user_id']

#     conn = mysql.connector.connect(**db_config)
#     cursor = conn.cursor(dictionary=True)

#     # Verify exam belongs to teacher
#     cursor.execute("SELECT exam_name FROM exams WHERE exam_id = %s AND teacher_id = %s", (exam_id, teacher_id))
#     exam = cursor.fetchone()
#     if not exam:
#         flash("Exam not found.", "error")
#         return redirect('/view_answers')

#     if request.method == 'POST':
#         answer_id = request.form.get('answer_id')
#         score = request.form.get('score')
#         feedback = request.form.get('feedback', '').strip()

#         try:
#             score = float(score) if score else None
#         except:
#             score = None

#         cursor.execute("""
#             INSERT INTO evaluations (answer_id, score, feedback)
#             VALUES (%s, %s, %s)
#             ON DUPLICATE KEY UPDATE score = %s, feedback = %s, evaluated_at = NOW()
#         """, (answer_id, score, feedback, score, feedback))
#         conn.commit()
#         flash("Answer saved!", "success")

#     # Fetch all unevaluated or partially evaluated answers
#     cursor.execute("""
#         SELECT 
#             sa.answer_id, sa.student_id, sa.answer_text, sa.submitted_at,
#             s.full_name, s.roll_no,
#             q.question_id, q.question_text, q.model_answer, q.max_score,
#             COALESCE(e.score, 0) AS current_score,
#             e.feedback
#         FROM student_answers sa
#         JOIN students s ON sa.student_id = s.student_id
#         JOIN questions q ON sa.question_id = q.question_id
#         LEFT JOIN evaluations e ON sa.answer_id = e.answer_id
#         WHERE q.exam_id = %s
#         ORDER BY s.roll_no, q.question_id
#     """, (exam_id,))
#     answers = cursor.fetchall()

#     # Add AI score
#     for ans in answers:
#         ans['ai_score'] = calculate_ai_score(ans['answer_text'], ans['model_answer'], ans['max_score'])

#     cursor.close()
#     conn.close()

#     return render_template(
#         'evaluate_exam.html',
#         exam=exam,
#         answers=answers,
#         name=session.get('email')
#     )
    
# @app.route('/finalize_evaluation/<int:exam_id>', methods=['POST'])
# def finalize_evaluation(exam_id):
#     if 'user_id' not in session or session.get('role') != 'teacher':
#         return redirect('/login')

#     teacher_id = session['user_id']

#     try:
#         conn = mysql.connector.connect(**db_config)
#         cursor = conn.cursor()

#         # Verify exam + all answers evaluated
#         cursor.execute("""
#             SELECT COUNT(sa.answer_id) AS total, COUNT(ev.answer_id) AS evaluated
#             FROM exams e
#             LEFT JOIN questions q ON e.exam_id = q.exam_id
#             LEFT JOIN student_answers sa ON q.question_id = sa.question_id
#             LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
#             WHERE e.exam_id = %s AND e.teacher_id = %s
#         """, (exam_id, teacher_id))
#         result = cursor.fetchone()

#         total, evaluated = result

#         if total and total == evaluated:
#             cursor.execute("""
#                 UPDATE exams SET evaluation_status = 1 WHERE exam_id = %s
#             """, (exam_id,))
#             conn.commit()
#             flash("Evaluation finalized! Exam is now fully evaluated.", "success")
#         else:
#             flash("Not all answers are evaluated yet.", "error")

#         cursor.close()
#         conn.close()
#     except mysql.connector.Error as err:
#         flash(f"Error: {err}", "error")

#     return redirect(url_for('view_answers'))


@app.route('/view_results')
def view_results():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')
    
    teacher_id = session['user_id']
    
    # Get filters
    course_id = request.args.get('course_id', type=int)
    semester_id = request.args.get('semester_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    selected_exam_id = request.args.get('exam_id', type=int)

    today = datetime.today().strftime('%Y-%m-%d')

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Fetch all courses
        cursor.execute("SELECT course_id, course_name FROM courses ORDER BY course_name")
        courses = cursor.fetchall()
        
        # Fetch semesters (filtered by course if selected)
        if course_id:
            cursor.execute("""
                SELECT DISTINCT s.semester_id, s.semester_name 
                FROM semesters s 
                WHERE s.course_id = %s
                ORDER BY s.semester_name
            """, (course_id,))
        else:
            cursor.execute("""
                SELECT DISTINCT s.semester_id, s.semester_name 
                FROM semesters s 
                JOIN exams e ON s.semester_id = e.semester_id 
                WHERE e.teacher_id = %s AND e.evaluation_status = 1
                ORDER BY s.semester_name
            """, (teacher_id,))
        semesters = cursor.fetchall()
        
        # Fetch subjects (filtered by course and semester if selected)
        subject_query = """
            SELECT DISTINCT sub.subject_id, sub.subject_name 
            FROM subjects sub 
            JOIN exams e ON sub.subject_id = e.subject_id 
            WHERE e.teacher_id = %s AND e.evaluation_status = 1
        """
        subject_params = [teacher_id]
        
        if course_id:
            subject_query += " AND sub.course_id = %s"
            subject_params.append(course_id)
        if semester_id:
            subject_query += " AND sub.semester_id = %s"
            subject_params.append(semester_id)
            
        subject_query += " ORDER BY sub.subject_name"
        cursor.execute(subject_query, subject_params)
        subjects = cursor.fetchall()
        
        # Fetch exams with filters
        exam_query = """
            SELECT DISTINCT e.exam_id, e.exam_name, e.topic, e.exam_date, 
                   e.start_time, e.end_time, e.max_marks, e.min_marks,
                   c.course_name, s.semester_name, sub.subject_name
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.teacher_id = %s AND e.evaluation_status = 1
        """
        exam_params = [teacher_id]

        if course_id:
            exam_query += " AND e.course_id = %s"
            exam_params.append(course_id)
        if semester_id:
            exam_query += " AND e.semester_id = %s"
            exam_params.append(semester_id)
        if subject_id:
            exam_query += " AND e.subject_id = %s"
            exam_params.append(subject_id)
        if date_from:
            exam_query += " AND e.exam_date >= %s"
            exam_params.append(date_from)
        if date_to:
            exam_query += " AND e.exam_date <= %s"
            exam_params.append(date_to)

        exam_query += " ORDER BY e.exam_date DESC, e.start_time DESC"
        cursor.execute(exam_query, exam_params)
        exams = cursor.fetchall()

        # Format exam display labels
        for exam in exams:
            start = td_to_time(exam['start_time']) if exam['start_time'] else None
            end = td_to_time(exam['end_time']) if exam['end_time'] else None
            time_str = f"{start.strftime('%I:%M %p')}-{end.strftime('%I:%M %p')}" if start and end else "N/A"
            exam['display_label'] = f"{exam['exam_name']} - {exam['subject_name']} ({exam['exam_date'].strftime('%b %d, %Y')}, {time_str})"

        results = []
        exam_stats = None

        if selected_exam_id:
            # Verify exam belongs to teacher and is finalized
            cursor.execute("""
                SELECT e.exam_name, e.topic, e.max_marks, e.min_marks, e.exam_date,
                       e.start_time, e.end_time,
                       c.course_name, s.semester_name, sub.subject_name
                FROM exams e
                JOIN courses c ON e.course_id = c.course_id
                JOIN semesters s ON e.semester_id = s.semester_id
                JOIN subjects sub ON e.subject_id = sub.subject_id
                WHERE e.exam_id = %s AND e.teacher_id = %s AND e.evaluation_status = 1
            """, (selected_exam_id, teacher_id))
            exam_stats = cursor.fetchone()

            if exam_stats:
                # Format exam times
                start = td_to_time(exam_stats['start_time']) if exam_stats['start_time'] else None
                end = td_to_time(exam_stats['end_time']) if exam_stats['end_time'] else None
                exam_stats['formatted_time'] = f"{start.strftime('%I:%M %p')} - {end.strftime('%I:%M %p')}" if start and end else "N/A"
                exam_stats['formatted_date'] = exam_stats['exam_date'].strftime('%B %d, %Y')
                
                # Fetch results from result table
                cursor.execute("""
                    SELECT 
                        s.student_id, s.full_name AS student_name, s.roll_no, 
                        s.enrollment_no,
                        r.score AS total_score,
                        r.final_score,
                        ex.max_marks, ex.min_marks,
                        CASE WHEN r.final_score >= ex.min_marks THEN 'Pass' ELSE 'Fail' END AS result_status
                    FROM result r
                    JOIN students s ON r.student_id = s.student_id
                    JOIN exams ex ON r.exam_id = ex.exam_id
                    WHERE r.exam_id = %s
                    ORDER BY r.final_score DESC, r.score DESC
                """, (selected_exam_id,))
                results = cursor.fetchall()

                # Process results
                for r in results:
                    r['total_score'] = round(float(r['total_score'] or 0), 2)
                    r['final_score'] = int(r['final_score'] or 0)
                    r['max_marks'] = int(r['max_marks'] or 0)
                    r['min_marks'] = int(r['min_marks'] or 0)
                    r['percentage'] = round((r['total_score'] / r['max_marks'] * 100), 2) if r['max_marks'] > 0 else 0

                # Calculate statistics
                if results:
                    total_students = len(results)
                    passed_students = sum(1 for r in results if r['result_status'] == 'Pass')
                    failed_students = total_students - passed_students
                    
                    scores = [r['total_score'] for r in results]
                    final_scores = [r['final_score'] for r in results]
                    
                    exam_stats.update({
                        'total_students': total_students,
                        'passed_students': passed_students,
                        'failed_students': failed_students,
                        'pass_percentage': round((passed_students / total_students * 100), 2) if total_students > 0 else 0,
                        'avg_score': round(sum(scores) / total_students, 2),
                        'avg_final': round(sum(final_scores) / total_students, 2),
                        'highest_score': max(scores),
                        'highest_final': max(final_scores),
                        'lowest_score': min(scores),
                        'lowest_final': min(final_scores)
                    })
                else:
                    # No results yet
                    exam_stats.update({
                        'total_students': 0,
                        'passed_students': 0,
                        'failed_students': 0,
                        'pass_percentage': 0,
                        'avg_score': 0,
                        'avg_final': 0,
                        'highest_score': 0,
                        'highest_final': 0,
                        'lowest_score': 0,
                        'lowest_final': 0
                    })

        cursor.close()
        conn.close()
        
    except Exception as e:
        flash(f"Database error: {e}", "error")
        courses = semesters = subjects = []
        exams = []
        results = []
        exam_stats = None

    return render_template('view_results.html',
                           exams=exams, 
                           results=results, 
                           exam_stats=exam_stats,
                           courses=courses, 
                           semesters=semesters, 
                           subjects=subjects,
                           selected_exam_id=selected_exam_id,
                           selected_course=course_id, 
                           selected_semester=semester_id, 
                           selected_subject=subject_id,
                           date_from=date_from, 
                           date_to=date_to,
                           today=today,
                           name=session.get('email'))


@app.route('/export_results_excel/<int:exam_id>')
def export_results_excel(exam_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')

    teacher_id = session['user_id']

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Verify exam belongs to teacher
        cursor.execute("""
            SELECT e.exam_name, e.topic, c.course_name, s.semester_name, sub.subject_name,
                   e.max_marks, e.min_marks, e.exam_date
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.exam_id = %s AND e.teacher_id = %s AND e.evaluation_status = 1
        """, (exam_id, teacher_id))
        exam_info = cursor.fetchone()

        if not exam_info:
            flash("Exam not found or not finalized.", "error")
            return redirect('/view_results')

        # Fetch results
        cursor.execute("""
            SELECT s.roll_no, s.enrollment_no, s.full_name, r.score, r.final_score,
                   CASE WHEN r.final_score >= e.min_marks THEN 'Pass' ELSE 'Fail' END AS status
            FROM result r
            JOIN students s ON r.student_id = s.student_id
            JOIN exams e ON r.exam_id = e.exam_id
            WHERE r.exam_id = %s
            ORDER BY r.final_score DESC, r.score DESC
        """, (exam_id,))
        results = cursor.fetchall()

        cursor.close()
        conn.close()

        # Create Excel
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()

        # Formats
        title_format = workbook.add_format({
            'bold': True, 'font_size': 18, 'align': 'center', 
            'valign': 'vcenter', 'font_color': '#4834d4'
        })
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4834d4', 'font_color': 'white', 
            'border': 1, 'align': 'center', 'valign': 'vcenter'
        })
        info_label_format = workbook.add_format({
            'bold': True, 'bg_color': '#e8e8e8', 'border': 1
        })
        info_value_format = workbook.add_format({'border': 1})
        cell_format = workbook.add_format({'border': 1, 'align': 'center'})
        pass_format = workbook.add_format({
            'border': 1, 'align': 'center', 'bg_color': '#d4edda', 
            'font_color': '#155724', 'bold': True
        })
        fail_format = workbook.add_format({
            'border': 1, 'align': 'center', 'bg_color': '#f8d7da', 
            'font_color': '#721c24', 'bold': True
        })

        # Title
        worksheet.merge_range('A1:H1', f"{exam_info['exam_name']} - Results Report", title_format)
        worksheet.set_row(0, 25)
        
        # Exam Details
        row = 2
        worksheet.write(row, 0, 'Course:', info_label_format)
        worksheet.write(row, 1, exam_info['course_name'], info_value_format)
        worksheet.write(row, 2, 'Semester:', info_label_format)
        worksheet.write(row, 3, exam_info['semester_name'], info_value_format)
        
        row += 1
        worksheet.write(row, 0, 'Subject:', info_label_format)
        worksheet.write(row, 1, exam_info['subject_name'], info_value_format)
        worksheet.write(row, 2, 'Topic:', info_label_format)
        worksheet.write(row, 3, exam_info['topic'] or 'N/A', info_value_format)
        
        row += 1
        worksheet.write(row, 0, 'Exam Date:', info_label_format)
        worksheet.write(row, 1, exam_info['exam_date'].strftime('%B %d, %Y'), info_value_format)
        worksheet.write(row, 2, 'Max Marks:', info_label_format)
        worksheet.write(row, 3, exam_info['max_marks'], info_value_format)
        
        row += 1
        worksheet.write(row, 0, 'Min Marks:', info_label_format)
        worksheet.write(row, 1, exam_info['min_marks'], info_value_format)
        worksheet.write(row, 2, 'Total Students:', info_label_format)
        worksheet.write(row, 3, len(results), info_value_format)

        # Column widths
        worksheet.set_column('A:A', 8)   # Rank
        worksheet.set_column('B:B', 12)  # Roll No
        worksheet.set_column('C:C', 18)  # Enrollment
        worksheet.set_column('D:D', 25)  # Name
        worksheet.set_column('E:E', 12)  # Score
        worksheet.set_column('F:F', 12)  # Final
        worksheet.set_column('G:G', 12)  # Percentage
        worksheet.set_column('H:H', 12)  # Status

        # Table Headers
        row += 2
        headers = ['Rank', 'Roll No', 'Enrollment No', 'Student Name', 
                   'Score', 'Final Score', 'Percentage', 'Status']
        worksheet.write_row(row, 0, headers, header_format)
        worksheet.set_row(row, 20)

        # Data rows
        row += 1
        for idx, r in enumerate(results, 1):
            score = round(float(r['score'] or 0), 2)
            final_score = int(r['final_score'] or 0)
            percentage = round((score / exam_info['max_marks'] * 100), 2) if exam_info['max_marks'] > 0 else 0
            
            worksheet.write(row, 0, idx, cell_format)
            worksheet.write(row, 1, r['roll_no'], cell_format)
            worksheet.write(row, 2, r['enrollment_no'], cell_format)
            worksheet.write(row, 3, r['full_name'], cell_format)
            worksheet.write(row, 4, score, cell_format)
            worksheet.write(row, 5, final_score, cell_format)
            worksheet.write(row, 6, f"{percentage}%", cell_format)
            
            status_format = pass_format if r['status'] == 'Pass' else fail_format
            worksheet.write(row, 7, r['status'], status_format)
            row += 1

        workbook.close()
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{exam_info['exam_name']}_Results.xlsx"
        )
        
    except Exception as e:
        flash(f"Export failed: {e}", "error")
        return redirect('/view_results')


@app.route('/export_results_pdf/<int:exam_id>')
def export_results_pdf(exam_id):
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')

    teacher_id = session['user_id']

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT e.exam_name, e.topic, c.course_name, s.semester_name, sub.subject_name,
                   e.max_marks, e.min_marks, e.exam_date
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.exam_id = %s AND e.teacher_id = %s AND e.evaluation_status = 1
        """, (exam_id, teacher_id))
        exam_info = cursor.fetchone()

        if not exam_info:
            flash("Exam not found or not finalized.", "error")
            return redirect('/view_results')

        cursor.execute("""
            SELECT s.roll_no, s.enrollment_no, s.full_name, r.score, r.final_score,
                   CASE WHEN r.final_score >= e.min_marks THEN 'Pass' ELSE 'Fail' END AS status
            FROM result r
            JOIN students s ON r.student_id = s.student_id
            JOIN exams e ON r.exam_id = e.exam_id
            WHERE r.exam_id = %s
            ORDER BY r.final_score DESC, r.score DESC
        """, (exam_id,))
        results = cursor.fetchall()

        cursor.close()
        conn.close()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40,
                                leftMargin=30, rightMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#4834d4'),
            spaceAfter=20,
            alignment=1
        )
        elements.append(Paragraph(f"{exam_info['exam_name']}", title_style))
        elements.append(Paragraph("Results Report", styles['Heading2']))
        elements.append(Spacer(1, 15))

        # Exam Info
        info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], fontSize=10, spaceAfter=6)
        elements.append(Paragraph(f"<b>Course:</b> {exam_info['course_name']} | <b>Semester:</b> {exam_info['semester_name']}", info_style))
        elements.append(Paragraph(f"<b>Subject:</b> {exam_info['subject_name']} | <b>Topic:</b> {exam_info['topic'] or 'N/A'}", info_style))
        elements.append(Paragraph(f"<b>Date:</b> {exam_info['exam_date'].strftime('%B %d, %Y')} | <b>Max Marks:</b> {exam_info['max_marks']} | <b>Min Marks:</b> {exam_info['min_marks']}", info_style))
        elements.append(Spacer(1, 20))

        # Table
        data = [['Rank', 'Roll', 'Enrollment', 'Name', 'Score', 'Final', '%', 'Status']]
        
        for idx, r in enumerate(results, 1):
            score = round(float(r['score'] or 0), 2)
            final_score = int(r['final_score'] or 0)
            percentage = round((score / exam_info['max_marks'] * 100), 2) if exam_info['max_marks'] > 0 else 0
            
            data.append([
                str(idx),
                str(r['roll_no']), 
                str(r['enrollment_no']),
                r['full_name'][:18],
                f"{score:.2f}",
                str(final_score),
                f"{percentage:.1f}%",
                r['status']
            ])

        table = Table(data, colWidths=[35, 40, 75, 110, 50, 45, 50, 45])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4834d4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 10),
            ('TOPPADDING', (0,0), (-1,0), 10),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        
        elements.append(table)
        
        # Add footer
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], 
                                      fontSize=8, textColor=colors.grey, alignment=1)
        elements.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", footer_style))

        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{exam_info['exam_name']}_Results.pdf"
        )
        
    except Exception as e:
        flash(f"PDF Export failed: {e}", "error")
        return redirect('/view_results')

# Fixed Teacher Analytics Route - Replace in your app.py

@app.route('/teacher/analytics')
def teacher_analytics():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')
    
    teacher_id = session['user_id']
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Get teacher info
        cursor.execute("""
            SELECT full_name, expertise, subjects_taught
            FROM teachers
            WHERE teacher_id = %s
        """, (teacher_id,))
        teacher_info = cursor.fetchone()
        
        # Overall Statistics
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT e.exam_id) as total_exams,
                COUNT(DISTINCT e.subject_id) as total_subjects,
                COUNT(DISTINCT q.question_id) as total_questions
            FROM exams e
            LEFT JOIN questions q ON e.exam_id = q.exam_id
            WHERE e.teacher_id = %s
        """, (teacher_id,))
        overall_stats = cursor.fetchone()
        
        # Count total unique students who attempted exams
        cursor.execute("""
            SELECT COUNT(DISTINCT sa.student_id) as total_students
            FROM exams e
            JOIN questions q ON e.exam_id = q.exam_id
            JOIN student_answers sa ON q.question_id = sa.question_id
            WHERE e.teacher_id = %s
        """, (teacher_id,))
        student_count = cursor.fetchone()
        overall_stats['total_students'] = student_count['total_students'] or 0
        
        # Exam Status Distribution
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN CURDATE() < exam_date THEN 'upcoming'
                    WHEN CURDATE() = exam_date AND CURTIME() BETWEEN start_time AND end_time THEN 'ongoing'
                    ELSE 'completed'
                END as status,
                COUNT(*) as count
            FROM exams
            WHERE teacher_id = %s
            GROUP BY status
        """, (teacher_id,))
        exam_status = cursor.fetchall()
        status_dict = {item['status']: item['count'] for item in exam_status}
        
        # Subject-wise Exam Distribution
        cursor.execute("""
            SELECT 
                sub.subject_name,
                COUNT(DISTINCT e.exam_id) as exam_count,
                AVG(e.max_marks) as avg_max_marks
            FROM exams e
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.teacher_id = %s
            GROUP BY sub.subject_id, sub.subject_name
            ORDER BY exam_count DESC
        """, (teacher_id,))
        subject_distribution = cursor.fetchall()
        
        # Add student count per subject
        for subject in subject_distribution:
            cursor.execute("""
                SELECT COUNT(DISTINCT sa.student_id) as student_count
                FROM exams e
                JOIN subjects sub ON e.subject_id = sub.subject_id
                JOIN questions q ON e.exam_id = q.exam_id
                JOIN student_answers sa ON q.question_id = sa.question_id
                WHERE e.teacher_id = %s AND sub.subject_name = %s
            """, (teacher_id, subject['subject_name']))
            count_result = cursor.fetchone()
            subject['student_count'] = count_result['student_count'] or 0
        
        # Student Performance Analysis - LIMITED TO TOP 8 EXAMS
        cursor.execute("""
            SELECT 
                e.exam_name,
                e.exam_id,
                e.exam_date,
                e.max_marks,
                e.min_marks,
                COUNT(DISTINCT student_scores.student_id) as total_students,
                AVG(student_scores.exam_score / e.max_marks * 100) as avg_percentage,
                SUM(CASE WHEN student_scores.exam_score >= e.min_marks THEN 1 ELSE 0 END) as passed_students
            FROM exams e
            LEFT JOIN (
                SELECT 
                    sa.student_id,
                    q.exam_id,
                    SUM(COALESCE(ev.score, 0)) as exam_score
                FROM student_answers sa
                JOIN questions q ON sa.question_id = q.question_id
                LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
                GROUP BY sa.student_id, q.exam_id
            ) as student_scores ON e.exam_id = student_scores.exam_id
            WHERE e.teacher_id = %s
            GROUP BY e.exam_id, e.exam_name, e.exam_date, e.max_marks, e.min_marks
            HAVING total_students > 0
            ORDER BY e.exam_date DESC
            LIMIT 8
        """, (teacher_id,))
        exam_performance = cursor.fetchall()
        
        # Calculate pass rates
        for exam in exam_performance:
            if exam['total_students'] and exam['total_students'] > 0:
                exam['pass_rate'] = (exam['passed_students'] / exam['total_students']) * 100
            else:
                exam['pass_rate'] = 0
                exam['avg_percentage'] = 0
        
        # Monthly Exam Trends - Last 12 months (FIXED)
        # cursor.execute("""
        #     SELECT 
        #         DATE_FORMAT(exam_date, '%%Y-%%m') AS month,
        #         DATE_FORMAT(exam_date, '%%b %%Y') AS month_name,
        #         COUNT(*) AS exam_count,
        #         AVG(max_marks) AS avg_marks
        #     FROM exams
        #     WHERE teacher_id = %s
        #     AND exam_date >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
        #     GROUP BY YEAR(exam_date), MONTH(exam_date)
        #     ORDER BY month ASC
        # """, (teacher_id,))
        # monthly_trends = cursor.fetchall()
    
        
        # Question Difficulty Analysis (based on average scores)
        cursor.execute("""
            SELECT 
                q.question_id,
                LEFT(q.question_text, 50) as question_preview,
                q.max_score,
                COUNT(DISTINCT sa.answer_id) as attempt_count,
                AVG(COALESCE(ev.score, 0)) as avg_score,
                (AVG(COALESCE(ev.score, 0)) / q.max_score) * 100 as avg_percentage
            FROM questions q
            JOIN exams e ON q.exam_id = e.exam_id
            LEFT JOIN student_answers sa ON q.question_id = sa.question_id
            LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
            WHERE e.teacher_id = %s
            GROUP BY q.question_id, q.question_text, q.max_score
            HAVING attempt_count > 0
            ORDER BY avg_percentage ASC
            LIMIT 10
        """, (teacher_id,))
        difficult_questions = cursor.fetchall()
        
        # Top Performing Students (across all exams)
        cursor.execute("""
            SELECT 
                st.full_name,
                st.roll_no,
                student_data.student_id,
                COUNT(DISTINCT student_data.exam_id) as exams_taken,
                AVG((student_data.exam_score / student_data.exam_max) * 100) as avg_percentage,
                SUM(CASE WHEN student_data.exam_score >= student_data.exam_min THEN 1 ELSE 0 END) as exams_passed
            FROM (
                SELECT 
                    sa.student_id,
                    q.exam_id,
                    SUM(COALESCE(ev.score, 0)) as exam_score,
                    MAX(e.max_marks) as exam_max,
                    MAX(e.min_marks) as exam_min
                FROM student_answers sa
                JOIN questions q ON sa.question_id = q.question_id
                JOIN exams e ON q.exam_id = e.exam_id
                LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
                WHERE e.teacher_id = %s
                GROUP BY sa.student_id, q.exam_id
            ) as student_data
            JOIN students st ON student_data.student_id = st.student_id
            GROUP BY st.student_id, st.full_name, st.roll_no, student_data.student_id
            HAVING exams_taken > 0
            ORDER BY avg_percentage DESC
            LIMIT 10
        """, (teacher_id,))
        top_students = cursor.fetchall()
        
        # Course-wise Performance
        cursor.execute("""
            SELECT 
                c.course_name,
                sem.semester_name,
                COUNT(DISTINCT e.exam_id) as exam_count,
                AVG(e.max_marks) as avg_max_marks
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters sem ON e.semester_id = sem.semester_id
            WHERE e.teacher_id = %s
            GROUP BY c.course_id, c.course_name, sem.semester_id, sem.semester_name
            ORDER BY exam_count DESC
        """, (teacher_id,))
        course_performance = cursor.fetchall()
        
        # Recent Activity
        cursor.execute("""
            SELECT 
                e.exam_name,
                e.exam_date,
                e.start_time,
                sub.subject_name,
                COUNT(DISTINCT sa.student_id) as submission_count
            FROM exams e
            JOIN subjects sub ON e.subject_id = sub.subject_id
            LEFT JOIN questions q ON e.exam_id = q.exam_id
            LEFT JOIN student_answers sa ON q.question_id = sa.question_id
            WHERE e.teacher_id = %s
            GROUP BY e.exam_id, e.exam_name, e.exam_date, e.start_time, sub.subject_name
            ORDER BY e.exam_date DESC
            LIMIT 5
        """, (teacher_id,))
        recent_activity = cursor.fetchall()
        
        # Convert timedelta to time string for display
        for activity in recent_activity:
            if activity['start_time']:
                if isinstance(activity['start_time'], timedelta):
                    total_seconds = int(activity['start_time'].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    activity['start_time_str'] = f"{hours:02d}:{minutes:02d}"
                else:
                    activity['start_time_str'] = activity['start_time'].strftime('%I:%M %p')
            else:
                activity['start_time_str'] = 'N/A'
        
        # NEW: Model Answer vs Student Answer Comparison (Top 10 Questions)
        # cursor.execute("""
        #     SELECT 
        #         q.question_id,
        #         LEFT(q.question_text, 40) as question_preview,
        #         q.max_score as model_max_score,
        #         AVG(COALESCE(ev.score, 0)) as avg_student_score,
        #         q.max_score - AVG(COALESCE(ev.score, 0)) as score_gap,
        #         COUNT(DISTINCT sa.answer_id) as total_attempts
        #     FROM questions q
        #     JOIN exams e ON q.exam_id = e.exam_id
        #     LEFT JOIN student_answers sa ON q.question_id = sa.question_id
        #     LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
        #     WHERE e.teacher_id = %s
        #     GROUP BY q.question_id, q.question_text, q.max_score
        #     HAVING total_attempts > 0
        #     ORDER BY total_attempts DESC, score_gap DESC
        #     LIMIT 10
        # """, (teacher_id,))
        # answer_comparison = cursor.fetchall()

        # NEW: Exam-wise Model Max Marks vs Average Student Score (Top 10 Exams by attempts)
        cursor.execute("""
            SELECT 
                e.exam_id,
                e.exam_name,
                e.max_marks AS exam_max_marks,
                COALESCE(AVG(r.score), 0) AS avg_student_score,
                e.max_marks - COALESCE(AVG(r.score), 0) AS score_gap,
                COUNT(DISTINCT r.student_id) AS total_students
            FROM exams e
            LEFT JOIN result r ON e.exam_id = r.exam_id
            WHERE e.teacher_id = %s
            AND e.evaluation_status = 1  -- Only evaluated exams
            GROUP BY e.exam_id, e.exam_name, e.max_marks
            HAVING total_students > 0
            ORDER BY total_students DESC, score_gap DESC
            LIMIT 10
        """, (teacher_id,))
        answer_comparison = cursor.fetchall()
        
        # Prepare analytics data
        analytics = {
            'teacher_info': teacher_info or {'full_name': session.get('email'), 'expertise': 'N/A', 'subjects_taught': 'N/A'},
            'total_exams': overall_stats['total_exams'] or 0,
            'total_subjects': overall_stats['total_subjects'] or 0,
            'total_questions': overall_stats['total_questions'] or 0,
            'total_students': overall_stats['total_students'] or 0,
            'exam_status': status_dict,
            'subject_distribution': subject_distribution,
            'exam_performance': exam_performance,
            # 'monthly_trends': monthly_trends,
            'difficult_questions': difficult_questions,
            'top_students': top_students,
            'course_performance': course_performance,
            'recent_activity': recent_activity,
            'answer_comparison': answer_comparison
        }
        
        cursor.close()
        conn.close()
        
        return render_template('teacher_analytics.html', 
                             analytics=analytics, 
                             name=session.get('email'))
        
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/teacher_dashboard')
    
    


@app.route('/student_dashboard')
def student_dashboard():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        # Fetch student info
        cursor.execute("""
            SELECT full_name, roll_no, enrollment_no, course_id, semester_id, contact, department
            FROM students WHERE student_id = %s
        """, (session['user_id'],))
        student = cursor.fetchone()
        if not student:
            flash("Student information not found.", "error")
            return redirect('/login')
        full_name, roll_no, enrollment_no, course_id, semester_id, contact, department = student
        
        # Fetch course name
        cursor.execute("SELECT course_name FROM courses WHERE course_id = %s", (course_id,))
        course_result = cursor.fetchone()
        course_name = course_result[0] if course_result else "N/A"
        
        # Fetch semester name
        cursor.execute("SELECT semester_name FROM semesters WHERE semester_id = %s", (semester_id,))
        semester_result = cursor.fetchone()
        semester_name = semester_result[0] if semester_result else "N/A"
        
        # Current date and time
        current_datetime = datetime.now()
        
        # Fetch all exams for the student's course and semester
        cursor.execute("""
            SELECT exam_id, exam_name, topic, exam_date, start_time, end_time, max_marks
            FROM exams 
            WHERE course_id = %s AND semester_id = %s
            ORDER BY exam_date ASC, start_time ASC
        """, (course_id, semester_id))
        exam_rows = cursor.fetchall()
        
        upcoming_exams = []
        ongoing_exams = []
        attempted_exams = []
        missed_exams = []
        
        num_upcoming = 0
        num_ongoing = 0
        
        for row in exam_rows:
            exam_id, exam_name, topic, exam_date, start_time_delta, end_time_delta, max_marks = row
            
            # Convert timedelta to time
            if isinstance(start_time_delta, timedelta):
                total_seconds = int(start_time_delta.total_seconds())
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                start_time = time(hours, minutes, seconds)
            else:
                start_time = start_time_delta
            
            if isinstance(end_time_delta, timedelta):
                total_seconds = int(end_time_delta.total_seconds())
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                end_time = time(hours, minutes, seconds)
            else:
                end_time = end_time_delta
            
            # Combine to create datetime objects
            start_dt = datetime.combine(exam_date, start_time)
            end_dt = datetime.combine(exam_date, end_time)
            
            # Determine status
            if current_datetime < start_dt:
                status = 'upcoming'
            elif start_dt <= current_datetime <= end_dt:
                status = 'ongoing'
            else:
                status = 'completed'

            
            # Check if submitted
            cursor.execute("""
                SELECT COUNT(*) FROM student_answers 
                WHERE student_id = %s AND question_id IN (
                    SELECT question_id FROM questions WHERE exam_id = %s
                )
            """, (session['user_id'], exam_id))
            submission_count = cursor.fetchone()[0]
            has_submitted = submission_count > 0
            
            # Correct counters based on final classification
            if status == 'upcoming':
                num_upcoming += 1
            elif status == 'ongoing' and not has_submitted:
                num_ongoing += 1

            
            exam_data = {
                'exam_id': exam_id,
                'exam_name': exam_name,
                'topic': topic,
                'exam_date': exam_date,
                'start_time': start_time,
                'end_time': end_time,
                'max_marks': max_marks,
                'status': status,
                'has_submitted': has_submitted,
                'start_datetime': start_dt,
                'end_datetime': end_dt
            }
            
            # Categorize exams
            # Categorize exams (FIXED LOGIC)
            if status == 'upcoming':
                upcoming_exams.append(exam_data)

            elif status == 'ongoing':
                if has_submitted:
                    attempted_exams.append(exam_data)
                else:
                    ongoing_exams.append(exam_data)

            elif status == 'completed':
                if has_submitted:
                    attempted_exams.append(exam_data)
                else:
                    missed_exams.append(exam_data)

        
        # Sort exams
        upcoming_exams.sort(key=lambda x: x['start_datetime'])
        ongoing_exams.sort(key=lambda x: x['start_datetime'])
        attempted_exams.sort(key=lambda x: x['end_datetime'], reverse=True)
        missed_exams.sort(key=lambda x: x['end_datetime'], reverse=True)
        
        # Limit to 5 for dashboard
        upcoming_exams = upcoming_exams[:5]
        ongoing_exams = ongoing_exams[:5]
        attempted_exams = attempted_exams[:5]
        missed_exams = missed_exams[:5]
        
        # Count attempted exams
        cursor.execute("""
            SELECT COUNT(DISTINCT q.exam_id) 
            FROM student_answers sa
            JOIN questions q ON sa.question_id = q.question_id
            WHERE sa.student_id = %s
        """, (session['user_id'],))
        num_attempted = cursor.fetchone()[0]
        
        num_missed = len([e for row in exam_rows for e in [row] 
                         if datetime.combine(row[3], row[5] if isinstance(row[5], time) else time(23,59)) < current_datetime
                         and not any(a['exam_id'] == row[0] for a in attempted_exams)])
        
        # Fetch recent submissions with scores
        cursor.execute("""
            SELECT ex.exam_id, ex.exam_name, ex.topic, ex.exam_date, 
                   MAX(sa.submitted_at) AS submission_date,
                   COALESCE(SUM(ev.score), 0) AS total_score,
                   ex.max_marks,
                   COUNT(DISTINCT sa.answer_id) AS questions_attempted
            FROM student_answers sa
            JOIN questions q ON sa.question_id = q.question_id
            JOIN exams ex ON q.exam_id = ex.exam_id
            LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
            WHERE sa.student_id = %s AND ex.course_id = %s AND ex.semester_id = %s
            GROUP BY ex.exam_id, ex.exam_name, ex.topic, ex.exam_date, ex.max_marks
            ORDER BY submission_date DESC
            LIMIT 5
        """, (session['user_id'], course_id, semester_id))
        recent_submissions = cursor.fetchall()
        
        # Calculate overall performance
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT q.exam_id) as total_exams_attempted,
                COALESCE(SUM(ev.score), 0) as total_marks_obtained,
                COALESCE(SUM(DISTINCT ex.max_marks), 0) as total_max_marks
            FROM student_answers sa
            JOIN questions q ON sa.question_id = q.question_id
            JOIN exams ex ON q.exam_id = ex.exam_id
            LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
            WHERE sa.student_id = %s
        """, (session['user_id'],))
        performance_result = cursor.fetchone()
        
        # Convert to float to avoid decimal issues
        performance = (
            int(performance_result[0]) if performance_result[0] else 0,
            float(performance_result[1]) if performance_result[1] else 0.0,
            float(performance_result[2]) if performance_result[2] else 0.0
        )
        
        cursor.close()
        conn.close()
        
        student_info = {
            'full_name': full_name if full_name and full_name != 'Not Provided' else 'Student',
            'roll_no': roll_no if roll_no else 'N/A',
            'enrollment_no': enrollment_no if enrollment_no else 'N/A',
            'course_name': course_name,
            'semester_name': semester_name,
            'contact': contact if contact and contact != 'Not Provided' else 'N/A',
            'department': department if department and department != 'Not Provided' else 'N/A'
        }
        
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        student_info = {
            'full_name': session.get('email', 'Student'),
            'roll_no': 'N/A',
            'enrollment_no': 'N/A',
            'course_name': 'N/A',
            'semester_name': 'N/A',
            'contact': 'N/A',
            'department': 'N/A'
        }
        upcoming_exams = []
        ongoing_exams = []
        attempted_exams = []
        missed_exams = []
        recent_submissions = []
        num_upcoming = 0
        num_ongoing = 0
        num_attempted = 0
        num_missed = 0
        performance = (0, 0.0, 0.0)
    except Exception as e:
        flash(f"An error occurred: {str(e)}", "error")
        student_info = {
            'full_name': session.get('email', 'Student'),
            'roll_no': 'N/A',
            'enrollment_no': 'N/A',
            'course_name': 'N/A',
            'semester_name': 'N/A',
            'contact': 'N/A',
            'department': 'N/A'
        }
        upcoming_exams = []
        ongoing_exams = []
        attempted_exams = []
        missed_exams = []
        recent_submissions = []
        num_upcoming = 0
        num_ongoing = 0
        num_attempted = 0
        num_missed = 0
        performance = (0, 0.0, 0.0)

    return render_template('student_dashboard.html', 
                         student_info=student_info,
                         upcoming_exams=upcoming_exams,
                         ongoing_exams=ongoing_exams,
                         attempted_exams=attempted_exams,
                         missed_exams=missed_exams,
                         recent_submissions=recent_submissions,
                         num_upcoming=num_upcoming,
                         num_ongoing=num_ongoing,
                         num_attempted=num_attempted,
                         num_missed=num_missed,
                         performance=performance)        



@app.route('/student/exams')
def student_exams():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')
    
    # Get filter parameters for each section
    upcoming_limit = int(request.args.get('upcoming_limit', 5))
    ongoing_limit = int(request.args.get('ongoing_limit', 5))
    attempted_limit = int(request.args.get('attempted_limit', 5))
    missed_limit = int(request.args.get('missed_limit', 5))
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        # Fetch student info
        cursor.execute("SELECT full_name, course_id, semester_id FROM students WHERE student_id = %s", (session['user_id'],))
        student = cursor.fetchone()
        if not student:
            flash("Student information not found.", "error")
            return redirect('/login')
        full_name, course_id, semester_id = student
        
        # Fetch exams for the student's course and semester
        cursor.execute("""
            SELECT exam_id, exam_name, topic, exam_date, start_time, end_time, max_marks
            FROM exams 
            WHERE course_id = %s AND semester_id = %s
            ORDER BY exam_date ASC
        """, (course_id, semester_id))
        exam_rows = cursor.fetchall()
        
        current_datetime = datetime.now()
        
        upcoming_exams = []
        ongoing_exams = []
        attempted_exams = []
        missed_exams = []
        
        for row in exam_rows:
            exam_id, exam_name, topic, exam_date, start_time_delta, end_time_delta, max_marks = row
            
            # Convert timedelta to time
            if isinstance(start_time_delta, timedelta):
                total_seconds = int(start_time_delta.total_seconds())
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                start_time = time(hours, minutes, seconds)
            else:
                start_time = start_time_delta
            
            if isinstance(end_time_delta, timedelta):
                total_seconds = int(end_time_delta.total_seconds())
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                end_time = time(hours, minutes, seconds)
            else:
                end_time = end_time_delta
            
            # Combine to create datetime objects
            start_dt = datetime.combine(exam_date, start_time)
            end_dt = datetime.combine(exam_date, end_time)
            
            # Determine exam status
            if current_datetime < start_dt:
                status = 'upcoming'
            elif start_dt <= current_datetime <= end_dt:
                status = 'ongoing'
            else:
                status = 'completed'
            
            # Check if the student has submitted answers for this exam
            cursor.execute("""
                SELECT COUNT(*) FROM student_answers 
                WHERE student_id = %s AND question_id IN (
                    SELECT question_id FROM questions WHERE exam_id = %s
                )
            """, (session['user_id'], exam_id))
            submission_count = cursor.fetchone()[0]
            has_submitted = submission_count > 0
            
            exam_data = {
                'exam_id': exam_id,
                'exam_name': exam_name,
                'topic': topic,
                'exam_date': exam_date,
                'start_time': start_time,
                'end_time': end_time,
                'max_marks': max_marks,
                'status': status,
                'has_submitted': has_submitted,
                'start_datetime': start_dt,
                'end_datetime': end_dt
            }
            
            # Categorize exams into 4 sections
            # Categorize exams into 4 sections
            if status == 'upcoming':
                upcoming_exams.append(exam_data)

            elif status == 'ongoing':
                if has_submitted:
                    # Ongoing but already submitted → treat as attempted
                    attempted_exams.append(exam_data)
                else:
                    ongoing_exams.append(exam_data)

            elif status == 'completed':
                if has_submitted:
                    attempted_exams.append(exam_data)
                else:
                    missed_exams.append(exam_data)

        
        # Sort upcoming exams: nearest exam first (ascending)
        upcoming_exams.sort(key=lambda x: x['start_datetime'])
        
        # Sort ongoing exams: current ones first (by start time, earliest first)
        ongoing_exams.sort(key=lambda x: x['start_datetime'])
        
        # Sort attempted exams: most recently completed first (descending)
        attempted_exams.sort(key=lambda x: x['end_datetime'], reverse=True)
        
        # Sort missed exams: most recently missed first (descending)
        missed_exams.sort(key=lambda x: x['end_datetime'], reverse=True)
        
        # Apply limits
        upcoming_exams = upcoming_exams[:upcoming_limit]
        ongoing_exams = ongoing_exams[:ongoing_limit]
        attempted_exams = attempted_exams[:attempted_limit]
        missed_exams = missed_exams[:missed_limit]
        
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        upcoming_exams = []
        ongoing_exams = []
        attempted_exams = []
        missed_exams = []
        full_name = session.get('email')  # Fallback

    return render_template('student_exams.html', 
                         upcoming_exams=upcoming_exams,
                         ongoing_exams=ongoing_exams,
                         attempted_exams=attempted_exams,
                         missed_exams=missed_exams,
                         name=full_name,
                         upcoming_limit=upcoming_limit,
                         ongoing_limit=ongoing_limit,
                         attempted_limit=attempted_limit,
                         missed_limit=missed_limit)
        
    

@app.route('/student/take_exam/<int:exam_id>', methods=['GET'])
def student_take_exam(exam_id):
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        # Fetch student info
        cursor.execute("SELECT course_id, semester_id FROM students WHERE student_id = %s", (session['user_id'],))
        student = cursor.fetchone()
        if not student:
            flash("Student information not found.", "error")
            return redirect('/login')
        student_course_id, student_semester_id = student
        
        # Fetch exam details
        cursor.execute("""
            SELECT exam_name, topic, exam_date, start_time, end_time, max_marks, course_id, semester_id
            FROM exams 
            WHERE exam_id = %s
        """, (exam_id,))
        exam = cursor.fetchone()
        if not exam:
            flash("Exam not found.", "error")
            return redirect('/student/exams')
        
        exam_name, topic, exam_date, start_time_delta, end_time_delta, max_marks, course_id, semester_id = exam
        
        if course_id != student_course_id or semester_id != student_semester_id:
            flash("You are not enrolled in this exam.", "error")
            return redirect('/student/exams')
        
        # Convert timedelta to time
        if isinstance(start_time_delta, timedelta):
            total_seconds = int(start_time_delta.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            start_time = time(hours, minutes, seconds)
        else:
            start_time = start_time_delta
        
        if isinstance(end_time_delta, timedelta):
            total_seconds = int(end_time_delta.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            end_time = time(hours, minutes, seconds)
        else:
            end_time = end_time_delta
        
        # Server-side time check
        current_datetime = datetime.now()
        start_dt = datetime.combine(exam_date, start_time)
        end_dt = datetime.combine(exam_date, end_time)
        if not (start_dt <= current_datetime <= end_dt):
            flash("The exam is not currently ongoing.", "error")
            return redirect('/student/exams')
        
        # Format times for display
        start_time_str = start_time.strftime('%H:%M:%S')
        end_time_str = end_time.strftime('%H:%M:%S')
        exam_date_str = exam_date.strftime('%Y-%m-%d')
        
        # Calculate end timestamp for JS countdown (in milliseconds)
        end_timestamp = int(end_dt.timestamp() * 1000)
        
        # Fetch questions with max_score
        cursor.execute("""
            SELECT question_id, question_text, max_score
            FROM questions 
            WHERE exam_id = %s
        """, (exam_id,))
        questions = cursor.fetchall()
        
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/student/exams')

    return render_template('student_take_exam.html', 
                           exam_id=exam_id, 
                           exam_name=exam_name, 
                           topic=topic, 
                           max_marks=max_marks, 
                           exam_date=exam_date_str,
                           start_time=start_time_str,
                           end_time=end_time_str,
                           end_timestamp=end_timestamp,
                           questions=questions, 
                           name=session.get('email'))  # Use email or fetch full_name if needed

@app.route('/student/submit_exam', methods=['POST'])
def student_submit_exam():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        exam_id = request.form.get('exam_id')
        if not exam_id:
            flash("Invalid exam.", "error")
            return redirect('/student/exams')
        
        # Fetch exam details
        cursor.execute("""
            SELECT exam_date, start_time, end_time, course_id, semester_id
            FROM exams 
            WHERE exam_id = %s
        """, (exam_id,))
        exam = cursor.fetchone()
        if not exam:
            flash("Exam not found.", "error")
            return redirect('/student/exams')
        
        # Check student's enrollment
        cursor.execute("SELECT course_id, semester_id FROM students WHERE student_id = %s", (session['user_id'],))
        student = cursor.fetchone()
        if not student or student[0] != exam[3] or student[1] != exam[4]:
            flash("You are not enrolled in this exam.", "error")
            return redirect('/student/exams')
        
        # Convert timedelta to time
        start_time_delta = exam[1]
        end_time_delta = exam[2]
        if isinstance(start_time_delta, timedelta):
            total_seconds = int(start_time_delta.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            start_time = time(hours, minutes, seconds)
        else:
            start_time = start_time_delta
        
        if isinstance(end_time_delta, timedelta):
            total_seconds = int(end_time_delta.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            end_time = time(hours, minutes, seconds)
        else:
            end_time = end_time_delta
        
        # Server-side time check
        current_datetime = datetime.now()
        start_dt = datetime.combine(exam[0], start_time)
        end_dt = datetime.combine(exam[0], end_time)
        if not (start_dt <= current_datetime <= end_dt):
            flash("The exam time has passed or not started yet.", "error")
            return redirect('/student/exams')
        
        # Fetch question IDs
        cursor.execute("SELECT question_id FROM questions WHERE exam_id = %s", (exam_id,))
        question_rows = cursor.fetchall()
        question_ids = [q[0] for q in question_rows]
        
        # Insert or update answers
        inserted = False
        for qid in question_ids:
            key = f"answer_{qid}"
            if key in request.form:
                answer_text = request.form[key].strip()
                if answer_text:
                    cursor.execute("""
                        SELECT answer_id FROM student_answers 
                        WHERE student_id = %s AND question_id = %s
                    """, (session['user_id'], qid))
                    existing = cursor.fetchone()
                    if existing:
                        cursor.execute("""
                            UPDATE student_answers 
                            SET answer_text = %s, submitted_at = CURRENT_TIMESTAMP
                            WHERE answer_id = %s
                        """, (answer_text, existing[0]))
                    else:
                        cursor.execute("""
                            INSERT INTO student_answers 
                            (student_id, question_id, answer_text)
                            VALUES (%s, %s, %s)
                        """, (session['user_id'], qid, answer_text))
                    inserted = True
        
        if inserted:
            conn.commit()
            flash("Answers submitted successfully!", "success")
        else:
            flash("No answers provided.", "warning")
        
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
    
    return redirect('/student/exams')



from datetime import datetime, time, timedelta
from flask import request, flash, redirect, render_template, session, url_for
import mysql.connector

import math          # <-- not strictly needed for round(), but good to have
# ... other imports ...

@app.route('/student/results')
def student_results():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')

    student_id = session['user_id']

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Fetch student info
        cursor.execute("""
            SELECT full_name, course_id, semester_id
            FROM students WHERE student_id = %s
        """, (student_id,))
        student = cursor.fetchone()
        if not student:
            flash("Student information not found.", "error")
            return redirect('/login')

        full_name = student['full_name']
        course_id = student['course_id']
        semester_id = student['semester_id']

        # Fetch exam results from result table
        cursor.execute("""
            SELECT 
                ex.exam_id,
                ex.exam_name,
                ex.topic,
                ex.exam_date,
                ex.max_marks,
                ex.min_marks,
                r.score AS total_score,
                r.final_score,
                COUNT(DISTINCT sa.answer_id) AS answered_questions,
                (SELECT COUNT(*) FROM questions WHERE exam_id = ex.exam_id) AS total_questions,
                MAX(sa.submitted_at) AS submission_date,
                CASE 
                    WHEN r.final_score >= ex.min_marks THEN 'Pass'
                    WHEN r.final_score IS NULL THEN 'Not Evaluated'
                    ELSE 'Fail'
                END AS result_status
            FROM student_answers sa
            JOIN questions q ON sa.question_id = q.question_id
            JOIN exams ex ON q.exam_id = ex.exam_id
            LEFT JOIN result r ON r.exam_id = ex.exam_id AND r.student_id = sa.student_id
            WHERE sa.student_id = %s
              AND ex.course_id = %s
              AND ex.semester_id = %s
            GROUP BY ex.exam_id, ex.exam_name, ex.topic,
                     ex.exam_date, ex.max_marks, ex.min_marks, r.score, r.final_score
            ORDER BY ex.exam_date DESC
        """, (student_id, course_id, semester_id))
        results = cursor.fetchall()

        # Process results
        for r in results:
            # Convert to proper types
            r['total_score'] = round(float(r['total_score']), 2) if r['total_score'] else 0.0
            r['final_score'] = int(r['final_score']) if r['final_score'] else 0
            r['max_marks'] = float(r['max_marks'] or 0)
            r['min_marks'] = float(r['min_marks'] or 0)

            # Determine evaluation status
            if r['total_score'] > 0:
                r['evaluation_status'] = 'Evaluated'
                r['percentage'] = (r['total_score'] / r['max_marks'] * 100) if r['max_marks'] else 0
            else:
                r['evaluation_status'] = 'Not Evaluated'
                r['percentage'] = 0
                r['result_status'] = 'Not Evaluated'

        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        results = []
        full_name = session.get('email', 'Student')
    except Exception as e:
        flash(f"An error occurred: {str(e)}", "error")
        results = []
        full_name = session.get('email', 'Student')

    return render_template('student_results.html',
                           results=results,
                           name=full_name)
    
    
@app.route('/student/view_result_detail/<int:exam_id>')
def student_view_result_detail(exam_id):
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')

    student_id = session['user_id']

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Fetch student name
        cursor.execute("SELECT full_name FROM students WHERE student_id = %s", (student_id,))
        student = cursor.fetchone()
        student_name = student['full_name'] if student else session.get('email')

        # Fetch exam header
        cursor.execute("""
            SELECT exam_name, topic, exam_date, max_marks, min_marks
            FROM exams WHERE exam_id = %s
        """, (exam_id,))
        exam = cursor.fetchone()
        if not exam:
            flash("Exam not found.", "error")
            return redirect('/student/results')

        # Fetch result from result table
        cursor.execute("""
            SELECT score, final_score
            FROM result
            WHERE student_id = %s AND exam_id = %s
        """, (student_id, exam_id))
        result_data = cursor.fetchone()

        # Fetch question details
        cursor.execute("""
            SELECT 
                q.question_id,
                q.question_text,
                q.model_answer,
                q.max_score,
                sa.answer_text,
                COALESCE(e.score, 0) AS score,
                sa.submitted_at
            FROM questions q
            LEFT JOIN student_answers sa 
                   ON sa.question_id = q.question_id 
                  AND sa.student_id = %s
            LEFT JOIN evaluations e 
                   ON e.answer_id = sa.answer_id
            WHERE q.exam_id = %s
            ORDER BY q.question_id
        """, (student_id, exam_id))
        questions = cursor.fetchall()

        # Process scores
        for q in questions:
            q['score'] = float(q['score'] or 0)
            q['max_score'] = float(q['max_score'] or 0)

        # Use result table data if available
        if result_data:
            total_score = round(float(result_data['score'] or 0), 2)
            final_score = int(result_data['final_score'] or 0)
            is_evaluated = True
        else:
            total_score = 0.0
            final_score = 0
            is_evaluated = False

        total_max = sum(float(q['max_score']) for q in questions)
        
        exam['max_marks'] = float(exam['max_marks'] or 0)
        exam['min_marks'] = float(exam['min_marks'] or 0)

        percentage = (total_score / total_max * 100) if total_max else 0
        result_status = ('Pass' if final_score >= exam['min_marks'] else 'Fail') if is_evaluated else 'Not Evaluated'

        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/student/results')
    except Exception as e:
        flash(f"An error occurred: {str(e)}", "error")
        return redirect('/student/results')

    return render_template('student_result_detail.html',
                           exam=exam,
                           questions=questions,
                           total_score=total_score,
                           final_score=final_score,
                           total_max_score=total_max,
                           percentage=percentage,
                           is_evaluated=is_evaluated,
                           result_status=result_status,
                           result_data=result_data,
                           name=student_name)
    
            

# Student Profile Route
@app.route('/student/profile')
def student_profile():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT s.full_name, s.roll_no, s.enrollment_no, s.contact, s.dob,
                   c.course_name, sem.semester_name, s.gender, s.address, 
                   s.department, s.university, u.email
            FROM students s
            JOIN users u ON s.student_id = u.uid
            LEFT JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN semesters sem ON s.semester_id = sem.semester_id
            WHERE s.student_id = %s
        """, (session['user_id'],))
        profile = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not profile:
            flash("No profile found. Please complete your profile.", "warning")
            return redirect('/student/complete_profile')
        
        return render_template('student_profile.html', 
                             profile=profile, 
                             email=profile[11], 
                             name=session.get('email'))
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/student_dashboard')


# Student Complete Profile Route
@app.route('/student/complete_profile', methods=['GET', 'POST'])
def complete_student_profile():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')
    
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    
    # Fetch existing profile
    cursor.execute("SELECT * FROM students WHERE student_id = %s", (session['user_id'],))
    existing_profile = cursor.fetchone()
    
    if request.method == 'POST':
        # Check if profile already complete
        if existing_profile and all([
            existing_profile.get('full_name'),
            existing_profile.get('dob'),
            existing_profile.get('contact'),
            existing_profile.get('gender'),
            existing_profile.get('address'),
            existing_profile.get('department'),
            existing_profile.get('university')
        ]):
            flash("Profile already completed.", "info")
            cursor.close()
            conn.close()
            return redirect('/student_dashboard')
        
        # Skip button
        if request.form.get('skip') == '1':
            cursor.close()
            conn.close()
            return redirect('/student_dashboard')
        
        # Get form inputs
        full_name = request.form.get('full_name', '').strip()
        dob = request.form.get('dob', '').strip()
        contact = request.form.get('contact', '').strip()
        gender = request.form.get('gender', '').strip()
        address = request.form.get('address', '').strip()
        department = request.form.get('department', '').strip()
        university = request.form.get('university', '').strip()
        
        # Validation
        errors = []
        
        if not re.match(r'^[A-Za-z ]{3,50}$', full_name):
            errors.append("Full name must be 3-50 alphabetic characters.")
        
        try:
            dob_date = datetime.strptime(dob, '%Y-%m-%d').date()
            age = (datetime.now().date() - dob_date).days // 365
            if age < 15:
                errors.append("Student must be at least 15 years old.")
        except ValueError:
            errors.append("Invalid date format for DOB. Use YYYY-MM-DD.")
        
        if not re.match(r'^\d{10}$', contact):
            errors.append("Contact number must be exactly 10 digits.")
        
        if gender not in ['Male', 'Female', 'Other']:
            errors.append("Invalid gender selected.")
        
        if len(address) < 5:
            errors.append("Address must be at least 5 characters.")
        
        if len(department) < 2:
            errors.append("Department must be at least 2 characters.")
        
        if len(university) < 2:
            errors.append("University name must be at least 2 characters.")
        
        if errors:
            for e in errors:
                flash(e, "error")
            cursor.close()
            conn.close()
            return render_template('complete_student_profile.html', name=session.get('email'))
        
        # Update database
        try:
            cursor.execute("""
                UPDATE students
                SET full_name=%s, dob=%s, contact=%s, gender=%s, 
                    address=%s, department=%s, university=%s
                WHERE student_id=%s
            """, (full_name, dob_date, contact, gender, address, 
                  department, university, session['user_id']))
            conn.commit()
            flash("Profile completed successfully!", "success")
        except mysql.connector.Error as err:
            flash(f"Database error: {err}", "error")
        
        cursor.close()
        conn.close()
        return redirect('/student_dashboard')
    
    cursor.close()
    conn.close()
    return render_template('complete_student_profile.html', name=session.get('email'))


# Student Edit Profile Route
@app.route('/student/edit_profile', methods=['GET', 'POST'])
def student_edit_profile():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT s.full_name, s.roll_no, s.enrollment_no, s.contact, s.dob,
                   c.course_name, sem.semester_name, s.gender, s.address, 
                   s.department, s.university, u.email
            FROM students s
            JOIN users u ON s.student_id = u.uid
            LEFT JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN semesters sem ON s.semester_id = sem.semester_id
            WHERE s.student_id = %s
        """, (session['user_id'],))
        profile = cursor.fetchone()
        
        if request.method == 'POST':
            full_name = request.form.get('full_name', '').strip()
            dob = request.form.get('dob', '').strip()
            contact = request.form.get('contact', '').strip()
            gender = request.form.get('gender', '').strip()
            address = request.form.get('address', '').strip()
            department = request.form.get('department', '').strip()
            university = request.form.get('university', '').strip()
            
            # Validation
            errors = []
            
            if not re.match(r'^[A-Za-z ]{3,50}$', full_name):
                errors.append("Full name must be 3-50 alphabetic characters.")
            
            try:
                dob_date = datetime.strptime(dob, '%Y-%m-%d').date()
                age = (datetime.now().date() - dob_date).days // 365
                if age < 15:
                    errors.append("Student must be at least 15 years old.")
            except ValueError:
                errors.append("Invalid date format for DOB. Use YYYY-MM-DD.")
            
            if not re.match(r'^\d{10}$', contact):
                errors.append("Contact number must be exactly 10 digits.")
            
            if gender not in ['Male', 'Female', 'Other']:
                errors.append("Invalid gender selected.")
            
            if len(address) < 5:
                errors.append("Address must be at least 5 characters.")
            
            if len(department) < 2:
                errors.append("Department must be at least 2 characters.")
            
            if len(university) < 2:
                errors.append("University name must be at least 2 characters.")
            
            if errors:
                for e in errors:
                    flash(e, "error")
                cursor.close()
                conn.close()
                return render_template('student_edit_profile.html', 
                                     profile=profile, 
                                     email=profile[11] if profile else session.get('email'), 
                                     name=session.get('email'))
            
            cursor.execute("""
                UPDATE students
                SET full_name=%s, dob=%s, contact=%s, gender=%s, 
                    address=%s, department=%s, university=%s
                WHERE student_id=%s
            """, (full_name, dob_date, contact, gender, address, 
                  department, university, session['user_id']))
            conn.commit()
            flash("Profile updated successfully!", "success")
            cursor.close()
            conn.close()
            return redirect('/student/profile')
        
        cursor.close()
        conn.close()
        return render_template('student_edit_profile.html', 
                             profile=profile, 
                             email=profile[11] if profile else session.get('email'), 
                             name=session.get('email'))
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/student_dashboard')


@app.route('/student/progress')
def student_progress():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')
    
    student_id = session['user_id']
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Fetch student info
        cursor.execute("""
            SELECT s.full_name, s.course_id, s.semester_id, c.course_name, sem.semester_name
            FROM students s
            LEFT JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN semesters sem ON s.semester_id = sem.semester_id
            WHERE s.student_id = %s
        """, (student_id,))
        student = cursor.fetchone()
        
        if not student:
            student = {
                'full_name': session.get('email'), 
                'course_id': None, 
                'semester_id': None,
                'course_name': 'N/A',
                'semester_name': 'N/A'
            }
        
        # Overall statistics with better calculations
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT ex.exam_id) AS total_exams_available,
                COUNT(DISTINCT CASE WHEN sa.student_id IS NOT NULL THEN ex.exam_id END) AS total_exams_attempted,
                COUNT(DISTINCT sa.answer_id) AS total_submissions,
                COUNT(DISTINCT CASE WHEN ev.evaluation_id IS NOT NULL THEN sa.answer_id END) AS evaluated_submissions
            FROM exams ex
            LEFT JOIN questions q ON ex.exam_id = q.exam_id
            LEFT JOIN student_answers sa ON q.question_id = sa.question_id AND sa.student_id = %s
            LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
            WHERE ex.course_id = %s AND ex.semester_id = %s
        """, (student_id, student['course_id'], student['semester_id']))
        overall_stats = cursor.fetchone()
        
        # Calculate exams passed and total score details
        cursor.execute("""
            SELECT 
                ex.exam_id,
                ex.exam_name,
                ex.max_marks,
                ex.min_marks,
                SUM(COALESCE(ev.score, 0)) AS total_score,
                CASE WHEN SUM(COALESCE(ev.score, 0)) >= ex.min_marks THEN 1 ELSE 0 END AS passed
            FROM student_answers sa
            JOIN questions q ON sa.question_id = q.question_id
            JOIN exams ex ON q.exam_id = ex.exam_id
            LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
            WHERE sa.student_id = %s 
                AND ex.course_id = %s 
                AND ex.semester_id = %s
            GROUP BY ex.exam_id, ex.exam_name, ex.max_marks, ex.min_marks
        """, (student_id, student['course_id'], student['semester_id']))
        exam_results = cursor.fetchall()
        
        exams_passed = sum(1 for exam in exam_results if exam['passed'] == 1)
        total_exams_taken = len(exam_results)
        
        # Calculate average percentage
        if exam_results:
            avg_percentage = sum((exam['total_score'] / exam['max_marks']) * 100 for exam in exam_results) / len(exam_results)
        else:
            avg_percentage = 0
        
        # Subject-wise performance with exam counts
        cursor.execute("""
            SELECT 
                sub.subject_id,
                sub.subject_name,
                COUNT(DISTINCT exam_data.exam_id) AS exams_count,
                ROUND(AVG((exam_data.exam_score / exam_data.exam_max) * 100), 2) AS average_percentage,
                SUM(exam_data.exam_score) AS total_score_obtained,
                SUM(exam_data.exam_max) AS total_max_marks,
                SUM(CASE WHEN exam_data.exam_score >= exam_data.min_marks THEN 1 ELSE 0 END) AS passed_count
            FROM (
                SELECT 
                    ex.exam_id,
                    ex.subject_id,
                    ex.min_marks,
                    SUM(COALESCE(ev.score, 0)) AS exam_score,
                    ex.max_marks AS exam_max
                FROM student_answers sa
                JOIN questions q ON sa.question_id = q.question_id
                JOIN exams ex ON q.exam_id = ex.exam_id
                LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
                WHERE sa.student_id = %s 
                    AND ex.course_id = %s 
                    AND ex.semester_id = %s
                GROUP BY ex.exam_id, ex.subject_id, ex.max_marks, ex.min_marks
            ) AS exam_data
            JOIN subjects sub ON exam_data.subject_id = sub.subject_id
            GROUP BY sub.subject_id, sub.subject_name
            ORDER BY average_percentage DESC
        """, (student_id, student['course_id'], student['semester_id']))
        subject_performance = cursor.fetchall()
        
        # Monthly trend with better date formatting
        # cursor.execute("""
        #     SELECT 
        #         DATE_FORMAT(exam_date, '%%Y-%%m') AS month,
        #         DATE_FORMAT(exam_date, '%%b %%Y') AS month_name,
        #         COUNT(DISTINCT exam_id) AS exams_count,
        #         ROUND(AVG((exam_score / exam_max) * 100), 2) AS avg_percentage,
        #         ROUND((SUM(CASE WHEN exam_score >= min_marks THEN 1 ELSE 0 END) / COUNT(DISTINCT exam_id)) * 100, 2) AS pass_rate,
        #         SUM(exam_score) AS total_score,
        #         SUM(exam_max) AS total_max
        #     FROM (
        #         SELECT 
        #             ex.exam_id,
        #             ex.exam_date,
        #             ex.min_marks,
        #             SUM(COALESCE(ev.score, 0)) AS exam_score,
        #             ex.max_marks AS exam_max
        #         FROM student_answers sa
        #         JOIN questions q ON sa.question_id = q.question_id
        #         JOIN exams ex ON q.exam_id = ex.exam_id
        #         LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
        #         WHERE sa.student_id = %s 
        #             AND ex.course_id = %s 
        #             AND ex.semester_id = %s
        #             AND ex.exam_date >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
        #         GROUP BY ex.exam_id, ex.exam_date, ex.max_marks, ex.min_marks
        #     ) AS exam_data
        #     GROUP BY DATE_FORMAT(exam_date, '%%Y-%%m'), DATE_FORMAT(exam_date, '%%b %%Y')
        #     ORDER BY month ASC
        # """, (student_id, student['course_id'], student['semester_id']))
        # monthly_trend = cursor.fetchall()
        
        # # Add trend indicators
        # for i in range(len(monthly_trend)):
        #     if i > 0:
        #         prev_score = monthly_trend[i-1]['avg_percentage'] or 0
        #         curr_score = monthly_trend[i]['avg_percentage'] or 0
        #         diff = curr_score - prev_score
        #         if diff > 5:
        #             monthly_trend[i]['trend'] = 'up'
        #             monthly_trend[i]['trend_icon'] = '📈'
        #         elif diff < -5:
        #             monthly_trend[i]['trend'] = 'down'
        #             monthly_trend[i]['trend_icon'] = '📉'
        #         else:
        #             monthly_trend[i]['trend'] = 'stable'
        #             monthly_trend[i]['trend_icon'] = '➡️'
        #         monthly_trend[i]['trend_diff'] = abs(diff)
        #     else:
        #         monthly_trend[i]['trend'] = 'stable'
        #         monthly_trend[i]['trend_icon'] = '➡️'
        #         monthly_trend[i]['trend_diff'] = 0
        
        # Get exam-wise detailed results
        cursor.execute("""
            SELECT 
                ex.exam_id,
                ex.exam_name,
                DATE_FORMAT(ex.exam_date, '%%d %%b %%Y') AS exam_date_formatted,
                ex.exam_date,
                SUM(COALESCE(ev.score, 0)) AS final_score,
                ex.max_marks,
                ex.min_marks,
                ROUND((SUM(COALESCE(ev.score, 0)) / ex.max_marks) * 100, 2) AS percentage,
                sub.subject_name,
                CASE WHEN SUM(COALESCE(ev.score, 0)) >= ex.min_marks THEN 'Pass' ELSE 'Fail' END AS status,
                COUNT(q.question_id) AS total_questions,
                COUNT(CASE WHEN ev.evaluation_id IS NOT NULL THEN 1 END) AS evaluated_questions
            FROM student_answers sa
            JOIN questions q ON sa.question_id = q.question_id
            JOIN exams ex ON q.exam_id = ex.exam_id
            JOIN subjects sub ON ex.subject_id = sub.subject_id
            LEFT JOIN evaluations ev ON sa.answer_id = ev.answer_id
            WHERE sa.student_id = %s
                AND ex.course_id = %s 
                AND ex.semester_id = %s
            GROUP BY ex.exam_id, ex.exam_name, ex.exam_date, ex.max_marks, ex.min_marks, sub.subject_name
            ORDER BY ex.exam_date DESC
        """, (student_id, student['course_id'], student['semester_id']))
        exam_details = cursor.fetchall()
        
        # Grade distribution
        grade_distribution = {'A+': 0, 'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        for exam in exam_details:
            percentage = exam['percentage']
            if percentage >= 90:
                grade_distribution['A+'] += 1
            elif percentage >= 80:
                grade_distribution['A'] += 1
            elif percentage >= 70:
                grade_distribution['B'] += 1
            elif percentage >= 60:
                grade_distribution['C'] += 1
            elif percentage >= 40:
                grade_distribution['D'] += 1
            else:
                grade_distribution['F'] += 1
        
        # Strengths (top subjects)
        strengths = sorted(subject_performance, key=lambda x: x['average_percentage'] or 0, reverse=True)[:3] if subject_performance else []
        
        # Weaknesses (subjects needing improvement)
        weaknesses = [s for s in subject_performance if (s['average_percentage'] or 0) < 60]
        weaknesses = sorted(weaknesses, key=lambda x: x['average_percentage'] or 0)[:3]
        
        # Calculate improvement rate
        if len(exam_details) >= 2:
            recent_avg = sum(exam['percentage'] for exam in exam_details[:5]) / min(5, len(exam_details))
            older_avg = sum(exam['percentage'] for exam in exam_details[-5:]) / min(5, len(exam_details[-5:]))
            improvement_rate = recent_avg - older_avg
        else:
            improvement_rate = 0
        
        # Prepare progress data
        progress = {
            'total_exams': total_exams_taken,
            'total_exams_available': overall_stats['total_exams_available'] or 0,
            'total_submissions': overall_stats['total_submissions'] or 0,
            'evaluated_submissions': overall_stats['evaluated_submissions'] or 0,
            'exams_passed': exams_passed,
            'exams_failed': total_exams_taken - exams_passed,
            'average_percentage': round(avg_percentage, 2),
            'subject_performance': subject_performance,
            # 'monthly_trend': monthly_trend,
            'strengths': [{'subject': s['subject_name'], 'score': s['average_percentage'] or 0, 'exams': s['exams_count']} for s in strengths],
            'weaknesses': [{'subject': s['subject_name'], 'score': s['average_percentage'] or 0, 'exams': s['exams_count']} for s in weaknesses],
            'exam_details': exam_details,
            'grade_distribution': grade_distribution,
            'improvement_rate': round(improvement_rate, 2),
            'pass_rate': round((exams_passed / total_exams_taken * 100), 2) if total_exams_taken > 0 else 0
        }
        
        cursor.close()
        conn.close()
        
    except mysql.connector.Error as err:
        print(f"Database error: {err}")
        flash(f"Database error: {err}", "error")
        progress = {
            'total_exams': 0,
            'total_exams_available': 0,
            'total_submissions': 0,
            'evaluated_submissions': 0,
            'exams_passed': 0,
            'exams_failed': 0,
            'average_percentage': 0,
            'subject_performance': [],
            'monthly_trend': [],
            'strengths': [],
            'weaknesses': [],
            'exam_details': [],
            'grade_distribution': {'A+': 0, 'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0},
            'improvement_rate': 0,
            'pass_rate': 0
        }
        student = {'full_name': session.get('email'), 'course_name': 'N/A', 'semester_name': 'N/A'}
    
    return render_template('student_progress.html', 
                         progress=progress, 
                         name=student.get('full_name', session.get('email')),
                         course=student.get('course_name', 'N/A'),
                         semester=student.get('semester_name', 'N/A'))


@app.route('/admin_dashboard')
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users")
        user_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'teacher'")
        teacher_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'student'")
        student_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM questions")
        question_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM student_answers")
        answer_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM evaluations")
        evaluation_count = cursor.fetchone()[0]
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        user_count = teacher_count = student_count = question_count = answer_count = evaluation_count = 0

    return render_template('admin_dashboard.html', 
                         user_count=user_count, 
                         teacher_count=teacher_count, 
                         student_count=student_count, 
                         question_count=question_count, 
                         answer_count=answer_count, 
                         evaluation_count=evaluation_count, 
                         name=session.get('email'))
    


@app.route('/admin/teachers')
def admin_teachers():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)  # dictionary=True so you can access by column name

        cursor.execute("""
            SELECT 
                t.teacher_id,
                u.email,
                u.role,
                t.full_name,
                t.dob,
                t.last_degree,
                t.contact,
                t.gender,
                t.address,
                t.expertise,
                t.subjects_taught,
                t.experience_years,
                t.industry_experience_years,
                t.research_papers,
                t.department,
                t.university
            FROM teachers t
            JOIN users u ON t.teacher_id = u.uid
        """)
        teachers = cursor.fetchall()

        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        teachers = []

    return render_template('admin_teachers.html', teachers=teachers, name=session.get('email'))



# ---------------- DOWNLOAD TEACHER TEMPLATE ----------------
@app.route('/admin/download_teacher_template')
def download_teacher_template():
    # Create a sample dataframe
    df = pd.DataFrame({
        "email": [
            "teacher1@example.com",
            "teacher2@example.com",
            "teacher3@example.com"
        ]
    })

    # Save to in-memory buffer
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name="Teachers")
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="teacher_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    
@app.route('/admin/download_student_template')
def download_student_template():
    # Sample student data
    df = pd.DataFrame({
        "email": [
            "student1@example.com",
            "student2@example.com",
            "student3@example.com"
        ],
        "roll_no": [
            1,
            2,
            3
        ],
        "enroll_no": [
            202528900101,
            202528900102,
            202528900103
        ]
    })

    # Write to memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name="Students")

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="student_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# ---------------- UPLOAD TEACHERS EXCEL ----------------
@app.route('/admin/upload_teachers_excel', methods=['POST'])
def upload_teachers_excel():
    if 'excel_file' not in request.files:
        flash("No file uploaded", "error")
        return redirect(url_for('admin_teachers'))

    file = request.files['excel_file']
    if file.filename == '':
        flash("No file selected", "error")
        return redirect(url_for('admin_teachers'))

    try:
        df = pd.read_excel(file)

        if 'email' not in df.columns:
            flash("Excel file must contain 'email' column", "error")
            return redirect(url_for('admin_teachers'))

        emails = df['email'].dropna().tolist()

        # Validate all emails
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        invalid_emails = [email for email in emails if not re.match(email_pattern, email)]

        if invalid_emails:
            flash(f"Invalid emails found: {', '.join(invalid_emails)}", "error")
            return redirect(url_for('admin_teachers'))

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        added_count = 0
        for email in emails:
            # Check if email already exists
            cursor.execute("SELECT uid FROM users WHERE email = %s", (email,))
            existing_user = cursor.fetchone()
            if existing_user:
                continue  # skip already registered teachers

            # Generate password + hash
            password_plain = generate_password(12)
            password_hashed = generate_password_hash(password_plain)

            # Insert into users table
            cursor.execute("""
                INSERT INTO users (email, password, role)
                VALUES (%s, %s, %s)
            """, (email, password_hashed, 'teacher'))
            uid = cursor.lastrowid  # get generated uid

            # Insert into teachers table with only uid (other cols NULL/empty)
            cursor.execute("""
                INSERT INTO teachers (
                    teacher_id, full_name, dob, last_degree, contact, gender,
                    address, expertise, subjects_taught, experience_years,
                    industry_experience_years, research_papers, department, university
                )
                VALUES (
                    %s, 'Not Provided', NULL, 'Not Provided', 'Not Provided', 'Not Provided',
                    'Not Provided', 'Not Provided', 'Not Provided', 0,
                    0, 0, 'Not Provided', 'Not Provided'
                )
            """, (uid,))


            # Send email with login details
            subject = "EduAI Teacher Registration Successful"
            body = f"""
            Dear Teacher,

            Your registration on EduAI was successful.

            Login Details:
            Email: {email}
            Temporary Password: {password_plain}

            Please log in and update your profile & change your password immediately.

            Regards,
            EduAI Team
            """
            send_email(email, subject, body)

            added_count += 1

        conn.commit()
        cursor.close()
        conn.close()

        if added_count > 0:
            flash(f"Successfully registered {added_count} teachers and sent credentials.", "success")
        else:
            flash("No new teachers were added (all emails already registered).", "info")

    except Exception as e:
        flash(f"Error processing file: {str(e)}", "error")

    return redirect(url_for('admin_teachers'))




@app.route('/admin/upload_students_excel', methods=['POST'])
def upload_students_excel():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect(url_for('login'))

    if 'excel_file' not in request.files:
        flash("No file uploaded", "error")
        return redirect(url_for('admin_students'))

    file = request.files['excel_file']
    if file.filename == '':
        flash("No file selected", "error")
        return redirect(url_for('admin_students'))

    course_id = request.form.get('course_id')
    semester_id = request.form.get('semester_id')

    if not course_id or not semester_id:
        flash("Course and semester must be selected.", "error")
        return redirect(url_for('admin_students'))

    try:
        # Step 1: Read Excel
        df = pd.read_excel(file)

        # Step 2: Check required columns
        required_columns = ['email', 'roll_no', 'enroll_no']
        if not all(col in df.columns for col in required_columns):
            flash("Excel file must contain 'email', 'roll_no', and 'enroll_no' columns", "error")
            return redirect(url_for('admin_students'))

        df = df[required_columns].dropna()
        emails = df['email'].tolist()
        roll_nos = df['roll_no'].tolist()
        enroll_nos = df['enroll_no'].tolist()

        # Step 3: Validate emails
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        invalid_emails = [email for email in emails if not re.match(email_pattern, email)]
        if invalid_emails:
            flash(f"Invalid emails found: {', '.join(invalid_emails)}", "error")
            return redirect(url_for('admin_students'))

        # Step 4: Validate DB uniqueness
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        existing_emails, existing_roll_nos, existing_enroll_nos = [], [], []
        for email, roll_no, enroll_no in zip(emails, roll_nos, enroll_nos):
            cursor.execute("SELECT uid FROM users WHERE email = %s", (email,))
            if cursor.fetchone():
                existing_emails.append(email)

            cursor.execute("SELECT student_id FROM students WHERE roll_no = %s", (roll_no,))
            if cursor.fetchone():
                existing_roll_nos.append(roll_no)

            cursor.execute("SELECT student_id FROM students WHERE enrollment_no = %s", (enroll_no,))
            if cursor.fetchone():
                existing_enroll_nos.append(enroll_no)

        if existing_emails or existing_roll_nos or existing_enroll_nos:
            errors = []
            if existing_emails:
                errors.append(f"Emails already registered: {', '.join(existing_emails)}")
            if existing_roll_nos:
                errors.append(f"Roll numbers already exist: {', '.join(map(str, existing_roll_nos))}")
            if existing_enroll_nos:
                errors.append(f"Enrollment numbers already exist: {', '.join(map(str, existing_enroll_nos))}")

            for error in errors:
                flash(error, "error")
            cursor.close()
            conn.close()
            return redirect(url_for('admin_students'))

        # Step 5: Insert into DB (prepare data first, no emails yet)
        added_students = []  # [(email, plain_password)]
        for email, roll_no, enroll_no in zip(emails, roll_nos, enroll_nos):
            password_plain = generate_password(12)
            password_hashed = generate_password_hash(password_plain)

            # Insert into users
            cursor.execute("""
                INSERT INTO users (email, password, role)
                VALUES (%s, %s, %s)
            """, (email, password_hashed, 'student'))
            uid = cursor.lastrowid

            # Insert into students
            cursor.execute("""
                INSERT INTO students (
                    student_id, full_name, roll_no, enrollment_no, contact, dob,
                    course_id, semester_id, gender, address, department, university
                )
                VALUES (
                    %s, 'Not Provided', %s, %s, 'Not Provided', NULL,
                    %s, %s, 'Not Provided', 'Not Provided', 'Not Provided', 'Not Provided'
                )
            """, (uid, roll_no, enroll_no, course_id, semester_id))

            added_students.append((email, password_plain))

        conn.commit()
        cursor.close()
        conn.close()

        # Step 6: Send emails after commit succeeded
        for email, password_plain in added_students:
            subject = "EduAI Student Registration Successful"
            body = f"""
            Dear Student,

            Your registration on EduAI was successful.

            Login Details:
            Email: {email}
            Temporary Password: {password_plain}

            Please log in and update your profile & change your password immediately.

            Regards,
            EduAI Team
            """
            send_email(email, subject, body)

        if added_students:
            flash(f"Successfully registered {len(added_students)} students and sent credentials.", "success")
        else:
            flash("No new students were added.", "info")

    except Exception as e:
        flash(f"Error processing file: {str(e)}", "error")

    return redirect(url_for('admin_students'))




@app.route('/admin/students', methods=['GET', 'POST'])
def admin_students():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect(url_for('login'))
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Fetch courses for the upload and filter forms
        cursor.execute("SELECT course_id, course_name FROM courses")
        courses = cursor.fetchall()
        
        # Get filter parameters from the request
        course_id = request.form.get('course_id') if request.method == 'POST' else request.args.get('course_id')
        semester_id = request.form.get('semester_id') if request.method == 'POST' else request.args.get('semester_id')
        
        # Convert to integers and validate
        try:
            course_id = int(course_id) if course_id else None
            semester_id = int(semester_id) if semester_id else None
        except (ValueError, TypeError):
            course_id = None
            semester_id = None
            flash("Invalid course or semester selection.", "error")
        
        # Build the SQL query with optional filters
        query = """
            SELECT 
                s.student_id, s.full_name, s.roll_no, s.enrollment_no, s.contact, s.dob,
                s.course_id, c.course_name, s.semester_id, sem.semester_name,
                s.gender, s.address, s.department, s.university
            FROM students s
            LEFT JOIN courses c ON s.course_id = c.course_id
            LEFT JOIN semesters sem ON s.semester_id = sem.semester_id
        """
        params = []
        conditions = []
        
        if course_id:
            conditions.append("s.course_id = %s")
            params.append(course_id)
        if semester_id:
            conditions.append("s.semester_id = %s")
            params.append(semester_id)
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        # Debug: Log query and parameters
        print(f"Query: {query}, Params: {params}")
        
        # Fetch students
        cursor.execute(query, params)
        students = cursor.fetchall()
        
        # Debug: Log number of students fetched
        print(f"Fetched {len(students)} students")
        
        cursor.close()
        conn.close()
        return render_template('admin_students.html', courses=courses, students=students, name=session.get('email'), 
                             selected_course_id=course_id, selected_semester_id=semester_id)
    except mysql.connector.Error as e:
        flash(f"Database error: {e}", "error")
        return redirect(url_for('admin_dashboard'))
    
            
    
    
    

@app.route('/admin/courses_semesters')
def admin_courses_semesters():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Fetch courses with semester count
        cursor.execute(
            """
            SELECT c.course_id, c.course_name, c.description, 
                   COUNT(s.semester_id) AS semester_count
            FROM courses c
            LEFT JOIN semesters s ON c.course_id = s.course_id
            GROUP BY c.course_id, c.course_name, c.description
            """
        )
        courses = cursor.fetchall()
        
        # Fetch semesters with course names
        cursor.execute(
            """
            SELECT s.semester_id, c.course_name, s.semester_name, s.start_date, s.end_date, s.course_id
            FROM semesters s
            JOIN courses c ON s.course_id = c.course_id
            """
        )
        semesters = cursor.fetchall()
        
        # Fetch subjects with course and semester names
        cursor.execute(
            """
            SELECT sub.subject_id, c.course_name, s.semester_name, sub.subject_name, sub.course_id, sub.semester_id
            FROM subjects sub
            JOIN courses c ON sub.course_id = c.course_id
            JOIN semesters s ON sub.semester_id = s.semester_id
            """
        )
        subjects = cursor.fetchall()
        
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        courses = []
        semesters = []
        subjects = []

    return render_template('admin_courses_semesters.html', courses=courses, semesters=semesters, 
                         subjects=subjects, name=session.get('email'), 
                         course_to_edit=None, semester_to_edit=None, subject_to_edit=None)

@app.route('/admin/add_course', methods=['POST'])
def admin_add_course():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    course_name = request.form.get('course_name', '').strip()
    description = request.form.get('description', '').strip()
    
    # Validation
    errors = []
    if not course_name or len(course_name) < 2 or len(course_name) > 100:
        errors.append("Course name must be 2–100 characters.")
    if len(description) > 65535:  # TEXT field limit in MySQL
        errors.append("Description is too long (max 65535 characters).")
    
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect('/admin/courses_semesters')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        # Check if course_name already exists
        cursor.execute("SELECT COUNT(*) FROM courses WHERE course_name = %s", (course_name,))
        if cursor.fetchone()[0] > 0:
            flash("Course name already exists.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Insert new course
        cursor.execute(
            "INSERT INTO courses (course_name, description) VALUES (%s, %s)",
            (course_name, description or None)
        )
        conn.commit()
        flash("Course added successfully!", "success")
        
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
    
    return redirect('/admin/courses_semesters')

@app.route('/admin/add_semester', methods=['POST'])
def admin_add_semester():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    course_id = request.form.get('course_id', '').strip()
    semester_name = request.form.get('semester_name', '').strip()
    start_date = request.form.get('start_date', '').strip()
    end_date = request.form.get('end_date', '').strip()
    
    # Validation
    errors = []
    if not course_id or not course_id.isdigit():
        errors.append("Invalid course selected.")
    if not semester_name or len(semester_name) < 2 or len(semester_name) > 50:
        errors.append("Semester name must be 2–50 characters.")
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        if start_date >= end_date:
            errors.append("End date must be after start date.")
    except ValueError:
        errors.append("Invalid date format for start or end date. Use YYYY-MM-DD.")
    
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect('/admin/courses_semesters')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        # Check if course_id exists
        cursor.execute("SELECT COUNT(*) FROM courses WHERE course_id = %s", (course_id,))
        if cursor.fetchone()[0] == 0:
            flash("Selected course does not exist.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Check if semester_name already exists for this course
        cursor.execute(
            "SELECT COUNT(*) FROM semesters WHERE course_id = %s AND semester_name = %s",
            (course_id, semester_name)
        )
        if cursor.fetchone()[0] > 0:
            flash("Semester name already exists for this course.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Insert new semester
        cursor.execute(
            "INSERT INTO semesters (course_id, semester_name, start_date, end_date) VALUES (%s, %s, %s, %s)",
            (course_id, semester_name, start_date, end_date)
        )
        conn.commit()
        flash("Semester added successfully!", "success")
        
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
    
    return redirect('/admin/courses_semesters')

@app.route('/admin/edit_course/<int:course_id>', methods=['GET', 'POST'])
def admin_edit_course(course_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Fetch the course to edit
        cursor.execute(
            "SELECT course_id, course_name, description FROM courses WHERE course_id = %s",
            (course_id,)
        )
        course_to_edit = cursor.fetchone()
        
        if not course_to_edit:
            flash("Course not found.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        if request.method == 'POST':
            course_name = request.form.get('course_name', '').strip()
            description = request.form.get('description', '').strip()
            
            # Validation
            errors = []
            if not course_name or len(course_name) < 2 or len(course_name) > 100:
                errors.append("Course name must be 2–100 characters.")
            if len(description) > 65535:
                errors.append("Description is too long (max 65535 characters).")
            
            # Check if course_name is unique (excluding the current course)
            cursor.execute(
                "SELECT COUNT(*) FROM courses WHERE course_name = %s AND course_id != %s",
                (course_name, course_id)
            )
            if cursor.fetchone()['COUNT(*)'] > 0:
                errors.append("Course name already exists.")
            
            if errors:
                for e in errors:
                    flash(e, "error")
                # Fetch courses, semesters, and subjects for re-render
                cursor.execute(
                    """
                    SELECT c.course_id, c.course_name, c.description, 
                           COUNT(s.semester_id) AS semester_count
                    FROM courses c
                    LEFT JOIN semesters s ON c.course_id = s.course_id
                    GROUP BY c.course_id, c.course_name, c.description
                    """
                )
                courses = cursor.fetchall()
                
                cursor.execute(
                    """
                    SELECT s.semester_id, c.course_name, s.semester_name, s.start_date, s.end_date, s.course_id
                    FROM semesters s
                    JOIN courses c ON s.course_id = c.course_id
                    """
                )
                semesters = cursor.fetchall()
                
                cursor.execute(
                    """
                    SELECT sub.subject_id, c.course_name, s.semester_name, sub.subject_name, sub.course_id, sub.semester_id
                    FROM subjects sub
                    JOIN courses c ON sub.course_id = c.course_id
                    JOIN semesters s ON sub.semester_id = s.semester_id
                    """
                )
                subjects = cursor.fetchall()
                
                cursor.close()
                conn.close()
                return render_template('admin_courses_semesters.html', courses=courses, semesters=semesters, 
                                     subjects=subjects, name=session.get('email'), 
                                     course_to_edit=course_to_edit, semester_to_edit=None, subject_to_edit=None)
            
            # Update course
            cursor.execute(
                "UPDATE courses SET course_name = %s, description = %s WHERE course_id = %s",
                (course_name, description or None, course_id)
            )
            conn.commit()
            flash("Course updated successfully!", "success")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Fetch courses, semesters, and subjects for display
        cursor.execute(
            """
            SELECT c.course_id, c.course_name, c.description, 
                   COUNT(s.semester_id) AS semester_count
            FROM courses c
            LEFT JOIN semesters s ON c.course_id = s.course_id
            GROUP BY c.course_id, c.course_name, c.description
            """
        )
        courses = cursor.fetchall()
        
        cursor.execute(
            """
            SELECT s.semester_id, c.course_name, s.semester_name, s.start_date, s.end_date, s.course_id
            FROM semesters s
            JOIN courses c ON s.course_id = c.course_id
            """
        )
        semesters = cursor.fetchall()
        
        cursor.execute(
            """
            SELECT sub.subject_id, c.course_name, s.semester_name, sub.subject_name, sub.course_id, sub.semester_id
            FROM subjects sub
            JOIN courses c ON sub.course_id = c.course_id
            JOIN semesters s ON sub.semester_id = s.semester_id
            """
        )
        subjects = cursor.fetchall()
        
        cursor.close()
        conn.close()
        return render_template('admin_courses_semesters.html', courses=courses, semesters=semesters, 
                             subjects=subjects, name=session.get('email'), 
                             course_to_edit=course_to_edit, semester_to_edit=None, subject_to_edit=None)
    
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/admin/courses_semesters')

@app.route('/admin/edit_semester/<int:semester_id>', methods=['GET', 'POST'])
def admin_edit_semester(semester_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Fetch the semester to edit
        cursor.execute(
            """
            SELECT s.semester_id, s.course_id, s.semester_name, s.start_date, s.end_date, c.course_name
            FROM semesters s
            JOIN courses c ON s.course_id = c.course_id
            WHERE s.semester_id = %s
            """,
            (semester_id,)
        )
        semester_to_edit = cursor.fetchone()
        
        if not semester_to_edit:
            flash("Semester not found.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        if request.method == 'POST':
            course_id = request.form.get('course_id', '').strip()
            semester_name = request.form.get('semester_name', '').strip()
            start_date = request.form.get('start_date', '').strip()
            end_date = request.form.get('end_date', '').strip()
            
            # Validation
            errors = []
            if not course_id or not course_id.isdigit():
                errors.append("Invalid course selected.")
            if not semester_name or len(semester_name) < 2 or len(semester_name) > 50:
                errors.append("Semester name must be 2–50 characters.")
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                if start_date >= end_date:
                    errors.append("End date must be after start date.")
            except ValueError:
                errors.append("Invalid date format for start or end date. Use YYYY-MM-DD.")
            
            # Check if course_id exists
            cursor.execute("SELECT COUNT(*) FROM courses WHERE course_id = %s", (course_id,))
            if cursor.fetchone()['COUNT(*)'] == 0:
                errors.append("Selected course does not exist.")
            
            # Check if semester_name is unique for the course (excluding the current semester)
            cursor.execute(
                "SELECT COUNT(*) FROM semesters WHERE course_id = %s AND semester_name = %s AND semester_id != %s",
                (course_id, semester_name, semester_id)
            )
            if cursor.fetchone()['COUNT(*)'] > 0:
                errors.append("Semester name already exists for this course.")
            
            if errors:
                for e in errors:
                    flash(e, "error")
                # Fetch courses, semesters, and subjects for re-render
                cursor.execute(
                    """
                    SELECT c.course_id, c.course_name, c.description, 
                           COUNT(s.semester_id) AS semester_count
                    FROM courses c
                    LEFT JOIN semesters s ON c.course_id = s.course_id
                    GROUP BY c.course_id, c.course_name, c.description
                    """
                )
                courses = cursor.fetchall()
                
                cursor.execute(
                    """
                    SELECT s.semester_id, c.course_name, s.semester_name, s.start_date, s.end_date, s.course_id
                    FROM semesters s
                    JOIN courses c ON s.course_id = c.course_id
                    """
                )
                semesters = cursor.fetchall()
                
                cursor.execute(
                    """
                    SELECT sub.subject_id, c.course_name, s.semester_name, sub.subject_name, sub.course_id, sub.semester_id
                    FROM subjects sub
                    JOIN courses c ON sub.course_id = c.course_id
                    JOIN semesters s ON sub.semester_id = s.semester_id
                    """
                )
                subjects = cursor.fetchall()
                
                cursor.close()
                conn.close()
                return render_template('admin_courses_semesters.html', courses=courses, semesters=semesters, 
                                     subjects=subjects, name=session.get('email'), 
                                     course_to_edit=None, semester_to_edit=semester_to_edit, subject_to_edit=None)
            
            # Update semester
            cursor.execute(
                """
                UPDATE semesters 
                SET course_id = %s, semester_name = %s, start_date = %s, end_date = %s 
                WHERE semester_id = %s
                """,
                (course_id, semester_name, start_date, end_date, semester_id)
            )
            conn.commit()
            flash("Semester updated successfully!", "success")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Fetch courses, semesters, and subjects for display
        cursor.execute(
            """
            SELECT c.course_id, c.course_name, c.description, 
                   COUNT(s.semester_id) AS semester_count
            FROM courses c
            LEFT JOIN semesters s ON c.course_id = s.course_id
            GROUP BY c.course_id, c.course_name, c.description
            """
        )
        courses = cursor.fetchall()
        
        cursor.execute(
            """
            SELECT s.semester_id, c.course_name, s.semester_name, s.start_date, s.end_date, s.course_id
            FROM semesters s
            JOIN courses c ON s.course_id = c.course_id
            """
        )
        semesters = cursor.fetchall()
        
        cursor.execute(
            """
            SELECT sub.subject_id, c.course_name, s.semester_name, sub.subject_name, sub.course_id, sub.semester_id
            FROM subjects sub
            JOIN courses c ON sub.course_id = c.course_id
            JOIN semesters s ON sub.semester_id = s.semester_id
            """
        )
        subjects = cursor.fetchall()
        
        cursor.close()
        conn.close()
        return render_template('admin_courses_semesters.html', courses=courses, semesters=semesters, 
                             subjects=subjects, name=session.get('email'), 
                             course_to_edit=None, semester_to_edit=semester_to_edit, subject_to_edit=None)
    
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/admin/courses_semesters')

@app.route('/admin/add_subject', methods=['POST'])
def admin_add_subject():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    course_id = request.form.get('course_id', '').strip()
    semester_id = request.form.get('semester_id', '').strip()
    subject_name = request.form.get('subject_name', '').strip()
    
    # Validation
    errors = []
    if not course_id or not course_id.isdigit():
        errors.append("Invalid course selected.")
    if not semester_id or not semester_id.isdigit():
        errors.append("Invalid semester selected.")
    if not subject_name or len(subject_name) < 2 or len(subject_name) > 100:
        errors.append("Subject name must be 2–100 characters.")
    
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect('/admin/courses_semesters')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        
        # Check if course_id exists
        cursor.execute("SELECT COUNT(*) FROM courses WHERE course_id = %s", (course_id,))
        if cursor.fetchone()[0] == 0:
            flash("Selected course does not exist.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Check if semester_id exists and belongs to the selected course
        cursor.execute(
            "SELECT COUNT(*) FROM semesters WHERE semester_id = %s AND course_id = %s",
            (semester_id, course_id)
        )
        if cursor.fetchone()[0] == 0:
            flash("Selected semester does not exist or does not belong to the selected course.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Check if subject_name already exists for this course and semester
        cursor.execute(
            "SELECT COUNT(*) FROM subjects WHERE course_id = %s AND semester_id = %s AND subject_name = %s",
            (course_id, semester_id, subject_name)
        )
        if cursor.fetchone()[0] > 0:
            flash("Subject name already exists for this course and semester.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Insert new subject
        cursor.execute(
            "INSERT INTO subjects (course_id, semester_id, subject_name) VALUES (%s, %s, %s)",
            (course_id, semester_id, subject_name)
        )
        conn.commit()
        flash("Subject added successfully!", "success")
        
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
    
    return redirect('/admin/courses_semesters')

@app.route('/admin/edit_subject/<int:subject_id>', methods=['GET', 'POST'])
def admin_edit_subject(subject_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Fetch the subject to edit
        cursor.execute(
            """
            SELECT sub.subject_id, sub.course_id, sub.semester_id, sub.subject_name, 
                   c.course_name, s.semester_name
            FROM subjects sub
            JOIN courses c ON sub.course_id = c.course_id
            JOIN semesters s ON sub.semester_id = s.semester_id
            WHERE sub.subject_id = %s
            """,
            (subject_id,)
        )
        subject_to_edit = cursor.fetchone()
        
        if not subject_to_edit:
            flash("Subject not found.", "error")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        if request.method == 'POST':
            course_id = request.form.get('course_id', '').strip()
            semester_id = request.form.get('semester_id', '').strip()
            subject_name = request.form.get('subject_name', '').strip()
            
            # Validation
            errors = []
            if not course_id or not course_id.isdigit():
                errors.append("Invalid course selected.")
            if not semester_id or not semester_id.isdigit():
                errors.append("Invalid semester selected.")
            if not subject_name or len(subject_name) < 2 or len(subject_name) > 100:
                errors.append("Subject name must be 2–100 characters.")
            
            # Check if course_id exists
            cursor.execute("SELECT COUNT(*) FROM courses WHERE course_id = %s", (course_id,))
            if cursor.fetchone()['COUNT(*)'] == 0:
                errors.append("Selected course does not exist.")
            
            # Check if semester_id exists and belongs to the selected course
            cursor.execute(
                "SELECT COUNT(*) FROM semesters WHERE semester_id = %s AND course_id = %s",
                (semester_id, course_id)
            )
            if cursor.fetchone()['COUNT(*)'] == 0:
                errors.append("Selected semester does not exist or does not belong to the selected course.")
            
            # Check if subject_name is unique for the course and semester (excluding the current subject)
            cursor.execute(
                """
                SELECT COUNT(*) FROM subjects 
                WHERE course_id = %s AND semester_id = %s AND subject_name = %s AND subject_id != %s
                """,
                (course_id, semester_id, subject_name, subject_id)
            )
            if cursor.fetchone()['COUNT(*)'] > 0:
                errors.append("Subject name already exists for this course and semester.")
            
            if errors:
                for e in errors:
                    flash(e, "error")
                # Fetch courses, semesters, and subjects for re-render
                cursor.execute(
                    """
                    SELECT c.course_id, c.course_name, c.description, 
                           COUNT(s.semester_id) AS semester_count
                    FROM courses c
                    LEFT JOIN semesters s ON c.course_id = s.course_id
                    GROUP BY c.course_id, c.course_name, c.description
                    """
                )
                courses = cursor.fetchall()
                
                cursor.execute(
                    """
                    SELECT s.semester_id, c.course_name, s.semester_name, s.start_date, s.end_date, s.course_id
                    FROM semesters s
                    JOIN courses c ON s.course_id = c.course_id
                    """
                )
                semesters = cursor.fetchall()
                
                cursor.execute(
                    """
                    SELECT sub.subject_id, c.course_name, s.semester_name, sub.subject_name, sub.course_id, sub.semester_id
                    FROM subjects sub
                    JOIN courses c ON sub.course_id = c.course_id
                    JOIN semesters s ON sub.semester_id = s.semester_id
                    """
                )
                subjects = cursor.fetchall()
                
                cursor.close()
                conn.close()
                return render_template('admin_courses_semesters.html', courses=courses, semesters=semesters, 
                                     subjects=subjects, name=session.get('email'), 
                                     course_to_edit=None, semester_to_edit=None, subject_to_edit=subject_to_edit)
            
            # Update subject
            cursor.execute(
                """
                UPDATE subjects 
                SET course_id = %s, semester_id = %s, subject_name = %s 
                WHERE subject_id = %s
                """,
                (course_id, semester_id, subject_name, subject_id)
            )
            conn.commit()
            flash("Subject updated successfully!", "success")
            cursor.close()
            conn.close()
            return redirect('/admin/courses_semesters')
        
        # Fetch courses, semesters, and subjects for display
        cursor.execute(
            """
            SELECT c.course_id, c.course_name, c.description, 
                   COUNT(s.semester_id) AS semester_count
            FROM courses c
            LEFT JOIN semesters s ON c.course_id = s.course_id
            GROUP BY c.course_id, c.course_name, c.description
            """
        )
        courses = cursor.fetchall()
        
        cursor.execute(
            """
            SELECT s.semester_id, c.course_name, s.semester_name, s.start_date, s.end_date, s.course_id
            FROM semesters s
            JOIN courses c ON s.course_id = c.course_id
            """
        )
        semesters = cursor.fetchall()
        
        cursor.execute(
            """
            SELECT sub.subject_id, c.course_name, s.semester_name, sub.subject_name, sub.course_id, sub.semester_id
            FROM subjects sub
            JOIN courses c ON sub.course_id = c.course_id
            JOIN semesters s ON sub.semester_id = s.semester_id
            """
        )
        subjects = cursor.fetchall()
        
        cursor.close()
        conn.close()
        return render_template('admin_courses_semesters.html', courses=courses, semesters=semesters, 
                             subjects=subjects, name=session.get('email'), 
                             course_to_edit=None, semester_to_edit=None, subject_to_edit=subject_to_edit)
    
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/admin/courses_semesters')


@app.route('/admin/api/semesters/<int:course_id>')
def admin_api_get_semesters(course_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401

    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT semester_id, semester_name
        FROM semesters
        WHERE course_id = %s
        ORDER BY semester_name
    """, (course_id,))

    semesters = cursor.fetchall()
    cursor.close()
    conn.close()

    return jsonify(semesters)

    
    


from datetime import datetime
from flask import request, render_template, flash, redirect, session, url_for, send_file
import mysql.connector
import pandas as pd
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

@app.route('/admin/evaluations')
def admin_evaluations():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')

    # Filters
    course_id = request.args.get('course_id')
    semester_id = request.args.get('semester_id')
    exam_id = request.args.get('exam_id')
    teacher_id = request.args.get('teacher_id')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    search = request.args.get('search', '').strip()

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Fetch filter options
        cursor.execute("SELECT course_id, course_name FROM courses ORDER BY course_name")
        courses = cursor.fetchall()

        cursor.execute("SELECT semester_id, semester_name FROM semesters ORDER BY semester_name")
        semesters = cursor.fetchall()

        cursor.execute("SELECT teacher_id, full_name FROM teachers WHERE full_name != 'Not Provided' ORDER BY full_name")
        teachers = cursor.fetchall()

        # Build exam query with filters (only finalized exams)
        exam_query = """
            SELECT e.exam_id, e.exam_name, e.exam_date, e.max_marks, e.min_marks, e.topic,
                   c.course_name, s.semester_name, sub.subject_name, t.full_name AS teacher_name
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            LEFT JOIN teachers t ON e.teacher_id = t.teacher_id
            WHERE e.evaluation_status = 1
        """
        exam_params = []

        if course_id:
            exam_query += " AND e.course_id = %s"
            exam_params.append(course_id)
        if semester_id:
            exam_query += " AND e.semester_id = %s"
            exam_params.append(semester_id)
        if teacher_id:
            exam_query += " AND e.teacher_id = %s"
            exam_params.append(teacher_id)
        if date_from:
            exam_query += " AND e.exam_date >= %s"
            exam_params.append(date_from)
        if date_to:
            exam_query += " AND e.exam_date <= %s"
            exam_params.append(date_to)

        exam_query += " ORDER BY e.exam_date DESC"
        cursor.execute(exam_query, exam_params)
        exams = cursor.fetchall()

        results = []
        exam_info = None
        
        if exam_id:
            # Get exam info
            cursor.execute("""
                SELECT e.exam_name, e.topic, e.max_marks, e.min_marks,
                       c.course_name, s.semester_name, sub.subject_name, t.full_name AS teacher_name
                FROM exams e
                JOIN courses c ON e.course_id = c.course_id
                JOIN semesters s ON e.semester_id = s.semester_id
                JOIN subjects sub ON e.subject_id = sub.subject_id
                LEFT JOIN teachers t ON e.teacher_id = t.teacher_id
                WHERE e.exam_id = %s
            """, (exam_id,))
            exam_info = cursor.fetchone()
            
            # Fetch results from result table
            result_query = """
                SELECT 
                    s.student_id, s.roll_no, s.enrollment_no, s.full_name,
                    r.score AS total_score,
                    r.final_score,
                    ex.max_marks, ex.min_marks,
                    CASE WHEN r.final_score >= ex.min_marks THEN 'Pass' ELSE 'Fail' END AS status
                FROM result r
                JOIN students s ON r.student_id = s.student_id
                JOIN exams ex ON r.exam_id = ex.exam_id
                WHERE r.exam_id = %s
            """
            params = [exam_id]
            
            if search:
                result_query += " AND (s.roll_no LIKE %s OR s.enrollment_no LIKE %s OR s.full_name LIKE %s)"
                pattern = f"%{search}%"
                params.extend([pattern, pattern, pattern])
            
            result_query += " ORDER BY r.final_score DESC, r.score DESC"
            cursor.execute(result_query, params)
            results = cursor.fetchall()
            
            # Convert to proper types
            for r in results:
                r['total_score'] = round(float(r['total_score'] or 0), 2)
                r['final_score'] = int(r['final_score'] or 0)
                r['max_marks'] = float(r['max_marks'] or 0)

        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        courses = semesters = teachers = exams = []
        results = []
        exam_info = None

    return render_template(
        'admin_evaluations.html',
        courses=courses, semesters=semesters, teachers=teachers,
        exams=exams, results=results, exam_info=exam_info,
        course_id=course_id, semester_id=semester_id,
        exam_id=exam_id, teacher_id=teacher_id,
        date_from=date_from, date_to=date_to,
        search_query=search, name=session.get('email')
    )
    
    

@app.route('/admin/export_results_excel')
def admin_export_results_excel():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Unauthorized.", "error")
        return redirect('/login')

    exam_id = request.args.get('exam_id')
    search = request.args.get('search', '').strip()
    if not exam_id:
        flash("Please select an exam first.", "error")
        return redirect('/admin/evaluations')

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Fetch exam details
        cursor.execute("""
            SELECT e.exam_name, e.topic, e.exam_date, e.max_marks, e.min_marks,
                   c.course_name, s.semester_name, sub.subject_name,
                   t.full_name AS teacher_name
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            LEFT JOIN teachers t ON e.teacher_id = t.teacher_id
            WHERE e.exam_id = %s
        """, (exam_id,))
        exam = cursor.fetchone()
        if not exam:
            flash("Exam not found.", "error")
            return redirect('/admin/evaluations')

        # Fetch student results from result table
        query = """
            SELECT s.roll_no, s.enrollment_no, s.full_name,
                   r.score, r.final_score,
                   ex.max_marks, ex.min_marks,
                   CASE WHEN r.final_score >= ex.min_marks THEN 'Pass' ELSE 'Fail' END AS status
            FROM result r
            JOIN students s ON r.student_id = s.student_id
            JOIN exams ex ON r.exam_id = ex.exam_id
            WHERE r.exam_id = %s
        """
        params = [exam_id]
        
        if search:
            query += " AND (s.roll_no LIKE %s OR s.enrollment_no LIKE %s OR s.full_name LIKE %s)"
            pattern = f"%{search}%"
            params.extend([pattern, pattern, pattern])
        
        query += " ORDER BY r.final_score DESC, r.score DESC"
        cursor.execute(query, params)
        results = cursor.fetchall()
        cursor.close()
        conn.close()

        # Create Excel with openpyxl
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Results")

        # Formats
        title_format = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'bg_color': '#4834d4', 'font_color': 'white'})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#4834d4', 'font_color': 'white', 'border': 1, 'align': 'center'})
        label_format = workbook.add_format({'bold': True, 'bg_color': '#E6E6FA', 'border': 1})
        cell_format = workbook.add_format({'border': 1})
        center_format = workbook.add_format({'border': 1, 'align': 'center'})
        pass_format = workbook.add_format({'border': 1, 'bg_color': '#D4EDDA', 'font_color': '#155724', 'align': 'center'})
        fail_format = workbook.add_format({'border': 1, 'bg_color': '#F8D7DA', 'font_color': '#721C24', 'align': 'center'})

        # Title
        worksheet.merge_range('A1:H1', 'Exam Results Report', title_format)
        
        # Exam Details
        row = 3
        exam_data = [
            ('Exam Name', exam['exam_name']),
            ('Topic', exam['topic'] or '—'),
            ('Course', exam['course_name']),
            ('Semester', exam['semester_name']),
            ('Subject', exam['subject_name']),
            ('Teacher', exam['teacher_name'] or '—'),
            ('Date', exam['exam_date'].strftime('%B %d, %Y') if exam['exam_date'] else '—'),
            ('Max Marks', exam['max_marks']),
            ('Min Marks', exam['min_marks']),
        ]

        for label, value in exam_data:
            worksheet.write(f'A{row}', label, label_format)
            worksheet.write(f'B{row}', value, cell_format)
            row += 1

        row += 2  # Gap

        # Table Headers
        headers = ['Roll No', 'Enrollment', 'Student Name', 'Score', 'Final Score', 'Max Marks', '%', 'Status']
        for col, header in enumerate(headers):
            worksheet.write(row - 1, col, header, header_format)

        # Data
        if not results:
            worksheet.write(row, 2, 'No students found', cell_format)
        else:
            for r in results:
                score = round(float(r['score'] or 0), 2)
                final_score = int(r['final_score'] or 0)
                max_marks = float(r['max_marks'] or 0)
                percentage = (score / max_marks * 100) if max_marks > 0 else 0
                
                worksheet.write(row, 0, r['roll_no'], center_format)
                worksheet.write(row, 1, r['enrollment_no'], center_format)
                worksheet.write(row, 2, r['full_name'], cell_format)
                worksheet.write(row, 3, score, center_format)
                worksheet.write(row, 4, final_score, center_format)
                worksheet.write(row, 5, max_marks, center_format)
                worksheet.write(row, 6, f"{percentage:.1f}%", center_format)
                
                status_format = pass_format if r['status'] == 'Pass' else fail_format
                worksheet.write(row, 7, r['status'], status_format)
                row += 1

        # Column widths
        worksheet.set_column('A:A', 10)
        worksheet.set_column('B:B', 15)
        worksheet.set_column('C:C', 25)
        worksheet.set_column('D:D', 12)
        worksheet.set_column('E:E', 12)
        worksheet.set_column('F:F', 12)
        worksheet.set_column('G:G', 10)
        worksheet.set_column('H:H', 10)

        workbook.close()
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Exam_{exam_id}_Results_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )

    except Exception as e:
        flash(f"Excel export failed: {e}", "error")
        return redirect('/admin/evaluations')    


@app.route('/admin/export_results_pdf')
def admin_export_results_pdf():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Unauthorized.", "error")
        return redirect('/login')

    exam_id = request.args.get('exam_id')
    search = request.args.get('search', '').strip()
    if not exam_id:
        flash("Please select an exam first.", "error")
        return redirect('/admin/evaluations')

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Exam details
        cursor.execute("""
            SELECT e.exam_name, e.topic, e.exam_date, e.max_marks, e.min_marks,
                   c.course_name, s.semester_name, sub.subject_name,
                   t.full_name AS teacher_name
            FROM exams e
            JOIN courses c ON e.course_id = c.course_id
            JOIN semesters s ON e.semester_id = s.semester_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            LEFT JOIN teachers t ON e.teacher_id = t.teacher_id
            WHERE e.exam_id = %s
        """, (exam_id,))
        exam = cursor.fetchone()
        if not exam:
            flash("Exam not found.", "error")
            return redirect('/admin/evaluations')

        # Student results from result table
        query = """
            SELECT s.roll_no, s.enrollment_no, s.full_name,
                   r.score, r.final_score,
                   ex.max_marks, ex.min_marks,
                   CASE WHEN r.final_score >= ex.min_marks THEN 'Pass' ELSE 'Fail' END AS status
            FROM result r
            JOIN students s ON r.student_id = s.student_id
            JOIN exams ex ON r.exam_id = ex.exam_id
            WHERE r.exam_id = %s
        """
        params = [exam_id]
        
        if search:
            query += " AND (s.roll_no LIKE %s OR s.enrollment_no LIKE %s OR s.full_name LIKE %s)"
            pattern = f"%{search}%"
            params.extend([pattern, pattern, pattern])
        
        query += " ORDER BY r.final_score DESC, r.score DESC"
        cursor.execute(query, params)
        results = cursor.fetchall()
        cursor.close()
        conn.close()

        # PDF Generation
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=50, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph("Exam Results Report", styles['Title']))
        elements.append(Spacer(1, 12))

        # Exam Details Table
        exam_data = [
            ['Exam Name', exam['exam_name']],
            ['Topic', exam['topic'] or '—'],
            ['Course', exam['course_name']],
            ['Semester', exam['semester_name']],
            ['Subject', exam['subject_name']],
            ['Teacher', exam['teacher_name'] or '—'],
            ['Date', exam['exam_date'].strftime('%B %d, %Y') if exam['exam_date'] else '—'],
            ['Max Marks', str(exam['max_marks'])],
            ['Min Marks', str(exam['min_marks'])]
        ]
        exam_table = Table(exam_data, colWidths=[120, 320])
        exam_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#4834d4')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ]))
        elements.append(exam_table)
        elements.append(Spacer(1, 20))

        # Generated on
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Student Results
        if results:
            data = [['Roll', 'Enrollment', 'Name', 'Score', 'Final', 'Max', '%', 'Status']]
            for r in results:
                score = round(float(r['score'] or 0), 2)
                final_score = int(r['final_score'] or 0)
                max_marks = float(r['max_marks'] or 0)
                percentage = (score / max_marks * 100) if max_marks > 0 else 0
                
                data.append([
                    str(r['roll_no']),
                    str(r['enrollment_no']),
                    r['full_name'][:18],  # Truncate long names
                    f"{score:.2f}",
                    str(final_score),
                    str(int(max_marks)),
                    f"{percentage:.1f}%",
                    r['status']
                ])
        else:
            data = [['—', '—', 'No students found', '—', '—', '—', '—', '—']]

        table = Table(data, colWidths=[45, 70, 100, 50, 45, 40, 45, 50])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4834d4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"Exam_{exam_id}_Results_{datetime.now().strftime('%Y%m%d')}.pdf"
        )

    except Exception as e:
        flash(f"PDF export failed: {e}", "error")
        return redirect('/admin/evaluations')
    


from flask import render_template, redirect, session, flash, request, jsonify, send_file
from datetime import datetime
import mysql.connector
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

@app.route('/admin/analytics')
def admin_analytics():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as an admin.", "error")
        return redirect('/login')
    
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Overall System Statistics
        cursor.execute("""
            SELECT 
                (SELECT COUNT(*) FROM users WHERE role = 'teacher') as total_teachers,
                (SELECT COUNT(*) FROM users WHERE role = 'student') as total_students,
                (SELECT COUNT(*) FROM users) as total_users,
                (SELECT COUNT(*) FROM exams) as total_exams,
                (SELECT COUNT(*) FROM questions) as total_questions,
                (SELECT COUNT(*) FROM student_answers) as total_submissions,
                (SELECT COUNT(*) FROM evaluations) as total_evaluations,
                (SELECT COUNT(*) FROM courses) as total_courses,
                (SELECT COUNT(*) FROM semesters) as total_semesters,
                (SELECT COUNT(*) FROM subjects) as total_subjects
        """)
        overall_stats = cursor.fetchone()
        
        # User Growth Trends (last 12 months) - based on users table (created_at)
        cursor.execute("""
            SELECT 
                DATE_FORMAT(created_at, '%%Y-%%m') AS ym,
                DATE_FORMAT(created_at, '%%b %%Y') AS month_name,
                SUM(CASE WHEN role = 'teacher' THEN 1 ELSE 0 END) AS teachers,
                SUM(CASE WHEN role = 'student' THEN 1 ELSE 0 END) AS students
            FROM users
            WHERE created_at >= DATE_SUB(CURDATE(), INTERVAL 11 MONTH)
            GROUP BY ym, month_name
            ORDER BY ym ASC
        """)
        user_growth = cursor.fetchall()
        
        # Course-wise Statistics
        cursor.execute("""
            SELECT 
                c.course_name,
                COUNT(DISTINCT s.semester_id) as semester_count,
                COUNT(DISTINCT sub.subject_id) as subject_count,
                COUNT(DISTINCT e.exam_id) as exam_count,
                COUNT(DISTINCT st.student_id) as student_count
            FROM courses c
            LEFT JOIN semesters s ON c.course_id = s.course_id
            LEFT JOIN subjects sub ON c.course_id = sub.course_id
            LEFT JOIN exams e ON c.course_id = e.course_id
            LEFT JOIN students st ON c.course_id = st.course_id
            GROUP BY c.course_id, c.course_name
            ORDER BY student_count DESC
        """)
        course_stats = cursor.fetchall()
        
        # Exam Activity Statistics
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN CURDATE() < exam_date THEN 'Upcoming'
                    WHEN CURDATE() = exam_date AND CURTIME() BETWEEN start_time AND end_time THEN 'Ongoing'
                    ELSE 'Completed'
                END as status,
                COUNT(*) as count
            FROM exams
            GROUP BY status
        """)
        exam_status_data = cursor.fetchall()
        exam_status = {item['status']: item['count'] for item in exam_status_data}
        
        # Subject-wise Question Distribution
        cursor.execute("""
            SELECT 
                sub.subject_name,
                COUNT(DISTINCT q.question_id) as question_count,
                COUNT(DISTINCT e.exam_id) as exam_count
            FROM subjects sub
            LEFT JOIN questions q ON sub.subject_id = q.subject_id
            LEFT JOIN exams e ON sub.subject_id = e.subject_id
            GROUP BY sub.subject_id, sub.subject_name
            HAVING question_count > 0
            ORDER BY question_count DESC
            LIMIT 10
        """)
        subject_distribution = cursor.fetchall()
        
        # Teacher Activity Analysis
        cursor.execute("""
            SELECT 
                t.full_name,
                t.teacher_id,
                COUNT(DISTINCT e.exam_id) as exams_created,
                COUNT(DISTINCT q.question_id) as questions_created,
                COUNT(DISTINCT sa.student_id) as students_reached,
                t.experience_years
            FROM teachers t
            LEFT JOIN exams e ON t.teacher_id = e.teacher_id
            LEFT JOIN questions q ON e.exam_id = q.exam_id
            LEFT JOIN student_answers sa ON q.question_id = sa.question_id
            GROUP BY t.teacher_id, t.full_name, t.experience_years
            HAVING exams_created > 0
            ORDER BY exams_created DESC
            LIMIT 10
        """)
        teacher_activity = cursor.fetchall()
        
        # Student Performance Overview (using result table for accuracy)
        cursor.execute("""
            SELECT 
                AVG(perf.pass_rate) AS avg_pass_rate,
                AVG(perf.avg_score) AS system_avg_score,
                COUNT(DISTINCT perf.student_id) AS active_students
            FROM (
                SELECT 
                    r.student_id,
                    AVG((r.final_score / e.max_marks) * 100) AS avg_score,
                    (SUM(CASE WHEN r.final_score >= e.min_marks THEN 1 ELSE 0 END) / COUNT(*)) * 100 AS pass_rate
                FROM result r
                JOIN exams e ON r.exam_id = e.exam_id
                GROUP BY r.student_id
            ) AS perf
        """)
        student_performance = cursor.fetchone() or {
            'avg_pass_rate': 0,
            'system_avg_score': 0,
            'active_students': 0
        }
        
        # Monthly Exam Creation Trends
        cursor.execute("""
            SELECT 
                DATE_FORMAT(exam_date, '%%Y-%%m') as month,
                DATE_FORMAT(exam_date, '%%b %%Y') as month_name,
                COUNT(*) as exam_count
            FROM exams
            WHERE exam_date >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
            GROUP BY DATE_FORMAT(exam_date, '%%Y-%%m'), DATE_FORMAT(exam_date, '%%b %%Y')
            ORDER BY month ASC
        """)
        exam_trends = cursor.fetchall()
        
        # Evaluation Status
        cursor.execute("""
            SELECT 
                COUNT(DISTINCT e.exam_id) as total_exams,
                SUM(CASE WHEN e.evaluation_status = 1 THEN 1 ELSE 0 END) as evaluated_exams,
                SUM(CASE WHEN e.evaluation_status = 0 THEN 1 ELSE 0 END) as pending_exams
            FROM exams e
        """)
        evaluation_status = cursor.fetchone()
        
        # Top Performing Courses (using result table)
        cursor.execute("""
            SELECT 
                c.course_name,
                COUNT(DISTINCT st.student_id) as student_count,
                COUNT(DISTINCT e.exam_id) as exam_count,
                AVG((r.final_score / e.max_marks) * 100) as avg_performance
            FROM courses c
            JOIN students st ON c.course_id = st.course_id
            JOIN result r ON st.student_id = r.student_id
            JOIN exams e ON r.exam_id = e.exam_id AND c.course_id = e.course_id
            GROUP BY c.course_id, c.course_name
            HAVING exam_count > 0
            ORDER BY avg_performance DESC
            LIMIT 5
        """)
        top_courses = cursor.fetchall()
        
        # System Health Metrics
        system_health = {
            'evaluation_completion': (evaluation_status['evaluated_exams'] / evaluation_status['total_exams'] * 100) if evaluation_status['total_exams'] > 0 else 0,
            'avg_questions_per_exam': (overall_stats['total_questions'] / overall_stats['total_exams']) if overall_stats['total_exams'] > 0 else 0,
            'avg_submissions_per_student': (overall_stats['total_submissions'] / overall_stats['total_students']) if overall_stats['total_students'] > 0 else 0,
            'teacher_to_student_ratio': (overall_stats['total_students'] / overall_stats['total_teachers']) if overall_stats['total_teachers'] > 0 else 0
        }
        
        # Get Courses
        cursor.execute("SELECT course_id, course_name FROM courses ORDER BY course_name")
        courses = cursor.fetchall()
        
        # Get Semesters with course relationship
        cursor.execute("""
            SELECT semester_id, semester_name, course_id 
            FROM semesters 
            ORDER BY course_id, semester_name
        """)
        semesters = cursor.fetchall()
        
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN evaluation_status = 1 THEN 1 ELSE 0 END) AS finalized,
                SUM(CASE WHEN evaluation_status = 0 THEN 1 ELSE 0 END) AS pending
            FROM exams
        """)
        exam_evaluation_status = cursor.fetchone()
        
        
        cursor.execute("""
            SELECT 
                t.full_name,
                COUNT(e.exam_id) AS exam_count
            FROM teachers t
            LEFT JOIN exams e 
                ON t.teacher_id = e.teacher_id
            WHERE 
                t.full_name IS NOT NULL
                AND TRIM(t.full_name) <> ''
                AND t.full_name <> 'Not Provided'
            GROUP BY t.teacher_id, t.full_name
            ORDER BY exam_count DESC;

        """)
        teacher_workload = cursor.fetchall()
        
        
        cursor.execute("""
            SELECT 
                e.exam_name,
                COUNT(DISTINCT sa.student_id) AS participants
            FROM exams e
            LEFT JOIN questions q ON e.exam_id = q.exam_id
            LEFT JOIN student_answers sa ON q.question_id = sa.question_id
            GROUP BY e.exam_id, e.exam_name
        """)
        exam_participation = cursor.fetchall()


        
        cursor.execute("""
            SELECT 
                e.exam_name,
                ROUND(AVG(r.final_score), 2) AS avg_score
            FROM exams e
            JOIN result r ON e.exam_id = r.exam_id
            GROUP BY e.exam_id, e.exam_name
        """)
        avg_score_per_exam = cursor.fetchall()
        
        
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN r.final_score >= e.min_marks THEN 1 ELSE 0 END) AS pass_count,
                SUM(CASE WHEN r.final_score < e.min_marks THEN 1 ELSE 0 END) AS fail_count
            FROM result r
            JOIN exams e ON r.exam_id = e.exam_id
        """)
        pass_fail_ratio = cursor.fetchone()
        
        # New query for AI vs Teacher Score Gap Analysis
        cursor.execute("""
            SELECT 
                e.exam_name,
                e.exam_date,
                AVG(r.score) as avg_ai_score,
                AVG(r.final_score) as avg_teacher_score,
                AVG(r.final_score - r.score) as avg_gap
            FROM result r
            JOIN exams e ON r.exam_id = e.exam_id
            GROUP BY e.exam_id, e.exam_name, e.exam_date
            ORDER BY e.exam_date ASC
        """)
        gap_analysis = cursor.fetchall()
        
        
       
     
        analytics = {
            'overall_stats': overall_stats,
            'user_growth': user_growth,
            'course_stats': course_stats,
            'exam_status': exam_status,
            'subject_distribution': subject_distribution,
            'teacher_activity': teacher_activity,
            'student_performance': student_performance,
            'exam_trends': exam_trends,
            'evaluation_status': evaluation_status,
            'top_courses': top_courses,
            'system_health': system_health,
            'courses': courses,
            'semesters': semesters
        }
        
        analytics.update({
            'exam_evaluation_status': exam_evaluation_status,
            'teacher_workload': teacher_workload,
            'exam_participation': exam_participation,
            'avg_score_per_exam': avg_score_per_exam,
            'pass_fail_ratio': pass_fail_ratio,
            'gap_analysis': gap_analysis
        })

        
        cursor.close()
        conn.close()
        
        return render_template('admin_analytics.html', 
                             analytics=analytics, 
                             name=session.get('email'))
        
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")
        return redirect('/admin_dashboard')
    

@app.route('/admin/generate_report', methods=['POST'])
def generate_report():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.json or {}
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        course_id = data.get('course_id')  # can be 'all'
        semester_id = data.get('semester_id')  # can be None
        status_filter = data.get('status')
        
        # Validate dates
        if not from_date or not to_date:
            return jsonify({'error': 'From date and To date are required'}), 400
        
        try:
            from_datetime = datetime.strptime(from_date, '%Y-%m-%d')
            to_datetime = datetime.strptime(to_date, '%Y-%m-%d')
            
            if from_datetime > to_datetime:
                return jsonify({'error': 'From date must be before To date'}), 400
            
            if from_datetime > datetime.now() or to_datetime > datetime.now():
                return jsonify({'error': 'Dates cannot be in the future'}), 400
        except ValueError:
            return jsonify({'error': 'Invalid date format'}), 400
        
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        
        # Build query using result table: aggregate per student per subject
        base_query = """
            SELECT 
                st.roll_no,
                st.enrollment_no,
                st.full_name,
                st.contact,
                st.course_id,
                c.course_name,
                st.semester_id,
                sem.semester_name,
                sub.subject_id,
                sub.subject_name,
                SUM(r.final_score) AS obtained_marks,
                SUM(e.max_marks) AS max_marks,
                ROUND((SUM(r.final_score) / SUM(e.max_marks)) * 100, 2) AS percentage,
                CASE 
                    WHEN (SUM(r.final_score) / SUM(e.max_marks)) * 100 >= 75 THEN 'Distinction'
                    WHEN (SUM(r.final_score) / SUM(e.max_marks)) * 100 >= 60 THEN 'First Class'
                    WHEN (SUM(r.final_score) / SUM(e.max_marks)) * 100 >= 50 THEN 'Second Class'
                    WHEN (SUM(r.final_score) / SUM(e.max_marks)) * 100 >= 40 THEN 'Pass'
                    ELSE 'Fail'
                END AS subject_status
            FROM students st
            JOIN courses c ON st.course_id = c.course_id
            JOIN semesters sem ON st.semester_id = sem.semester_id
            JOIN result r ON st.student_id = r.student_id
            JOIN exams e ON r.exam_id = e.exam_id
            JOIN subjects sub ON e.subject_id = sub.subject_id
            WHERE e.exam_date >= %s
              AND e.exam_date <= %s
        """
        
        params = [from_date, to_date]
        
        # Optional course filter (ignore when 'all')
        if course_id and course_id != 'all':
            base_query += " AND st.course_id = %s"
            params.append(course_id)
        
        # Optional semester filter (used only when provided)
        if semester_id:
            base_query += " AND st.semester_id = %s"
            params.append(semester_id)
        
        base_query += """
            GROUP BY 
                st.student_id, st.roll_no, st.enrollment_no, st.full_name, st.contact,
                st.course_id, c.course_name, st.semester_id, sem.semester_name,
                sub.subject_id, sub.subject_name
            ORDER BY c.course_name, sem.semester_name, st.roll_no, sub.subject_name
        """
        
        cursor.execute(base_query, params)
        report_data = cursor.fetchall()
        
        # Filter by status if provided
        if status_filter:
            report_data = [row for row in report_data if row['subject_status'] == status_filter]
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'status': 'success',
            'data': report_data,
            'count': len(report_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/admin/export_report', methods=['POST'])
def export_report():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.json or {}
        report_format = data.get('format')
        report_data = data.get('data') or []
        filters = data.get('filters', {})
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 400
        
        if report_format == 'excel':
            return generate_excel_report(report_data, filters)
        elif report_format == 'pdf':
            return generate_pdf_report(report_data, filters)
        else:
            return jsonify({'error': 'Invalid format'}), 400
            
    except Exception as e:
        print(f"Export error: {str(e)}")
        return jsonify({'error': str(e)}), 500


def _prepare_grouped_report(report_data):
    """
    Transform raw report_data (list of dicts) into grouped structure:
    groups[(course_name, semester_name)] = {
        'course_name': ...,
        'semester_name': ...,
        'subjects': [ { 'id', 'name', 'max_marks' }, ... ],
        'students': {
            student_key: {
                'roll_no', 'enrollment_no', 'full_name', 'contact',
                'marks': { subject_id: obtained_marks }
            }
        }
    }
    """
    groups = {}

    for record in report_data:
        course_name = record.get('course_name') or 'N/A'
        semester_name = record.get('semester_name') or 'N/A'
        key = (course_name, semester_name)

        if key not in groups:
            groups[key] = {
                'course_name': course_name,
                'semester_name': semester_name,
                'subjects': {},   # subject_id -> {name, max_marks}
                'students': {}    # student_key -> student_data
            }

        group = groups[key]

        subject_id = record.get('subject_id')
        subject_name = record.get('subject_name') or 'Subject'
        max_marks = float(record.get('max_marks') or 0.0)

        if subject_id not in group['subjects']:
            group['subjects'][subject_id] = {
                'id': subject_id,
                'name': subject_name,
                'max_marks': max_marks
            }
        else:
            # Ensure max_marks is the max across rows
            group['subjects'][subject_id]['max_marks'] = max(
                group['subjects'][subject_id]['max_marks'],
                max_marks
            )

        roll_no = record.get('roll_no')
        enrollment_no = record.get('enrollment_no')
        full_name = record.get('full_name')
        contact = record.get('contact')

        student_key = f"{roll_no}-{enrollment_no}"
        if student_key not in group['students']:
            group['students'][student_key] = {
                'roll_no': roll_no,
                'enrollment_no': enrollment_no,
                'full_name': full_name,
                'contact': contact,
                'marks': {}  # subject_id -> obtained_marks
            }

        obtained_marks = float(record.get('obtained_marks') or 0.0)
        group['students'][student_key]['marks'][subject_id] = obtained_marks

    # Convert subjects dicts to ordered lists
    for key, group in groups.items():
        # Sort subjects by name
        subjects_list = sorted(
            group['subjects'].values(),
            key=lambda s: s['name']
        )
        group['subjects'] = subjects_list

    return groups


def generate_excel_report(report_data, filters):
    try:
        groups = _prepare_grouped_report(report_data)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Performance"

        title_font = Font(bold=True, size=16, color='4834D4')
        subtitle_font = Font(bold=True, size=12, color='333333')
        info_font = Font(size=10, italic=True)
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4834D4', end_color='4834D4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1
        ws.cell(row=row, column=1, value='Department of Computer Science').font = title_font
        row += 1
        ws.cell(row=row, column=1, value='Gujarat University').font = title_font
        row += 1
        ws.cell(row=row, column=1, value='Student Performance Report').font = subtitle_font
        row += 2

        ws.cell(row=row, column=1, value=f"Report Generated: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}").font = info_font
        row += 1

        filter_parts = []
        if filters.get('from_date'):
            filter_parts.append(f"From: {filters['from_date']}")
        if filters.get('to_date'):
            filter_parts.append(f"To: {filters['to_date']}")
        if filters.get('course'):
            filter_parts.append(f"Course: {filters['course']}")
        if filters.get('semester'):
            filter_parts.append(f"Semester: {filters['semester']}")

        if filter_parts:
            ws.cell(row=row, column=1, value=" | ".join(filter_parts)).font = Font(size=9, italic=True, color='666666')
            row += 2
        else:
            row += 1

        # For each course-semester group, create a block
        for (course_name, semester_name), group in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
            subjects = group['subjects']
            students = list(group['students'].values())

            # Course heading
            ws.cell(row=row, column=1, value=f"Course: {course_name} | Semester: {semester_name}").font = Font(bold=True, size=11, color='4834D4')
            row += 1

            # Header row: Roll, Enrollment, Name, Contact, each subject, Total
            headers = ['Roll No', 'Enrollment No', 'Student Name', 'Contact']
            for subj in subjects:
                headers.append(f"{subj['name']} (Max {subj['max_marks']:.0f})")
            total_max = sum(s['max_marks'] for s in subjects)
            headers.append(f"Total (Max {total_max:.0f})")

            start_col = 1
            for col_offset, header in enumerate(headers):
                col = start_col + col_offset
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            row += 1

            # Data rows
            for student in students:
                roll_no = student['roll_no']
                enrollment_no = student['enrollment_no']
                full_name = student['full_name']
                contact = student['contact']
                marks = student['marks']

                row_values = [
                    roll_no,
                    enrollment_no,
                    full_name,
                    contact
                ]

                total_obtained = 0.0
                for subj in subjects:
                    m = float(marks.get(subj['id'], 0.0))
                    row_values.append(m)
                    total_obtained += m

                if total_max > 0:
                    perc = (total_obtained / total_max) * 100
                else:
                    perc = 0.0

                total_cell_value = f"{total_obtained:.2f}/{total_max:.0f} ({perc:.2f}%)"
                row_values.append(total_cell_value)

                for col_offset, value in enumerate(row_values):
                    col = start_col + col_offset
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = border
                    if col_offset < 4:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1

            row += 2  # space between groups

        # Adjust some basic column widths (first four)
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'student_report_{datetime.now().strftime("%d%m%Y_%H%M%S")}.xlsx'
        )
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        raise


def generate_pdf_report(report_data, filters):
    try:
        from reportlab.lib import colors

        groups = _prepare_grouped_report(report_data)

        output = io.BytesIO()
        pdf = SimpleDocTemplate(
            output, 
            pagesize=A4, 
            leftMargin=0.5*inch, 
            rightMargin=0.5*inch,
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#4834D4'),  # ✅ FIXED
            spaceAfter=4,
            alignment=1,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#333333'),  # ✅ FIXED
            spaceAfter=6,
            alignment=1
        )
        
        info_style = styles['Normal']
        info_style.fontSize = 8

        # Header
        story.append(Paragraph('Department of Computer Science', title_style))
        story.append(Paragraph('Gujarat University', title_style))
        story.append(Paragraph('Student Performance Report', subtitle_style))
        story.append(Spacer(1, 0.15*inch))
        
        info_text = f"<b>Report Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
        story.append(Paragraph(info_text, styles['Normal']))
        
        filter_parts = []
        if filters.get('from_date'):
            filter_parts.append(f"<b>From:</b> {filters['from_date']}")
        if filters.get('to_date'):
            filter_parts.append(f"<b>To:</b> {filters['to_date']}")
        if filters.get('course'):
            filter_parts.append(f"<b>Course:</b> {filters['course']}")
        if filters.get('semester'):
            filter_parts.append(f"<b>Semester:</b> {filters['semester']}")

        if filter_parts:
            story.append(Paragraph(" | ".join(filter_parts), styles['Normal']))
        
        story.append(Spacer(1, 0.15*inch))

        # For each course-semester group, create a table
        for (course_name, semester_name), group in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
            subjects = group['subjects']
            students = list(group['students'].values())

            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(
                f"<b>Course:</b> {course_name} &nbsp;&nbsp; <b>Semester:</b> {semester_name}", 
                styles['Normal']
            ))
            story.append(Spacer(1, 0.05*inch))

            # Header row
            header_cells = ['Roll No', 'Enroll No', 'Student Name', 'Contact']
            total_max = sum(s['max_marks'] for s in subjects)
            for subj in subjects:
                short_name = subj['name'][:15]
                header_cells.append(f"{short_name} (Max {subj['max_marks']:.0f})")
            header_cells.append(f"Total (Max {total_max:.0f})")

            table_data = [header_cells]

            # Rows: one line per student
            for student in students:
                row_cells = [
                    str(student['roll_no']),
                    str(student['enrollment_no']),
                    str(student['full_name']),
                    str(student['contact'])
                ]
                total_obtained = 0.0
                marks = student['marks']

                for subj in subjects:
                    obtained = float(marks.get(subj['id'], 0.0))
                    total_obtained += obtained
                    row_cells.append(f"{obtained:.2f}")

                if total_max > 0:
                    perc = (total_obtained / total_max) * 100
                else:
                    perc = 0.0
                row_cells.append(f"{total_obtained:.2f}/{total_max:.0f} ({perc:.1f}%)")

                table_data.append(row_cells)

            # Build table with proper header color
            col_count = len(header_cells)
            base_width = 7.5 * inch
            col_widths = [0.6*inch, 0.9*inch, 1.4*inch, 0.9*inch]
            remaining_cols = col_count - 4
            if remaining_cols > 0:
                remaining_width = base_width - sum(col_widths)
                per = max(0.6*inch, remaining_width / remaining_cols)
                for _ in range(remaining_cols):
                    col_widths.append(per)

            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4834D4')),  # ✅ FIXED
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('VALIGN', (0, 1), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))

            story.append(table)
            story.append(Spacer(1, 0.2*inch))

        pdf.build(story)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'student_report_{datetime.now().strftime("%d%m%Y_%H%M%S")}.pdf'
        )
    except Exception as e:
        print(f"PDF export error: {str(e)}")
        raise

@app.route('/admins/profile')
def admin_profile():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash("Please log in as admin.", "error")
        return redirect('/login')

    admin_id = session['user_id']

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # Fetch admin details
        cursor.execute("""
            SELECT a.full_name, a.contact, u.email
            FROM admins a
            JOIN users u ON a.admin_id = u.uid
            WHERE a.admin_id = %s
        """, (admin_id,))
        admin = cursor.fetchone()

        cursor.close()
        conn.close()

        if not admin:
            flash("Admin profile not found.", "error")
            return redirect('/admin_dashboard')

        full_name, contact, email = admin

        return render_template(
            'admin_profile.html',
            name=email,
            email=email,
            profile={
                "full_name": full_name,
                "contact": contact
            }
        )

    except Exception as e:
        flash(f"Error loading profile: {str(e)}", "error")
        return redirect('/admin_dashboard')    



@app.route('/teacher/upload', methods=['POST'])
def teacher_upload():
    if 'user_id' not in session or session.get('role') != 'teacher':
        flash("Please log in as a teacher.", "error")
        return redirect('/login')

    question_text = request.form.get('question_text', '').strip()
    model_answer = request.form.get('model_answer', '').strip()
    max_score = request.form.get('max_score', '').strip()
    course_id = request.form.get('course_id', '').strip()
    semester_id = request.form.get('semester_id', '').strip()

    if not question_text or not model_answer or not max_score or not course_id or not semester_id:
        flash("All fields are required for uploading a question!", "error")
        return redirect('/teacher_dashboard')

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO questions (course_id, semester_id, question_text, model_answer, max_score) VALUES (%s, %s, %s, %s, %s)",
            (course_id, semester_id, question_text, model_answer, max_score)
        )
        conn.commit()
        cursor.close()
        conn.close()
        flash("Question uploaded successfully!", "success")
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")

    return redirect('/teacher_dashboard')





@app.route('/student/submit', methods=['POST'])
def student_submit():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Please log in as a student.", "error")
        return redirect('/login')

    answer_text = request.form.get('answer_text', '').strip()
    question_id = request.form.get('question_id', '').strip()

    if not answer_text or not question_id:
        flash("Answer and question selection are required!", "error")
        return redirect('/student_dashboard')

    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO student_answers (student_id, question_id, answer_text) VALUES (%s, %s, %s)",
            (session['user_id'], question_id, answer_text)
        )
        conn.commit()
        cursor.close()
        conn.close()
        flash("Answer submitted successfully!", "success")
    except mysql.connector.Error as err:
        flash(f"Database error: {err}", "error")

    return redirect('/student_dashboard')

if __name__ == '__main__':
    app.run(debug=False, port=5000)